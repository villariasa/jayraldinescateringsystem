import os
from datetime import datetime

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, Image, KeepTogether
    )
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


if REPORTLAB_OK:
    _C_RED    = colors.HexColor("#E11D48")
    _C_DARK   = colors.HexColor("#0B1220")
    _C_GRAY   = colors.HexColor("#374151")
    _C_LIGHT  = colors.HexColor("#F9FAFB")
    _C_WHITE  = colors.white
    _C_MUTED  = colors.HexColor("#6B7280")
    _C_BORDER = colors.HexColor("#E5E7EB")
    _C_GREEN  = colors.HexColor("#22C55E")
    _C_AMBER  = colors.HexColor("#F59E0B")
    _C_BLUE   = colors.HexColor("#3B82F6")
else:
    _C_RED = _C_DARK = _C_GRAY = _C_LIGHT = _C_WHITE = _C_MUTED = _C_BORDER = _C_GREEN = _C_AMBER = _C_BLUE = None

from utils.paths import resource_path
_LOGO_PATH = resource_path("assets", "logo.png")
if not os.path.exists(_LOGO_PATH):
    _LOGO_PATH = resource_path("assets", "logo.jpg")

_PAGE_W = A4[0] if REPORTLAB_OK else 595
_MARGIN = 1.5 * cm if REPORTLAB_OK else 0
_CONTENT_W = _PAGE_W - 2 * _MARGIN


def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle("Brand",
        fontName="Helvetica-Bold", fontSize=20, textColor=_C_RED,
        spaceAfter=2, alignment=TA_LEFT, leading=24))
    s.add(ParagraphStyle("BrandSub",
        fontName="Helvetica", fontSize=8.5, textColor=_C_MUTED,
        spaceAfter=0, alignment=TA_LEFT, leading=12))
    s.add(ParagraphStyle("BrandSubRight",
        fontName="Helvetica", fontSize=8.5, textColor=_C_MUTED,
        spaceAfter=0, alignment=TA_RIGHT, leading=12))
    s.add(ParagraphStyle("SectionHead",
        fontName="Helvetica-Bold", fontSize=10, textColor=_C_DARK,
        spaceBefore=14, spaceAfter=6, alignment=TA_LEFT, leading=14))
    s.add(ParagraphStyle("KpiLabel",
        fontName="Helvetica", fontSize=7.5, textColor=_C_MUTED,
        spaceAfter=1, alignment=TA_CENTER, leading=10))
    s.add(ParagraphStyle("KpiValue",
        fontName="Helvetica-Bold", fontSize=15, textColor=_C_DARK,
        spaceAfter=0, alignment=TA_CENTER, leading=18))
    s.add(ParagraphStyle("Footer",
        fontName="Helvetica", fontSize=7, textColor=_C_MUTED,
        alignment=TA_CENTER, leading=10))
    s.add(ParagraphStyle("ReceiptTitle",
        fontName="Helvetica-Bold", fontSize=18, textColor=_C_WHITE,
        alignment=TA_LEFT, leading=22, spaceAfter=4))
    s.add(ParagraphStyle("ReceiptSub",
        fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#CBD5E1"),
        alignment=TA_LEFT, leading=12))
    s.add(ParagraphStyle("DetailLabel",
        fontName="Helvetica-Bold", fontSize=9, textColor=_C_GRAY,
        leading=12, wordWrap="CJK"))
    s.add(ParagraphStyle("DetailValue",
        fontName="Helvetica", fontSize=9, textColor=_C_DARK,
        leading=12, wordWrap="CJK"))
    s.add(ParagraphStyle("DetailValueBold",
        fontName="Helvetica-Bold", fontSize=9, textColor=_C_DARK,
        leading=12, wordWrap="CJK"))
    s.add(ParagraphStyle("TableHead",
        fontName="Helvetica-Bold", fontSize=8, textColor=_C_WHITE,
        alignment=TA_CENTER, leading=10))
    s.add(ParagraphStyle("TableCell",
        fontName="Helvetica", fontSize=8, textColor=_C_DARK,
        leading=11, wordWrap="CJK"))
    s.add(ParagraphStyle("TableCellRight",
        fontName="Helvetica", fontSize=8, textColor=_C_DARK,
        alignment=TA_RIGHT, leading=11, wordWrap="CJK"))
    s.add(ParagraphStyle("TableCellCenter",
        fontName="Helvetica", fontSize=8, textColor=_C_DARK,
        alignment=TA_CENTER, leading=11, wordWrap="CJK"))
    s.add(ParagraphStyle("StatusPaid",
        fontName="Helvetica-Bold", fontSize=8, textColor=_C_GREEN,
        alignment=TA_CENTER, leading=11))
    s.add(ParagraphStyle("StatusPartial",
        fontName="Helvetica-Bold", fontSize=8, textColor=_C_AMBER,
        alignment=TA_CENTER, leading=11))
    s.add(ParagraphStyle("StatusUnpaid",
        fontName="Helvetica-Bold", fontSize=8, textColor=_C_RED,
        alignment=TA_CENTER, leading=11))
    return s


def _status_style(status: str, styles):
    mapping = {
        "Paid":    styles["StatusPaid"],
        "Partial": styles["StatusPartial"],
        "Unpaid":  styles["StatusUnpaid"],
    }
    return mapping.get(status, styles["TableCell"])


def _header_block(story, styles, biz_name: str, title: str, period: str = "All Time"):
    header_left = []
    if os.path.exists(_LOGO_PATH):
        try:
            logo = Image(_LOGO_PATH, width=2*cm, height=2*cm)
            header_left.append(logo)
        except Exception:
            header_left.append(Paragraph(biz_name[:1], styles["Brand"]))
    else:
        header_left.append(Paragraph(biz_name[:1], styles["Brand"]))

    header_center = [
        Paragraph(biz_name, styles["Brand"]),
        Paragraph("Professional Catering Services", styles["BrandSub"]),
    ]

    now_str = datetime.now().strftime("%b %d, %Y  %I:%M %p")
    header_right = [
        Paragraph(f"<b>{title}</b>", styles["BrandSubRight"]),
        Paragraph(f"Period: {period}", styles["BrandSubRight"]),
        Paragraph(f"Generated: {now_str}", styles["BrandSubRight"]),
    ]

    tbl = Table([[header_left, header_center, header_right]],
                colWidths=[2.2*cm, 10.5*cm, 5.5*cm])
    tbl.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (2, 0), (2, 0),   "RIGHT"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(HRFlowable(width="100%", thickness=2.5, color=_C_RED, spaceAfter=12))


def _kpi_row(story, styles, kpis: dict):
    items = [
        ("Total Bookings",  str(kpis.get("total_bookings", 0))),
        ("Total Pax",       f"{int(kpis.get('total_pax', 0)):,}"),
        ("Total Revenue",   f"PHP {float(kpis.get('total_revenue', 0)):,.0f}"),
        ("Unpaid Amount",   f"PHP {float(kpis.get('unpaid_amount', 0)):,.0f}"),
    ]
    ncols = len(items)
    col_w = [_CONTENT_W / ncols] * ncols

    labels_row = [Paragraph(lbl, styles["KpiLabel"]) for lbl, _ in items]
    values_row = [Paragraph(val, styles["KpiValue"]) for _, val in items]

    t = Table([labels_row, values_row], colWidths=col_w)
    t.setStyle(TableStyle([
        ("BOX",           (0, 0), (-1, -1), 0.5, _C_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, _C_BORDER),
        ("BACKGROUND",    (0, 0), (-1, 0),  _C_LIGHT),
        ("BACKGROUND",    (0, 1), (-1, 1),  _C_WHITE),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))


def _bookings_table(story, styles, bookings: list):
    story.append(Paragraph("Booking Statistics", styles["SectionHead"]))

    headers = ["Booking Ref", "Client", "Event Date", "Pax", "Total Amount", "Status"]
    col_w   = [2.8*cm, 5.5*cm, 2.8*cm, 1.5*cm, 3.2*cm, 2.4*cm]

    header_row = [Paragraph(h, styles["TableHead"]) for h in headers]
    rows = [header_row]

    status_styles = {
        "CONFIRMED": styles["StatusPaid"],
        "PENDING":   styles["StatusPartial"],
        "CANCELLED": styles["StatusUnpaid"],
    }

    for b in bookings:
        st_key = b.get("status", "").upper()
        st_style = status_styles.get(st_key, styles["TableCell"])
        rows.append([
            Paragraph(b.get("id", ""), styles["TableCell"]),
            Paragraph(b.get("name", ""), styles["TableCell"]),
            Paragraph(b.get("date", ""), styles["TableCellCenter"]),
            Paragraph(str(b.get("pax", "")), styles["TableCellCenter"]),
            Paragraph(b.get("total", ""), styles["TableCellRight"]),
            Paragraph(b.get("status", "").capitalize(), st_style),
        ])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  _C_RED),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [_C_WHITE, _C_LIGHT]),
        ("BOX",           (0, 0), (-1, -1), 0.4, _C_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, _C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tbl)


def _footer(story, styles, biz_name: str = "Jayraldine's Catering"):
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_C_BORDER))
    story.append(Spacer(1, 5))
    story.append(Paragraph(
        f"{biz_name}  •  This report is system-generated and confidential.",
        styles["Footer"]
    ))


def export_pdf(path: str, kpis: dict, bookings: list,
               title: str = "Business Report", period: str = "All Time",
               biz_name: str = "Jayraldine's Catering") -> bool:
    if not REPORTLAB_OK:
        return False
    try:
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=_MARGIN, rightMargin=_MARGIN,
            topMargin=_MARGIN, bottomMargin=_MARGIN,
            title=f"{biz_name} — {title}",
            author=biz_name,
        )
        styles = _styles()
        story = []
        _header_block(story, styles, biz_name, title, period)
        story.append(Paragraph("Key Performance Indicators", styles["SectionHead"]))
        _kpi_row(story, styles, kpis)
        _bookings_table(story, styles, bookings)
        _footer(story, styles, biz_name)
        doc.build(story)
        return True
    except Exception as exc:
        print(f"[exporter] PDF failed: {exc}")
        return False


def export_receipt_pdf(path: str, inv: dict, business: dict) -> bool:
    """Generate a professional PDF receipt for a single invoice."""
    if not REPORTLAB_OK:
        return False
    try:
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=_MARGIN, rightMargin=_MARGIN,
            topMargin=_MARGIN, bottomMargin=_MARGIN,
            title=f"Receipt — {inv.get('invoice', '')}",
        )
        styles = _styles()
        story = []

        biz_name    = business.get("name", "Jayraldine's Catering")
        biz_address = business.get("address", "")
        biz_contact = business.get("contact", "")
        biz_email   = business.get("email", "")

        total   = float(inv.get("amount", 0))
        paid    = float(inv.get("paid", 0))
        balance = total - paid
        status  = inv.get("status", "Unpaid")

        status_color = {"Paid": _C_GREEN, "Partial": _C_AMBER, "Unpaid": _C_RED}.get(status, _C_MUTED)

        header_w = _CONTENT_W

        logo_cell = ""
        if os.path.exists(_LOGO_PATH):
            try:
                logo_cell = Image(_LOGO_PATH, width=2.0*cm, height=2.0*cm)
            except Exception:
                logo_cell = ""

        biz_cell = [
            Paragraph(biz_name, styles["Brand"]),
            Paragraph("Professional Catering Services", styles["BrandSub"]),
            Spacer(1, 2),
            Paragraph(biz_address, styles["BrandSub"]),
            Paragraph(f"Tel: {biz_contact}  ·  {biz_email}", styles["BrandSub"]),
        ]

        receipt_cell = [
            Paragraph("OFFICIAL RECEIPT", ParagraphStyle(
                "rcpt_lbl", fontName="Helvetica-Bold", fontSize=8.5,
                textColor=_C_MUTED, alignment=TA_RIGHT, leading=11,
                spaceAfter=4)),
            Paragraph(inv.get("invoice", "—"), ParagraphStyle(
                "rcpt_no", fontName="Helvetica-Bold", fontSize=18,
                textColor=_C_RED, alignment=TA_RIGHT, leading=22)),
        ]

        hdr_cols = [2.2*cm, 10*cm, 6*cm] if logo_cell else [12.2*cm, 6*cm]
        hdr_data = [[logo_cell, biz_cell, receipt_cell]] if logo_cell else [[biz_cell, receipt_cell]]

        hdr_tbl = Table(hdr_data, colWidths=hdr_cols)
        hdr_tbl.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("ALIGN",         (-1, 0), (-1, 0), "RIGHT"),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=2.5, color=_C_RED, spaceAfter=0.4*cm))

        def _det_row(label, value, alt=False, value_style=None):
            bg = colors.HexColor("#F8FAFC") if alt else _C_WHITE
            vs = value_style or styles["DetailValue"]
            return [
                Paragraph(label, styles["DetailLabel"]),
                Paragraph(str(value), vs),
            ], bg

        detail_rows_data = [
            ("Receipt #",    inv.get("invoice", "—"),                   False, None),
            ("Customer",     inv.get("customer", "—"),                   True,  None),
            ("Event Date",   inv.get("event_date", "—"),                 False, None),
        ]

        detail_rows = []
        tbl_style_cmds = [
            ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("TOPPADDING",   (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 7),
            ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]

        for i, (lbl, val, alt, vstyle) in enumerate(detail_rows_data):
            row_data, bg = _det_row(lbl, val, alt, vstyle)
            detail_rows.append(row_data)
            if alt:
                tbl_style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAFC")))

        det_tbl = Table(detail_rows, colWidths=[4.5*cm, _CONTENT_W - 4.5*cm])
        det_tbl.setStyle(TableStyle(tbl_style_cmds))
        story.append(det_tbl)
        story.append(Spacer(1, 0.3*cm))

        amount_rows = [
            [Paragraph("Total Amount", styles["DetailLabel"]),
             Paragraph(f"PHP {total:,.2f}", ParagraphStyle(
                 "amt", fontName="Helvetica", fontSize=10,
                 textColor=_C_DARK, alignment=TA_RIGHT, leading=13))],
            [Paragraph("Amount Paid", styles["DetailLabel"]),
             Paragraph(f"PHP {paid:,.2f}", ParagraphStyle(
                 "paid", fontName="Helvetica", fontSize=10,
                 textColor=_C_GREEN, alignment=TA_RIGHT, leading=13))],
            [Paragraph("Balance Due", ParagraphStyle(
                 "bal_lbl", fontName="Helvetica-Bold", fontSize=10,
                 textColor=_C_DARK, leading=13)),
             Paragraph(f"PHP {balance:,.2f}", ParagraphStyle(
                 "bal_val", fontName="Helvetica-Bold", fontSize=12,
                 textColor=_C_RED if balance > 0 else _C_GREEN,
                 alignment=TA_RIGHT, leading=15))],
        ]

        amt_style = [
            ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("BACKGROUND",   (0, 0), (-1, 0),  _C_LIGHT),
            ("BACKGROUND",   (0, 1), (-1, 1),  _C_WHITE),
            ("BACKGROUND",   (0, 2), (-1, 2),  colors.HexColor("#FFF1F2") if balance > 0 else colors.HexColor("#F0FDF4")),
            ("TOPPADDING",   (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 9),
            ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]

        amt_tbl = Table(amount_rows, colWidths=[_CONTENT_W * 0.55, _CONTENT_W * 0.45])
        amt_tbl.setStyle(TableStyle(amt_style))
        story.append(amt_tbl)
        story.append(Spacer(1, 0.3*cm))

        status_row = Table(
            [[Paragraph("Payment Status:", ParagraphStyle(
                "st_lbl", fontName="Helvetica-Bold", fontSize=9,
                textColor=_C_GRAY, leading=12)),
              Paragraph(status.upper(), ParagraphStyle(
                "st_val", fontName="Helvetica-Bold", fontSize=11,
                textColor=status_color, alignment=TA_RIGHT, leading=14))]],
            colWidths=[_CONTENT_W * 0.5, _CONTENT_W * 0.5],
        )
        status_row.setStyle(TableStyle([
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("BOX",           (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("BACKGROUND",    (0, 0), (-1, 0),  _C_LIGHT),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(status_row)

        story.append(Spacer(1, 0.6*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=_C_BORDER))
        story.append(Spacer(1, 0.3*cm))

        from datetime import datetime as _dt
        printed_str = _dt.now().strftime("%B %d, %Y at %I:%M %p")
        story.append(Paragraph(
            f"Printed: {printed_str}  ·  Thank you for choosing <b>{biz_name}</b>!",
            styles["Footer"],
        ))
        if balance > 0:
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                f"Please settle the remaining balance of PHP {balance:,.2f} before your event date.",
                ParagraphStyle("bal_note", fontName="Helvetica", fontSize=7.5,
                    textColor=_C_RED, alignment=TA_CENTER, leading=10)
            ))

        doc.build(story)
        return True
    except Exception as exc:
        print(f"[exporter] Receipt PDF failed: {exc}")
        return False


def export_excel(path: str, kpis: dict, bookings: list,
                 title: str = "Business Report", period: str = "All Time",
                 biz_name: str = "Jayraldine's Catering") -> bool:
    if not OPENPYXL_OK:
        return False
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        RED   = "E11D48"
        DARK  = "0B1220"
        GRAY  = "374151"
        LIGHT = "F9FAFB"
        WHITE = "FFFFFF"

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _border():
            s = Side(style="thin", color="E5E7EB")
            return Border(left=s, right=s, top=s, bottom=s)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 14

        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = biz_name.upper()
        c.font = Font(name="Calibri", bold=True, size=18, color=RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = f"{title}  |  Period: {period}  |  Generated: {datetime.now().strftime('%b %d, %Y %I:%M %p')}"
        c.font = Font(name="Calibri", size=9, color=GRAY)
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 16
        row += 2

        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "KEY PERFORMANCE INDICATORS"
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = _fill(DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        kpi_pairs = [
            ("Total Bookings",  kpis.get("total_bookings", 0)),
            ("Total Pax",       kpis.get("total_pax", 0)),
            ("Total Revenue",   f"PHP {float(kpis.get('total_revenue', 0)):,.0f}"),
            ("Unpaid Amount",   f"PHP {float(kpis.get('unpaid_amount', 0)):,.0f}"),
            ("Today's Bookings", kpis.get("today_bookings", 0)),
            ("This Week's Bookings", kpis.get("week_bookings", 0)),
        ]
        for label, value in kpi_pairs:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(name="Calibri", bold=True, size=10, color=DARK)
            lc.fill = _fill(LIGHT)
            lc.border = _border()
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = Font(name="Calibri", size=10, color=DARK)
            vc.fill = _fill(WHITE)
            vc.border = _border()
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "BOOKING STATISTICS"
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = _fill(RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        headers = ["Booking Ref", "Client", "Event Date", "Pax", "Total Amount", "Status"]
        for col, hdr in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
            c.fill = _fill(GRAY)
            c.border = _border()
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for i, b in enumerate(bookings):
            bg = LIGHT if i % 2 == 0 else WHITE
            vals = [
                b.get("id", ""), b.get("name", ""), b.get("date", ""),
                b.get("pax", ""), b.get("total", ""), b.get("status", "").capitalize()
            ]
            aligns = ["left","left","center","center","right","center"]
            for col, (val, aln) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Calibri", size=9)
                c.fill = _fill(bg)
                c.border = _border()
                c.alignment = Alignment(horizontal=aln, vertical="center", indent=1 if aln=="left" else 0)
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = f"{biz_name}  •  System-generated report  •  Confidential"
        c.font = Font(name="Calibri", size=8, italic=True, color=GRAY)
        c.alignment = Alignment(horizontal="center")

        wb.save(path)
        return True
    except Exception as exc:
        print(f"[exporter] Excel failed: {exc}")
        return False
