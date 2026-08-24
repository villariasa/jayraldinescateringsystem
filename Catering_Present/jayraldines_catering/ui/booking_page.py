from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QDialog, QFileDialog, QMessageBox, QInputDialog, QScrollArea,
    QCheckBox, QComboBox, QLineEdit, QDateEdit, QSpinBox, QDoubleSpinBox,
    QTabWidget
)
from PySide6.QtCore import Qt, QSize, QDate, QTimer
from PySide6.QtGui import QColor
import csv

from utils.icons import btn_icon_primary, btn_icon_secondary, btn_icon_muted, btn_icon_red, get_icon
from utils.theme import ThemeManager
from components.booking_modal import BookingModal
from components.dialogs import confirm, success, prompt_file_saved
from components.filter_popover import FilterPopover
import utils.repository as repo
import utils.db as _db
from utils.session import get_actor
from utils.signals import app_events
from utils.data_loader import run_async


_STATUS_COLORS = {
    "CONFIRMED": ("#22C55E", "rgba(34,197,94,.15)", "rgba(34,197,94,.3)"),
    "PENDING":   ("#F59E0B", "rgba(245,158,11,.15)", "rgba(245,158,11,.3)"),
    "CANCELLED": ("#EF4444", "rgba(239,68,68,.15)",  "rgba(239,68,68,.3)"),
}


class AnimatedCard(QFrame):
    def __init__(self):
        super().__init__()
        self.setObjectName("card")


def _status_badge(text):
    color, bg, border = _STATUS_COLORS.get(text, ("#9CA3AF", "rgba(156,163,175,.15)", "rgba(156,163,175,.3)"))
    lbl = QLabel(text)
    lbl.setStyleSheet(
        f"font-weight:700;font-size:11px;padding:4px 10px;border-radius:12px;"
        f"background:{bg};color:{color};border:1px solid {border};"
    )
    lbl.setAlignment(Qt.AlignCenter)
    return lbl


def _action_buttons(status, on_approve, on_decline):
    widget = QWidget()
    row = QHBoxLayout(widget)
    row.setContentsMargins(4, 0, 4, 0)
    row.setSpacing(6)
    if status == "PENDING":
        approve_btn = QPushButton()
        approve_btn.setIcon(get_icon("check", color="#22C55E", size=QSize(16, 16)))
        approve_btn.setIconSize(QSize(16, 16))
        approve_btn.setFixedSize(28, 28)
        approve_btn.setStyleSheet(
            "border-radius:14px;"
            "background:rgba(34,197,94,.15);border:1px solid rgba(34,197,94,.3);"
        )
        approve_btn.setCursor(Qt.PointingHandCursor)
        approve_btn.setToolTip("Approve")
        approve_btn.clicked.connect(on_approve)
        row.addWidget(approve_btn)

        decline_btn = QPushButton()
        decline_btn.setIcon(get_icon("x-circle", color="#EF4444", size=QSize(16, 16)))
        decline_btn.setIconSize(QSize(16, 16))
        decline_btn.setFixedSize(28, 28)
        decline_btn.setStyleSheet(
            "border-radius:14px;"
            "background:rgba(239,68,68,.15);border:1px solid rgba(239,68,68,.3);"
        )
        decline_btn.setCursor(Qt.PointingHandCursor)
        decline_btn.setToolTip("Decline")
        decline_btn.clicked.connect(on_decline)
        row.addWidget(decline_btn)
    else:
        locked_lbl = QLabel("—")
        locked_lbl.setStyleSheet("font-size:13px;")
        row.addWidget(locked_lbl)
    row.addStretch()
    return widget


_OCCASIONS_LIST = [
    "Wedding", "Birthday", "Debut", "Corporate Event", "Anniversary", "Christening", "Graduation", "Holiday Party", "Party"
]


class MultiMenuSelectionDialog(QDialog):
    """Interactive modal to select multiple custom dishes / menu offerings for an order."""
    def __init__(self, selected_items=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Custom Menu Dishes")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setMinimumSize(780, 580)
        self.setModal(True)
        self._initial_selected = set(selected_items or [])
        self._checkboxes = []
        self._all_items = repo.get_available_menu_items() or []
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 16, 16, 16)

        container = QFrame()
        container.setObjectName("card")
        lay = QVBoxLayout(container)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(14)

        # Header
        head = QHBoxLayout()
        title = QLabel("Select Custom Menu Dishes")
        title.setObjectName("h3")
        head.addWidget(title)
        head.addStretch()
        close_btn = QPushButton()
        close_btn.setIcon(get_icon("close", color="#6B7280", size=QSize(14, 14)))
        close_btn.setIconSize(QSize(14, 14))
        close_btn.setFixedSize(28, 28)
        close_btn.setStyleSheet("background: transparent; border: none;")
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.clicked.connect(self.reject)
        head.addWidget(close_btn)
        lay.addLayout(head)

        sub = QLabel("Pick multiple culinary dishes to create a custom catering menu for this order:")
        sub.setObjectName("subtitle")
        lay.addWidget(sub)

        # Search Bar + Quick Actions
        ctrl_row = QHBoxLayout()
        ctrl_row.setSpacing(10)
        self._search = QLineEdit()
        self._search.setPlaceholderText("Search dishes (e.g. Pork, Chicken, Seafood, Dessert)...")
        self._search.setFixedHeight(36)
        self._search.textChanged.connect(self._filter_items)
        ctrl_row.addWidget(self._search, 2)

        btn_all = QPushButton("Select All")
        btn_all.setObjectName("secondaryButton")
        btn_all.setFixedHeight(36)
        btn_all.setCursor(Qt.PointingHandCursor)
        btn_all.clicked.connect(self._select_all)

        btn_none = QPushButton("Clear All")
        btn_none.setObjectName("secondaryButton")
        btn_none.setFixedHeight(36)
        btn_none.setCursor(Qt.PointingHandCursor)
        btn_none.clicked.connect(self._clear_all)

        ctrl_row.addWidget(btn_all)
        ctrl_row.addWidget(btn_none)
        lay.addLayout(ctrl_row)

        # Scrollable Categorized Dish Cards
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("background: transparent;")

        content = QWidget()
        self._dishes_lay = QVBoxLayout(content)
        self._dishes_lay.setContentsMargins(0, 0, 0, 0)
        self._dishes_lay.setSpacing(12)

        # Group items by category
        grouped = {}
        for item in self._all_items:
            cat = item.get("category") or "Main Course"
            grouped.setdefault(cat, []).append(item)

        if not grouped:
            empty_lbl = QLabel("No menu items found in the menu database.\nAdd dishes in the Menu module first.")
            empty_lbl.setObjectName("subtitle")
            empty_lbl.setAlignment(Qt.AlignCenter)
            self._dishes_lay.addWidget(empty_lbl)
        else:
            for cat, items in grouped.items():
                cat_group = QFrame()
                cat_group.setObjectName("entryCard")
                cg_lay = QVBoxLayout(cat_group)
                cg_lay.setContentsMargins(14, 12, 14, 12)
                cg_lay.setSpacing(6)

                cat_hdr = QLabel(f"● {cat.upper()}")
                cat_hdr.setStyleSheet("font-size: 11px; font-weight: 800; color: #E11D48; letter-spacing: 1px;")
                cg_lay.addWidget(cat_hdr)

                for it in items:
                    i_name = it.get("name") or it.get("item", "Dish")
                    i_price = float(it.get("price") or 0.0)
                    i_desc = it.get("description") or ""

                    row_frame = QFrame()
                    row_frame._dish_name = i_name
                    row_frame._dish_cat = cat
                    r_lay = QHBoxLayout(row_frame)
                    r_lay.setContentsMargins(6, 4, 6, 4)
                    r_lay.setSpacing(10)

                    cb = QCheckBox(i_name)
                    cb.setStyleSheet("QCheckBox { font-size: 13px; font-weight: 600; color: #F9FAFB; } QCheckBox::indicator { width: 18px; height: 18px; }")
                    if i_name in self._initial_selected:
                        cb.setChecked(True)
                    cb.stateChanged.connect(self._update_counter)

                    if i_desc:
                        desc_lbl = QLabel(f"— {i_desc}")
                        desc_lbl.setStyleSheet("font-size: 11px; color: #64748B;")
                        r_lay.addWidget(cb)
                        r_lay.addWidget(desc_lbl, 1)
                    else:
                        r_lay.addWidget(cb, 1)

                    price_lbl = QLabel(f"₱ {i_price:,.2f} / pax")
                    price_lbl.setStyleSheet("font-size: 12px; font-weight: 700; color: #F59E0B;")
                    r_lay.addWidget(price_lbl)

                    cg_lay.addWidget(row_frame)
                    self._checkboxes.append((cb, it, row_frame))

                self._dishes_lay.addWidget(cat_group)

        self._dishes_lay.addStretch()
        scroll.setWidget(content)
        lay.addWidget(scroll, 1)

        # Bottom Summary & Action Buttons
        bot_bar = QHBoxLayout()
        self._summary_lbl = QLabel("Selected: 0 dishes | Total Unit Rate: ₱ 0.00 / pax")
        self._summary_lbl.setStyleSheet("font-size: 13px; font-weight: 700; color: #38BDF8;")
        bot_bar.addWidget(self._summary_lbl)
        bot_bar.addStretch()

        cancel = QPushButton("Cancel")
        cancel.setObjectName("secondaryButton")
        cancel.setCursor(Qt.PointingHandCursor)
        cancel.clicked.connect(self.reject)
        bot_bar.addWidget(cancel)

        apply_btn = QPushButton("  Apply Dishes to Order")
        apply_btn.setObjectName("primaryButton")
        apply_btn.setIcon(btn_icon_primary("check"))
        apply_btn.setIconSize(QSize(15, 15))
        apply_btn.setCursor(Qt.PointingHandCursor)
        apply_btn.clicked.connect(self.accept)
        bot_bar.addWidget(apply_btn)
        lay.addLayout(bot_bar)

        outer.addWidget(container)
        self._update_counter()

    def _filter_items(self, text: str):
        query = text.strip().lower()
        for cb, it, rf in self._checkboxes:
            dish_name = it.get("name") or it.get("item", "")
            match = (query in dish_name.lower() or query in it.get("category", "").lower() or query in it.get("description", "").lower())
            rf.setVisible(match)

    def _select_all(self):
        for cb, it, rf in self._checkboxes:
            if rf.isVisible():
                cb.setChecked(True)
        self._update_counter()

    def _clear_all(self):
        for cb, it, rf in self._checkboxes:
            cb.setChecked(False)
        self._update_counter()

    def _update_counter(self):
        cnt = sum(1 for cb, it, _ in self._checkboxes if cb.isChecked())
        tot = sum(float(it.get("price") or 0.0) for cb, it, _ in self._checkboxes if cb.isChecked())
        self._summary_lbl.setText(f"Selected: {cnt} dishes | Total Rate: ₱ {tot:,.2f} / pax")

    def get_selected_dishes(self) -> tuple[list[str], float]:
        selected_names = []
        sum_rate = 0.0
        for cb, it, _ in self._checkboxes:
            if cb.isChecked():
                dish_name = it.get("name") or it.get("item", "")
                selected_names.append(dish_name)
                sum_rate += float(it.get("price") or 0.0)
        return selected_names, sum_rate


class AddMultipleBookingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Multiple Orders & Bookings")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.resize(1360, 680)
        self.setMinimumSize(980, 560)
        self.setModal(True)
        self._added_count = 0
        self._customers = repo.get_all_customers() or []
        self._packages = repo.get_all_packages() or []
        self._menu_items = repo.get_available_menu_items() or []
        self._kpi_summary_lbl = QLabel("")
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 16, 16, 16)
        container = QFrame()
        container.setObjectName("card")
        lay = QVBoxLayout(container)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(14)

        header = QHBoxLayout()
        title = QLabel("Add Multiple Orders & Bookings")
        title.setObjectName("h3")
        header.addWidget(title)
        header.addStretch()
        close_btn = QPushButton()
        close_btn.setIcon(get_icon("close", color="#6B7280", size=QSize(14, 14)))
        close_btn.setIconSize(QSize(14, 14))
        close_btn.setFixedSize(28, 28)
        close_btn.setStyleSheet("background: transparent; border: none;")
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.clicked.connect(self.reject)
        header.addWidget(close_btn)
        lay.addLayout(header)

        sub = QLabel("Easily select registered customers, choose packages or multiple custom dishes, and set event dates & down payments:")
        sub.setObjectName("subtitle")
        lay.addWidget(sub)

        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels([
            "Customer (Select or Type) *", "Contact / Phone", "Occasion", "Package / Custom Menu *",
            "Event Date *", "Pax *", "Total Amount (₱) *", "Down Payment (₱)", "Action"
        ])
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setMinimumSectionSize(60)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

        self.table.setColumnWidth(0, 260)  # Customer
        self.table.setColumnWidth(1, 150)  # Contact / Phone
        self.table.setColumnWidth(2, 140)  # Occasion
        self.table.setColumnWidth(3, 340)  # Package / Custom Menu
        self.table.setColumnWidth(4, 150)  # Event Date
        self.table.setColumnWidth(5, 90)   # Pax
        self.table.setColumnWidth(6, 160)  # Total Amount
        self.table.setColumnWidth(7, 160)  # Down Payment
        self.table.setColumnWidth(8, 70)   # Action

        self.table.verticalHeader().setDefaultSectionSize(48)
        lay.addWidget(self.table)

        # Add 5 initial rows
        for _ in range(5):
            self._add_row()

        lay.addWidget(self.table)

        row_actions = QHBoxLayout()
        add_row_btn = QPushButton("  + Add Row")
        add_row_btn.setObjectName("secondaryButton")
        add_row_btn.clicked.connect(self._add_row)
        add_5_btn = QPushButton("  + Add 5 Rows")
        add_5_btn.setObjectName("secondaryButton")
        add_5_btn.clicked.connect(lambda: [self._add_row() for _ in range(5)])
        clear_btn = QPushButton("  Clear Empty Rows")
        clear_btn.setObjectName("secondaryButton")
        clear_btn.clicked.connect(self._clear_empty_rows)

        row_actions.addWidget(add_row_btn)
        row_actions.addWidget(add_5_btn)
        row_actions.addWidget(clear_btn)
        row_actions.addStretch()

        self._kpi_summary_lbl = QLabel("")
        self._kpi_summary_lbl.setStyleSheet("font-size: 12px; font-weight: 700; color: #9CA3AF;")
        row_actions.addWidget(self._kpi_summary_lbl)
        lay.addLayout(row_actions)

        self._err = QLabel("")
        self._err.setStyleSheet("color: #E11D48; font-size: 12px; font-weight: 600;")
        self._err.hide()
        lay.addWidget(self._err)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel = QPushButton("Cancel")
        cancel.setObjectName("secondaryButton")
        cancel.setCursor(Qt.PointingHandCursor)
        cancel.clicked.connect(self.reject)
        save = QPushButton("  Save All Orders")
        save.setObjectName("primaryButton")
        save.setIcon(btn_icon_primary("check"))
        save.setIconSize(QSize(15, 15))
        save.setCursor(Qt.PointingHandCursor)
        save.clicked.connect(self._save_all)
        btn_row.addWidget(cancel)
        btn_row.addWidget(save)
        lay.addLayout(btn_row)

        outer.addWidget(container)
        self._update_grand_totals()

    def _add_row(self):
        r = self.table.rowCount()
        self.table.insertRow(r)

        # 0. Customer Selector (Editable QComboBox with auto-complete & auto-fill)
        cust_combo = QComboBox()
        cust_combo.setEditable(True)
        cust_combo.setFixedHeight(34)
        cust_combo.lineEdit().setPlaceholderText("Select or Type Customer Name *")
        cust_combo.addItem("+ Type New Customer Name...", None)
        for c in self._customers:
            c_name = c.get("name", "Client")
            c_phone = c.get("contact") or "No phone"
            cust_combo.addItem(f"{c_name} ({c_phone})", c)
        cust_combo.setCurrentIndex(0)
        self.table.setCellWidget(r, 0, cust_combo)

        # 1. Contact / Phone
        contact_edit = QLineEdit()
        contact_edit.setPlaceholderText("09171234567")
        contact_edit.setFixedHeight(34)
        self.table.setCellWidget(r, 1, contact_edit)

        # Auto-fill contact on customer selection
        def _on_cust_selected(idx):
            data = cust_combo.currentData()
            if isinstance(data, dict):
                phone = data.get("contact", "")
                if phone:
                    contact_edit.setText(phone)
        cust_combo.currentIndexChanged.connect(_on_cust_selected)

        # 2. Occasion
        occ_combo = QComboBox()
        occ_combo.addItems(_OCCASIONS_LIST)
        occ_combo.setCurrentText("Party")
        occ_combo.setFixedHeight(34)
        self.table.setCellWidget(r, 2, occ_combo)

        # 3. Package / Custom Menu Smart Cell (Dropdown + 'Dishes' multi-select button)
        menu_widget = QWidget()
        m_lay = QHBoxLayout(menu_widget)
        m_lay.setContentsMargins(0, 0, 0, 0)
        m_lay.setSpacing(4)

        menu_combo = QComboBox()
        menu_combo.setFixedHeight(34)

        # Add Packages
        if self._packages:
            for pkg in self._packages:
                p_name = pkg.get("name", "Package")
                p_price = float(pkg.get("price_per_pax") or 0.0)
                menu_combo.addItem(f"📦 {p_name} (₱{p_price:,.0f}/pax)", {
                    "type": "package", "name": p_name, "rate": p_price, "id": pkg.get("id")
                })
        else:
            menu_combo.addItem("📦 Standard Package (₱350/pax)", {
                "type": "package", "name": "Standard Package", "rate": 350.0, "id": None
            })

        # Add Custom Multiple Dishes Entry
        menu_combo.addItem("🍽️ Custom Menu (Select Dishes...)", {
            "type": "custom", "items": [], "rate": 450.0, "name": "Custom Menu"
        })

        # Add Single Dishes
        for it in self._menu_items:
            i_name = it.get("name", "Dish")
            i_price = float(it.get("price") or 0.0)
            menu_combo.addItem(f"🍲 {i_name} (₱{i_price:,.0f}/pax)", {
                "type": "custom", "items": [i_name], "rate": i_price, "name": i_name
            })

        btn_dishes = QPushButton("🍽️ Dishes")
        btn_dishes.setObjectName("secondaryButton")
        btn_dishes.setFixedHeight(34)
        btn_dishes.setToolTip("Open multi-dish selector to pick custom menu dishes")

        m_lay.addWidget(menu_combo, 1)
        m_lay.addWidget(btn_dishes)
        self.table.setCellWidget(r, 3, menu_widget)

        # 4. Event Date
        date_edit = QDateEdit(QDate.currentDate())
        date_edit.setCalendarPopup(True)
        date_edit.setDisplayFormat("yyyy-MM-dd")
        date_edit.setFixedHeight(34)
        self.table.setCellWidget(r, 4, date_edit)

        # 5. Pax
        pax_spin = QSpinBox()
        pax_spin.setRange(1, 10000)
        pax_spin.setValue(50)
        pax_spin.setFixedHeight(34)
        self.table.setCellWidget(r, 5, pax_spin)

        # 6. Total Amount
        total_spin = QDoubleSpinBox()
        total_spin.setRange(0, 9999999)
        total_spin.setPrefix("₱ ")
        total_spin.setDecimals(2)
        total_spin.setSingleStep(500)
        total_spin.setFixedHeight(34)
        self.table.setCellWidget(r, 6, total_spin)

        # 7. Down Payment
        down_spin = QDoubleSpinBox()
        down_spin.setRange(0, 9999999)
        down_spin.setPrefix("₱ ")
        down_spin.setDecimals(2)
        down_spin.setSingleStep(500)
        down_spin.setFixedHeight(34)
        self.table.setCellWidget(r, 7, down_spin)

        # 8. Action: Delete Row
        del_btn = QPushButton()
        del_btn.setIcon(get_icon("trash", color="#EF4444", size=QSize(14, 14)))
        del_btn.setFixedSize(28, 28)
        del_btn.setStyleSheet("background: transparent; border: none;")
        del_btn.setCursor(Qt.PointingHandCursor)
        del_btn.setToolTip("Remove row")
        del_btn.clicked.connect(lambda _, w=menu_widget: self._delete_row_by_widget(w))
        self.table.setCellWidget(r, 8, del_btn)

        # Multi-dish selector handler
        def _open_dish_picker():
            curr_data = menu_combo.currentData() or {}
            initial_dishes = curr_data.get("items", [])
            dlg = MultiMenuSelectionDialog(selected_items=initial_dishes, parent=self)
            if dlg.exec():
                sel_names, sel_rate = dlg.get_selected_dishes()
                if sel_names:
                    label_desc = f"🍽️ Custom ({len(sel_names)} Dishes) - ₱{sel_rate:,.0f}/pax"
                    custom_payload = {
                        "type": "custom", "items": sel_names, "rate": sel_rate, "name": ", ".join(sel_names)
                    }
                    menu_combo.insertItem(0, label_desc, custom_payload)
                    menu_combo.setCurrentIndex(0)
                    _recompute_price()

        btn_dishes.clicked.connect(_open_dish_picker)

        # Auto-recompute total & down payment
        def _recompute_price():
            m_data = menu_combo.currentData() or {}
            rate = float(m_data.get("rate", 350.0))
            pax_val = pax_spin.value()
            computed_tot = rate * pax_val
            total_spin.setValue(computed_tot)
            down_spin.setValue(round(computed_tot * 0.30, 2))
            self._update_grand_totals()

        menu_combo.currentIndexChanged.connect(_recompute_price)
        pax_spin.valueChanged.connect(_recompute_price)
        total_spin.valueChanged.connect(self._update_grand_totals)
        down_spin.valueChanged.connect(self._update_grand_totals)
        _recompute_price()

    def _delete_row_by_widget(self, cell_w: QWidget):
        for r in range(self.table.rowCount()):
            if self.table.cellWidget(r, 3) == cell_w:
                self.table.removeRow(r)
                break
        self._update_grand_totals()

    def _delete_selected_row(self):
        curr = self.table.currentRow()
        if curr >= 0:
            self.table.removeRow(curr)
        self._update_grand_totals()

    def _clear_empty_rows(self):
        r = 0
        while r < self.table.rowCount():
            name_w = self.table.cellWidget(r, 0)
            name_text = ""
            if isinstance(name_w, QComboBox):
                data = name_w.currentData()
                name_text = data.get("name", "") if isinstance(data, dict) else name_w.currentText().strip()
                if "+ Type" in name_text:
                    name_text = ""
            elif isinstance(name_w, QLineEdit):
                name_text = name_w.text().strip()

            if not name_text:
                self.table.removeRow(r)
            else:
                r += 1
        self._update_grand_totals()

    def _update_grand_totals(self):
        count = self.table.rowCount()
        tot_sales = 0.0
        tot_dp = 0.0
        for r in range(count):
            tot_w = self.table.cellWidget(r, 6)
            dp_w = self.table.cellWidget(r, 7)
            if isinstance(tot_w, QDoubleSpinBox):
                tot_sales += tot_w.value()
            if isinstance(dp_w, QDoubleSpinBox):
                tot_dp += dp_w.value()
        self._kpi_summary_lbl.setText(f"Rows: {count}  |  Total Estimated: ₱ {tot_sales:,.2f}  |  Total Down Payment: ₱ {tot_dp:,.2f}")

    def _save_all(self):
        rows_to_save = []
        for r in range(self.table.rowCount()):
            cust_w = self.table.cellWidget(r, 0)
            contact_w = self.table.cellWidget(r, 1)
            occ_w = self.table.cellWidget(r, 2)
            menu_cell = self.table.cellWidget(r, 3)
            date_w = self.table.cellWidget(r, 4)
            pax_w = self.table.cellWidget(r, 5)
            total_w = self.table.cellWidget(r, 6)
            down_w = self.table.cellWidget(r, 7)

            # Customer Name extraction
            cust_name = ""
            existing_cid = None
            if isinstance(cust_w, QComboBox):
                c_data = cust_w.currentData()
                if isinstance(c_data, dict):
                    cust_name = c_data.get("name", "").strip()
                    existing_cid = c_data.get("id")
                else:
                    cust_name = cust_w.currentText().strip()
                    if "+ Type" in cust_name:
                        cust_name = ""
            elif isinstance(cust_w, QLineEdit):
                cust_name = cust_w.text().strip()

            if not cust_name:
                continue

            contact = contact_w.text().strip() if isinstance(contact_w, QLineEdit) else ""
            occ = occ_w.currentText().strip() if isinstance(occ_w, QComboBox) else "Party"

            # Menu & Package extraction
            menu_combo = menu_cell.findChild(QComboBox) if menu_cell else None
            m_data = menu_combo.currentData() if menu_combo else {}
            menu_type = m_data.get("type", "package") if isinstance(m_data, dict) else "package"
            pkg_id = m_data.get("id") if (isinstance(m_data, dict) and menu_type == "package") else None
            menu_val = ""
            if isinstance(m_data, dict):
                if menu_type == "custom":
                    items_list = m_data.get("items", [])
                    menu_val = ", ".join(items_list) if items_list else (m_data.get("name") or "Custom Menu")
                else:
                    menu_val = m_data.get("name", "Standard Package")
            else:
                menu_val = menu_combo.currentText() if menu_combo else "Standard Package"

            date_val = date_w.date().toString("yyyy-MM-dd") if isinstance(date_w, QDateEdit) else ""
            pax = pax_w.value() if isinstance(pax_w, QSpinBox) else 50
            tot = total_w.value() if isinstance(total_w, QDoubleSpinBox) else 0.0
            down = down_w.value() if isinstance(down_w, QDoubleSpinBox) else 0.0

            rows_to_save.append({
                "name": cust_name,
                "customer_id": existing_cid,
                "contact": contact,
                "occasion": occ,
                "date": date_val,
                "event_time": "18:00",
                "venue": "Client Venue",
                "pax": pax,
                "total": tot,
                "amount_paid": down,
                "down_payment": down,
                "payment_mode": "Cash",
                "menu_type": menu_type,
                "menu_value": menu_val,
                "package_id": pkg_id,
                "status": "PENDING",
                "notes": f"Multi-Order entry with {menu_val}",
            })

        if not rows_to_save:
            self._err.setText("Please enter or select at least one customer.")
            self._err.show()
            return

        saved = 0
        for b_data in rows_to_save:
            cid = b_data.get("customer_id")
            if not cid:
                cust = repo.get_customer_by_name(b_data["name"])
                if not cust:
                    cid = repo.add_customer({
                        "name": b_data["name"],
                        "contact": b_data.get("contact", ""),
                        "email": "",
                        "address": b_data.get("venue", "Cebu"),
                        "status": "Active"
                    })
                else:
                    cid = cust["id"]

            b_data["customer_id"] = cid
            if "address" not in b_data:
                b_data["address"] = b_data.get("venue", "Cebu")
            ref = repo.create_booking(b_data)
            if ref:
                saved += 1

        self._added_count = saved
        self.accept()


class BookingPage(QWidget):
    def __init__(self):
        super().__init__()
        self._dirty = True
        self._bookings = []
        self._selected_refs: set[str] = set()
        self._card_checkboxes: dict[str, QCheckBox] = {}
        self._active_filter = "All"
        self._filter_popover = None
        self._search_query = ""
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._on_search_timer_fired)
        self._build_ui()
        db_rows = repo.get_all_bookings()
        self._bookings = db_rows if db_rows else []
        self._populate_table()
        self._dirty = False
        app_events().booking_saved.connect(self._mark_dirty_and_reload)
        app_events().data_changed.connect(self._mark_dirty)

    def _mark_dirty(self):
        self._dirty = True

    def _mark_dirty_and_reload(self):
        self._dirty = True
        if self.isVisible():
            self._refresh_bookings()

    def showEvent(self, event):
        super().showEvent(event)
        if self._dirty:
            self._refresh_bookings()

    def reload(self):
        self._mark_dirty()
        if self.isVisible():
            self._refresh_bookings()

    def _refresh_bookings(self):
        self._dirty = False
        run_async(self, repo.get_all_bookings, self._on_bookings_loaded)

    def _on_bookings_loaded(self, data):
        try:
            from shiboken6 import isValid
            if not isValid(self):
                return
        except Exception:
            pass
        if data is not None:
            self._bookings = data
            self._selected_refs.clear()
            self._populate_table()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(20)

        # Header Row
        header_row = QHBoxLayout()
        v = QVBoxLayout()
        v.setSpacing(4)
        title = QLabel("Orders & Bookings")
        title.setObjectName("pageTitle")
        sub = QLabel("Manage all catering reservations, pending approvals, and confirmed events.")
        sub.setObjectName("subtitle")
        v.addWidget(title)
        v.addWidget(sub)
        header_row.addLayout(v)
        header_row.addStretch()

        self.btn_new = QPushButton("  New Booking")
        self.btn_new.setObjectName("primaryButton")
        self.btn_new.setIcon(btn_icon_primary("plus"))
        self.btn_new.setIconSize(QSize(15, 15))
        self.btn_new.clicked.connect(self._open_modal)
        header_row.addWidget(self.btn_new)

        self.btn_multi_add = QPushButton("  + Quick Multi-Order")
        self.btn_multi_add.setObjectName("secondaryButton")
        self.btn_multi_add.setCursor(Qt.PointingHandCursor)
        self.btn_multi_add.clicked.connect(self._open_multi_add_dialog)
        header_row.addWidget(self.btn_multi_add)

        self.btn_import = QPushButton("  Import")
        self.btn_import.setObjectName("secondaryButton")
        self.btn_import.setIcon(btn_icon_secondary("export"))
        self.btn_import.setIconSize(QSize(15, 15))
        self.btn_import.clicked.connect(self._open_import_dialog)
        header_row.addWidget(self.btn_import)
        layout.addLayout(header_row)

        # Search and Global Filter Bar
        search_filter_row = QHBoxLayout()
        search_filter_row.setSpacing(12)

        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText("Search bookings by client name, reference ID, date, or pax...")
        self._search_input.setFixedHeight(38)
        self._search_input.textChanged.connect(self.filter_search)
        search_filter_row.addWidget(self._search_input, 1)

        self._btn_filter = QPushButton("  Filter")
        self._btn_filter.setObjectName("secondaryButton")
        self._btn_filter.setIcon(btn_icon_secondary("filter"))
        self._btn_filter.setIconSize(QSize(14, 14))
        self._btn_filter.setFixedHeight(38)
        self._btn_filter.clicked.connect(self._open_filter)
        search_filter_row.addWidget(self._btn_filter)

        btn_export = QPushButton("  Export")
        btn_export.setObjectName("secondaryButton")
        btn_export.setIcon(btn_icon_secondary("export"))
        btn_export.setIconSize(QSize(14, 14))
        btn_export.setFixedHeight(38)
        btn_export.clicked.connect(self._export_csv)
        search_filter_row.addWidget(btn_export)

        layout.addLayout(search_filter_row)

        # Two Main Tabs: Pending Bookings & Confirmed Bookings (plus All Bookings)
        self._tabs = QTabWidget()
        self._tabs.currentChanged.connect(self._on_tab_changed)

        # Tab 0: Pending Bookings Tab
        pending_tab, self._pending_cards_layout, self._pending_toolbar = self._create_tab_page("pending")
        self._tabs.addTab(pending_tab, "⏳ Pending Bookings (0)")

        # Tab 1: Confirmed Bookings Tab
        confirmed_tab, self._confirmed_cards_layout, self._confirmed_toolbar = self._create_tab_page("confirmed")
        self._tabs.addTab(confirmed_tab, "✅ Confirmed Bookings (0)")

        # Tab 2: All Bookings Tab
        all_tab, self._all_cards_layout, self._all_toolbar = self._create_tab_page("all")
        self._tabs.addTab(all_tab, "📋 All Bookings (0)")

        layout.addWidget(self._tabs, 1)
        self._populate_table()

    def _create_tab_page(self, tab_type: str):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(20, 18, 20, 18)
        lay.setSpacing(12)

        # Batch toolbar
        batch_toolbar = QHBoxLayout()
        batch_toolbar.setSpacing(12)

        cb_select_all = QCheckBox("Select All")
        cb_select_all.setStyleSheet("QCheckBox { font-weight: 600; font-size: 13px; color: #9CA3AF; }")
        batch_toolbar.addWidget(cb_select_all)

        lbl_selected_count = QLabel("0 selected")
        lbl_selected_count.setStyleSheet("font-size: 12px; color: #6B7280; font-weight: 600;")
        batch_toolbar.addWidget(lbl_selected_count)

        batch_toolbar.addStretch()

        btn_batch_approve = None
        if tab_type in ("pending", "all"):
            btn_batch_approve = QPushButton("  Batch Confirm")
            btn_batch_approve.setObjectName("secondaryButton")
            btn_batch_approve.setIcon(get_icon("check", color="#22C55E", size=QSize(13, 13)))
            btn_batch_approve.setIconSize(QSize(13, 13))
            btn_batch_approve.setEnabled(False)
            btn_batch_approve.setCursor(Qt.PointingHandCursor)
            btn_batch_approve.clicked.connect(self._batch_approve_bookings)
            batch_toolbar.addWidget(btn_batch_approve)

        btn_batch_cancel = QPushButton("  Batch Cancel")
        btn_batch_cancel.setObjectName("secondaryButton")
        btn_batch_cancel.setIcon(get_icon("close", color="#F59E0B", size=QSize(13, 13)))
        btn_batch_cancel.setIconSize(QSize(13, 13))
        btn_batch_cancel.setEnabled(False)
        btn_batch_cancel.setCursor(Qt.PointingHandCursor)
        btn_batch_cancel.clicked.connect(self._batch_cancel_bookings)
        batch_toolbar.addWidget(btn_batch_cancel)

        btn_delete_selected = QPushButton("  Delete Selected")
        btn_delete_selected.setIcon(btn_icon_red("trash"))
        btn_delete_selected.setIconSize(QSize(13, 13))
        btn_delete_selected.setCursor(Qt.PointingHandCursor)
        btn_delete_selected.setEnabled(False)
        btn_delete_selected.setStyleSheet(
            "QPushButton { background: rgba(225,29,72,0.15); border: 1px solid rgba(225,29,72,0.3); color: #E11D48; border-radius: 8px; padding: 6px 14px; font-weight: 600; font-size: 12px; }"
            "QPushButton:hover { background: rgba(225,29,72,0.25); border-color: #E11D48; }"
            "QPushButton:disabled { opacity: 0.35; background: rgba(255,255,255,0.04); border-color: transparent; color: #6B7280; }"
        )
        btn_delete_selected.clicked.connect(self._delete_selected_bookings)
        batch_toolbar.addWidget(btn_delete_selected)

        lay.addLayout(batch_toolbar)

        div = QFrame()
        div.setObjectName("divider")
        div.setFixedHeight(1)
        lay.addWidget(div)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet("background: transparent;")

        cards_container = QWidget()
        cards_container.setStyleSheet("background: transparent;")
        cards_layout = QVBoxLayout(cards_container)
        cards_layout.setContentsMargins(0, 0, 10, 0)
        cards_layout.setSpacing(12)

        scroll_area.setWidget(cards_container)
        lay.addWidget(scroll_area, 1)

        toolbar_bundle = {
            "select_all": cb_select_all,
            "selected_lbl": lbl_selected_count,
            "batch_approve": btn_batch_approve,
            "batch_cancel": btn_batch_cancel,
            "delete_selected": btn_delete_selected,
            "container": cards_container,
            "tab_type": tab_type
        }

        cb_select_all.stateChanged.connect(lambda state, t=tab_type: self._toggle_select_tab(t, state))

        return page, cards_layout, toolbar_bundle

    def _on_tab_changed(self, index: int):
        self._update_selection_ui()

    def _visible_bookings(self):
        rows = self._bookings
        f = self._active_filter
        if f and f != "All":
            if isinstance(f, list):
                rows = [b for b in rows if b.get("status") in f]
            else:
                rows = [b for b in rows if b.get("status") == f]

        q = (getattr(self, "_search_query", "") or "").strip().lower()
        if q:
            def _match(b):
                terms = [
                    str(b.get("name", "")),
                    str(b.get("id", "")),
                    str(b.get("date", "")),
                    str(b.get("pax", "")),
                    str(b.get("total", "")),
                    str(b.get("occasion", "")),
                    str(b.get("venue", "")),
                    str(b.get("notes", "")),
                    str(b.get("status", "")),
                    str(b.get("payment_mode", "")),
                    str(b.get("contact", "")),
                ]
                combined = " ".join(terms).lower()
                return q in combined

            rows = [b for b in rows if _match(b)]
        return rows

    def _populate_table(self, data=None):
        self.setUpdatesEnabled(False)
        try:
            raw_rows = data if data is not None else self._visible_bookings()
            self._card_checkboxes.clear()

            # Partition rows into Pending, Confirmed/Completed, and All
            pending_rows = [b for b in raw_rows if b.get("status") == "PENDING"]
            confirmed_rows = [b for b in raw_rows if b.get("status") in ("CONFIRMED", "COMPLETED")]
            all_rows = raw_rows

            # Populate Pending Tab
            self._populate_card_layout(self._pending_cards_layout, pending_rows, "No pending bookings found.")

            # Populate Confirmed Tab
            self._populate_card_layout(self._confirmed_cards_layout, confirmed_rows, "No confirmed bookings found.")

            # Populate All Tab
            self._populate_card_layout(self._all_cards_layout, all_rows, "No bookings found.")

            # Update Tab Title Counts
            if hasattr(self, "_tabs"):
                self._tabs.setTabText(0, f"⏳ Pending Bookings ({len(pending_rows)})")
                self._tabs.setTabText(1, f"✅ Confirmed Bookings ({len(confirmed_rows)})")
                self._tabs.setTabText(2, f"📋 All Bookings ({len(all_rows)})")

            self._update_selection_ui()
        finally:
            self.setUpdatesEnabled(True)

    def _populate_card_layout(self, layout: QVBoxLayout, rows: list[dict], empty_msg: str):
        if not layout:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item:
                w = item.widget()
                if w:
                    w.hide()
                    w.deleteLater()

        if not rows:
            empty_lbl = QLabel(empty_msg)
            empty_lbl.setObjectName("subtitle")
            empty_lbl.setAlignment(Qt.AlignCenter)
            empty_lbl.setStyleSheet("font-size: 13px; color: #64748B; padding: 24px;")
            layout.addWidget(empty_lbl)
        else:
            for b in rows:
                card = self._create_booking_card(b)
                layout.addWidget(card)

        layout.addStretch()

    def _create_booking_card(self, b: dict) -> QFrame:
        bref = b["id"]
        card = QFrame()
        card.setObjectName("entryCard")
        lay = QHBoxLayout(card)
        lay.setContentsMargins(14, 14, 18, 14)
        lay.setSpacing(14)

        # Col 0: Checkbox
        cb = QCheckBox()
        cb.setChecked(bref in self._selected_refs)
        cb.stateChanged.connect(lambda state, ref=bref: self._on_card_checked(ref, state))
        self._card_checkboxes[bref] = cb
        lay.addWidget(cb, alignment=Qt.AlignVCenter)

        # Col 1: Date & Reference ID
        c1 = QVBoxLayout()
        c1.setSpacing(2)
        ref_lbl = QLabel(b["id"])
        ref_lbl.setStyleSheet("font-weight: 800; font-size: 13px; color: #E11D48;")
        date_lbl = QLabel(b["date"])
        date_lbl.setObjectName("subtitle")
        c1.addWidget(ref_lbl)
        c1.addWidget(date_lbl)
        lay.addLayout(c1, 1)

        # Col 2: Client Name & Pax
        c2 = QVBoxLayout()
        c2.setSpacing(2)
        name_lbl = QLabel(b["name"])
        name_lbl.setStyleSheet("font-weight: 700; font-size: 14px;")
        pax_lbl = QLabel(f"{b['pax']} pax")
        pax_lbl.setObjectName("subtitle")
        c2.addWidget(name_lbl)
        c2.addWidget(pax_lbl)
        lay.addLayout(c2, 2)

        # Col 3: Total Amount
        c3 = QVBoxLayout()
        c3.setSpacing(2)
        tot_title = QLabel("TOTAL")
        tot_title.setStyleSheet("font-size: 10px; font-weight: 700; color: #6B7280; letter-spacing: 0.5px;")
        tot_val = QLabel(str(b["total"]))
        tot_val.setStyleSheet("font-weight: 800; font-size: 14px; color: #F59E0B;")
        c3.addWidget(tot_title)
        c3.addWidget(tot_val)
        lay.addLayout(c3, 1)

        # Col 4: Status Badge & Reason/Approvals
        c4 = QVBoxLayout()
        c4.setSpacing(4)
        c4.setAlignment(Qt.AlignCenter)
        c4.addWidget(_status_badge(b["status"]), alignment=Qt.AlignLeft)
        if b["status"] == "CANCELLED" and b.get("cancellation_reason"):
            reason_lbl = QLabel(b["cancellation_reason"])
            reason_lbl.setStyleSheet("color:#DC2626;font-size:10px;font-style:italic;")
            reason_lbl.setWordWrap(True)
            c4.addWidget(reason_lbl)
        elif b["status"] == "PENDING":
            bref = b["id"]
            c4.addWidget(_action_buttons(
                b["status"],
                on_approve=lambda _, r=bref: self._approve_booking(r),
                on_decline=lambda _, r=bref: self._decline_booking(r)
            ))
        lay.addLayout(c4, 2)

        # Col 5: Actions (Edit, Delete, Confirmation)
        actions_w = QFrame()
        actions_w.setStyleSheet("background: transparent;")
        actions_l = QHBoxLayout(actions_w)
        actions_l.setContentsMargins(0, 0, 0, 0)
        actions_l.setSpacing(6)

        bref = b["id"]

        edit_btn = QPushButton()
        edit_btn.setIcon(get_icon("edit", color="#9CA3AF", size=QSize(13, 13)))
        edit_btn.setIconSize(QSize(13, 13))
        edit_btn.setFixedSize(30, 30)
        edit_btn.setStyleSheet("background:transparent;border:none;")
        edit_btn.setCursor(Qt.PointingHandCursor)
        edit_btn.setToolTip("Edit booking")
        edit_btn.setEnabled(b["status"] == "PENDING")
        if b["status"] != "PENDING":
            edit_btn.setStyleSheet("background:transparent;border:none;opacity:0.3;")
        edit_btn.clicked.connect(lambda _, r=bref: self._edit_booking(r))

        del_btn = QPushButton()
        del_btn.setIcon(btn_icon_red("trash"))
        del_btn.setIconSize(QSize(13, 13))
        del_btn.setFixedSize(30, 30)
        del_btn.setStyleSheet("border:none;background:transparent;")
        del_btn.setCursor(Qt.PointingHandCursor)
        del_btn.setToolTip("Delete booking")
        del_btn.clicked.connect(lambda _, r=bref: self._delete_booking(r))

        confirm_btn = QPushButton()
        confirm_btn.setIcon(get_icon("bell", color="#9CA3AF", size=QSize(13, 13)))
        confirm_btn.setIconSize(QSize(13, 13))
        confirm_btn.setFixedSize(30, 30)
        confirm_btn.setStyleSheet("background:transparent;border:none;")
        confirm_btn.setCursor(Qt.PointingHandCursor)
        confirm_btn.setToolTip("Send Confirmation Email")
        confirm_btn.setEnabled(b["status"] == "CONFIRMED")
        if b["status"] != "CONFIRMED":
            confirm_btn.setStyleSheet("background:transparent;border:none;opacity:0.3;")
        confirm_btn.clicked.connect(lambda _, r=bref: self._send_confirmation(r))

        actions_l.addWidget(edit_btn)
        actions_l.addWidget(del_btn)
        actions_l.addWidget(confirm_btn)
        lay.addWidget(actions_w)

        return card

    def _approve_booking(self, ref):
        b = next((x for x in self._bookings if x["id"] == ref), None)
        if not b:
            return

        db_id = b.get("db_id")
        detail = repo.get_booking_detail(db_id) if db_id else b
        if not detail:
            detail = b

        from components.confirm_booking_dialog import ConfirmBookingDialog, _parse_amount
        from datetime import date as _d

        dlg = ConfirmBookingDialog(detail, parent=self)
        if not dlg.exec():
            return

        try:
            pay_amt = dlg.get_payment_amount()
            is_auto_pay = dlg.is_auto_pay_checked()
            pay_method = dlg.get_payment_method()
            remarks = dlg.get_payment_remarks()

            tot_val = _parse_amount(detail.get("total") or detail.get("total_amount") or b.get("total", 0.0))
            paid_val = _parse_amount(detail.get("amount_paid") or detail.get("down_payment") or 0.0)
            rem = max(0.0, tot_val - paid_val)

            if db_id:
                if pay_amt > 0:
                    try:
                        repo.pay_invoice(db_id, payment_amount=pay_amt, payment_date=_d.today(), method=pay_method, note=remarks)
                    except Exception as p_err:
                        print(f"[booking] pay_invoice error: {p_err}")
                
                # Explicitly guarantee booking status is persisted as CONFIRMED in database
                repo.update_booking_status(db_id, "CONFIRMED")

                try:
                    if detail.get("customer_id"):
                        repo.recalculate_loyalty(detail["customer_id"])
                except Exception:
                    pass

            b["status"] = "CONFIRMED"
            self._populate_table()
            if pay_amt >= rem and rem > 0:
                msg = f"Order {b['id']} confirmed & marked as FULLY PAID (₱{tot_val:,.2f})!"
            elif pay_amt > 0:
                msg = f"Order {b['id']} confirmed & payment of ₱{pay_amt:,.2f} recorded!"
            else:
                msg = f"Order {b['id']} confirmed successfully!"
            success(self, message=msg)
            repo.write_audit_log(get_actor(), "APPROVE", "bookings", db_id, None, {"status": "CONFIRMED", "payment_amount": pay_amt})

            try:
                repo.push_notification(
                    "success",
                    "Booking Confirmed",
                    f"Booking for {b.get('name', '')} on {b.get('date', '')} has been confirmed.",
                    "#22C55E",
                )
            except Exception:
                pass

            if db_id:
                self._send_confirmation_auto(b)
            app_events().booking_updated.emit()
            app_events().booking_saved.emit()
            app_events().data_changed.emit()
        except Exception as exc:
            QMessageBox.warning(self, "Cannot Approve", str(exc))

    def _decline_booking(self, ref):
        b = next((x for x in self._bookings if x["id"] == ref), None)
        if not b or b["status"] != "PENDING":
            return
        reason, ok = QInputDialog.getText(
            self, "Cancellation Reason",
            f"Enter reason for declining booking for '{b['name']}' (optional):"
        )
        if not ok:
            return
        if not confirm(self, title="Decline Booking",
                       message=f"Decline booking for '{b['name']}'? This will mark it as Cancelled.",
                       confirm_label="Decline", danger=True):
            return
        b["status"] = "CANCELLED"
        if b.get("db_id"):
            repo.update_booking_status(b["db_id"], "CANCELLED", reason.strip() or None)
            repo.write_audit_log(get_actor(), "CANCEL", "bookings", b["db_id"], None, {"status": "CANCELLED", "reason": reason.strip()})
        self._populate_table()
        success(self, message="Booking declined.")

    def _send_confirmation_auto(self, b: dict) -> None:
        """Auto-trigger confirmation email on approval (best-effort, silent on failure)."""
        try:
            detail = repo.get_booking_detail(b["db_id"]) if b.get("db_id") else None
            if not detail:
                return
            biz = repo.get_business_info()
            booking_data = {**detail, "business_contact": biz.get("contact", "")}
            smtp = repo.get_smtp_config()
            if detail.get("email") and smtp.get("smtp_host"):
                from utils.mailer import send_booking_confirmation_email
                ok, _ = send_booking_confirmation_email(smtp, detail["email"], booking_data)
                if ok and detail.get("db_id"):
                    repo.log_confirmation_sent(detail["db_id"], "email")
        except Exception as exc:
            print(f"[booking] auto-confirm send failed: {exc}")

    def _send_confirmation(self, ref: str) -> None:
        """Manual resend of confirmation for a CONFIRMED booking."""
        b = next((x for x in self._bookings if x["id"] == ref), None)
        if not b or b["status"] != "CONFIRMED" or not b.get("db_id"):
            QMessageBox.information(self, "Not Available",
                "Confirmation can only be resent for confirmed bookings with a database record.")
            return
        detail = repo.get_booking_detail(b["db_id"])
        if not detail:
            return
        biz = repo.get_business_info()
        booking_data = {**detail, "business_contact": biz.get("contact", "")}
        smtp = repo.get_smtp_config()
        errors = []
        sent_email = False
        if detail.get("email") and smtp.get("smtp_host"):
            from utils.mailer import send_booking_confirmation_email
            ok, err = send_booking_confirmation_email(smtp, detail["email"], booking_data)
            if ok:
                repo.log_confirmation_sent(detail["db_id"], "email")
                sent_email = True
            else:
                errors.append(f"Email: {err}")
        if sent_email:
            QMessageBox.information(self, "Confirmation Sent",
                f"Booking confirmation has been sent via email to:\n{detail['email']}")
        elif errors:
            QMessageBox.warning(self, "Send Failed", "\n".join(errors))
        else:
            QMessageBox.information(self, "No Channel",
                "No email configured for this booking.\n"
                "Ensure customer has an email and SMTP is configured in Settings.")

    def _delete_booking(self, ref):
        b = next((x for x in self._bookings if x["id"] == ref), None)
        if not b:
            return
        if not confirm(self, title="Delete Booking",
                       message=f"Are you sure you want to delete booking for '{b['name']}'? This cannot be undone.",
                       confirm_label="Delete", danger=True):
            return
        if b.get("db_id"):
            repo.delete_booking(b["db_id"])
        self._bookings = [x for x in self._bookings if x["id"] != ref]
        self._populate_table()
        success(self, message="Booking deleted successfully.")

    def _edit_booking(self, ref):
        b = next((x for x in self._bookings if x["id"] == ref), None)
        if not b:
            return
        db_id = b.get("db_id")
        detail = repo.get_booking_detail(db_id) if db_id else None
        if not detail:
            detail = b
        modal = BookingModal(self, booking_data=detail)
        modal.booking_saved.connect(lambda data, orig=b: self._update_booking(orig, data))
        modal.exec()

    def _update_booking(self, orig, data):
        db_id = orig.get("db_id") or data.get("db_id")
        if db_id:
            repo.update_booking(db_id, data)
        self.reload()
        success(self, message="Order details updated successfully.")
        app_events().booking_updated.emit()

    def _open_modal(self):
        modal = BookingModal(self)
        modal.booking_saved.connect(self._add_booking)
        modal.exec()

    def _add_booking(self, data):
        result = repo.create_booking(data)
        if not result:
            QMessageBox.warning(self, "Booking Failed", "Failed to save booking to database. Please check application logs.")
            return
        bkg_id = result["booking_ref"]
        db_id  = result["booking_id"]

        self._bookings.append({
            "db_id":  db_id,
            "date":   data["date"],
            "id":     bkg_id,
            "name":   data["name"],
            "pax":    str(data["pax"]),
            "total":  f"₱ {data['total']:,}",
            "status": data["status"],
        })
        self._populate_table()

        email_status = self._send_approval_request(data, bkg_id)
        if email_status is True:
            msg = f"Booking created successfully.\nConfirmation email sent to {data.get('email', '')}."
        elif email_status is False:
            msg = "Booking created successfully.\nCould not send confirmation email — check SMTP settings."
        else:
            msg = "Booking created successfully."
        success(self, message=msg)

        try:
            repo.push_notification(
                type_="info",
                title="New Booking Request",
                message=f"{data['name']} submitted a new booking request for {data['pax']} pax on {data['date']}.",
                color="#3B82F6"
            )
        except Exception as exc:
            print(f"[Notification] Failed to create in-app notification: {exc}")

        app_events().booking_saved.emit()

    def _send_approval_request(self, data: dict, bkg_ref: str):
        """Send a booking approval request email to the customer.
        Returns True on success, False on failure, None if email/SMTP not configured."""
        try:
            email = data.get("email", "")
            if not email or "@" not in email:
                return None
            smtp = repo.get_smtp_config()
            if not smtp.get("smtp_host"):
                return None
            biz = repo.get_business_info()
            booking_data = {
                **data,
                "booking_ref":      bkg_ref,
                "business_contact": biz.get("contact", ""),
                "business_name":    biz.get("name", "Jayraldine's Catering"),
                "event_date":       data.get("date", "—"),
                "event_time":       data.get("time", "—"),
            }
            from utils.mailer import send_booking_approval_request_email
            ok, err = send_booking_approval_request_email(smtp, email, booking_data)
            if ok:
                print(f"[booking] Approval request email sent to {email}")
                return True
            else:
                print(f"[booking] Approval request email failed: {err}")
                return False
        except Exception as exc:
            print(f"[booking] send_approval_request failed: {exc}")
            return False

    def _open_filter(self):
        if self._filter_popover is None:
            win = self.window()
            self._filter_popover = FilterPopover(
                parent=win if win else self,
                statuses=["All", "PENDING", "CONFIRMED", "CANCELLED"],
            )
            self._filter_popover.filter_applied.connect(self._on_filter_applied)
        self._filter_popover.toggle_anchored(self._btn_filter)

    def _on_filter_applied(self, result):
        status = result.get("statuses", ["All"])[0]
        self._active_filter = "All" if not status or status == "All" else status
        self._populate_table()

    def _export_csv(self):
        from components.export_dialog import ExportWizardDialog
        dlg = ExportWizardDialog(parent=self)
        dlg.exec()

    def filter_search(self, text):
        self._search_query = str(text or "")
        if hasattr(self, "_search_timer"):
            self._search_timer.start(80)
        else:
            self._populate_table()

    def _on_search_timer_fired(self):
        self._populate_table()

    def search(self, query: str):
        if hasattr(self, "_search_input"):
            self._search_input.setText(query)
        else:
            self.filter_search(query)

    def _open_import_dialog(self):
        from components.import_dialog import ImportWizardDialog
        dlg = ImportWizardDialog(default_entity="bookings", parent=self)
        if dlg.exec():
            self._refresh_bookings()

    def _on_card_checked(self, ref: str, state):
        if not ref:
            return
        if state:
            self._selected_refs.add(ref)
        else:
            self._selected_refs.discard(ref)
        self._update_selection_ui()

    def _toggle_select_tab(self, tab_type: str, state):
        if tab_type == "pending":
            target_refs = [b["id"] for b in self._visible_bookings() if b.get("status") == "PENDING"]
        elif tab_type == "confirmed":
            target_refs = [b["id"] for b in self._visible_bookings() if b.get("status") in ("CONFIRMED", "COMPLETED")]
        else:
            target_refs = [b["id"] for b in self._visible_bookings()]

        if state:
            self._selected_refs.update(target_refs)
        else:
            self._selected_refs.difference_update(target_refs)

        for ref, cb in self._card_checkboxes.items():
            if ref in target_refs:
                cb.blockSignals(True)
                cb.setChecked(bool(state))
                cb.blockSignals(False)

        self._update_selection_ui()

    def _update_selection_ui(self):
        count = len(self._selected_refs)
        pending_count = sum(1 for r in self._selected_refs if any(b["id"] == r and b.get("status") == "PENDING" for b in self._bookings))
        cancellable_count = sum(1 for r in self._selected_refs if any(b["id"] == r and b.get("status") != "CANCELLED" for b in self._bookings))

        toolbars = [
            getattr(self, "_pending_toolbar", None),
            getattr(self, "_confirmed_toolbar", None),
            getattr(self, "_all_toolbar", None),
        ]

        for tb in toolbars:
            if not tb:
                continue
            tb["selected_lbl"].setText(f"{count} selected")
            tb["delete_selected"].setEnabled(count > 0)
            tb["delete_selected"].setText(f"  Delete Selected ({count})" if count > 0 else "  Delete Selected")

            if tb.get("batch_approve"):
                tb["batch_approve"].setEnabled(pending_count > 0)
                tb["batch_approve"].setText(f"  Batch Confirm ({pending_count})" if pending_count > 0 else "  Batch Confirm")

            if tb.get("batch_cancel"):
                tb["batch_cancel"].setEnabled(cancellable_count > 0)
                tb["batch_cancel"].setText(f"  Batch Cancel ({cancellable_count})" if cancellable_count > 0 else "  Batch Cancel")

            tab_t = tb.get("tab_type")
            if tab_t == "pending":
                t_refs = [b["id"] for b in self._visible_bookings() if b.get("status") == "PENDING"]
            elif tab_t == "confirmed":
                t_refs = [b["id"] for b in self._visible_bookings() if b.get("status") in ("CONFIRMED", "COMPLETED")]
            else:
                t_refs = [b["id"] for b in self._visible_bookings()]

            all_checked = len(t_refs) > 0 and all(r in self._selected_refs for r in t_refs)
            tb["select_all"].blockSignals(True)
            tb["select_all"].setChecked(all_checked)
            tb["select_all"].blockSignals(False)

    def _delete_selected_bookings(self):
        if not self._selected_refs:
            return
        count = len(self._selected_refs)
        if not confirm(self, title="Delete Multiple Bookings",
                       message=f"Are you sure you want to permanently delete {count} selected booking(s)/order(s)?\nThis cannot be undone.",
                       confirm_label=f"Delete {count} Orders", danger=True):
            return

        db_ids_or_refs = []
        for ref in list(self._selected_refs):
            b = next((x for x in self._bookings if x["id"] == ref), None)
            if b and b.get("db_id"):
                db_ids_or_refs.append(b["db_id"])
            else:
                db_ids_or_refs.append(ref)

        deleted = repo.delete_multiple_bookings(db_ids_or_refs)
        self._selected_refs.clear()
        self.reload()
        app_events().booking_saved.emit()
        app_events().data_changed.emit()
        success(self, message=f"Successfully deleted {deleted} order(s).")

    def _batch_approve_bookings(self):
        if not self._selected_refs:
            return

        pending_refs = [
            ref for ref in list(self._selected_refs)
            if any(b["id"] == ref and b.get("status") == "PENDING" for b in self._bookings)
        ]

        if not pending_refs:
            QMessageBox.information(
                self,
                "Already Confirmed",
                "None of the selected orders are PENDING.\nAll selected bookings are already confirmed or completed."
            )
            return

        count = len(pending_refs)
        if not confirm(self, title="Batch Confirm Bookings",
                       message=f"Confirm and approve {count} pending booking(s)?\nThis will mark them as CONFIRMED and mark them as FULLY PAID by default.",
                       confirm_label=f"Confirm & Fully Pay {count} Bookings"):
            return

        from datetime import date as _d
        from components.confirm_booking_dialog import _parse_amount

        approved_cnt = 0
        for ref in pending_refs:
            try:
                b = next((x for x in self._bookings if x["id"] == ref), None)
                if b and b.get("db_id"):
                    db_id = b["db_id"]
                    detail = repo.get_booking_detail(db_id) or b
                    tot_val = _parse_amount(detail.get("total") or detail.get("total_amount") or b.get("total", 0.0))
                    paid_val = _parse_amount(detail.get("amount_paid") or detail.get("down_payment") or 0.0)
                    rem = max(0.0, tot_val - paid_val)
                    if rem > 0:
                        try:
                            repo.pay_invoice(db_id, payment_amount=rem, payment_date=_d.today(), method="Cash", note="Batch confirmed & auto fully paid")
                        except Exception as p_err:
                            print(f"[booking] batch pay error: {p_err}")
                    repo.update_booking_status(db_id, "CONFIRMED")
                    approved_cnt += 1
            except Exception as exc:
                print(f"[booking] batch confirm error for {ref}: {exc}")

        self._selected_refs.clear()
        self.reload()
        app_events().booking_saved.emit()
        app_events().data_changed.emit()
        success(self, message=f"Successfully batch confirmed {approved_cnt} booking(s) as fully paid.")

    def _batch_cancel_bookings(self):
        if not self._selected_refs:
            return

        cancellable_refs = [
            ref for ref in list(self._selected_refs)
            if any(b["id"] == ref and b.get("status") != "CANCELLED" for b in self._bookings)
        ]

        if not cancellable_refs:
            QMessageBox.information(
                self,
                "Already Cancelled",
                "None of the selected orders can be cancelled.\nAll selected bookings are already cancelled."
            )
            return

        count = len(cancellable_refs)
        if not confirm(self, title="Batch Cancel Bookings",
                       message=f"Cancel {count} selected booking(s)?",
                       confirm_label=f"Cancel {count} Bookings", danger=True):
            return

        cancelled_cnt = 0
        for ref in cancellable_refs:
            b = next((x for x in self._bookings if x["id"] == ref), None)
            if b and b.get("db_id"):
                repo.update_booking_status(b["db_id"], "CANCELLED", "Batch cancelled by user")
                cancelled_cnt += 1

        self._selected_refs.clear()
        self.reload()
        app_events().booking_saved.emit()
        app_events().data_changed.emit()
        success(self, message=f"Batch cancelled {cancelled_cnt} booking(s).")

    def _open_multi_add_dialog(self):
        dlg = AddMultipleBookingsDialog(self)
        if dlg.exec():
            self.reload()
            app_events().booking_saved.emit()
            app_events().data_changed.emit()
            success(self, message=f"Added {dlg._added_count} booking(s) successfully.")

    def _export_csv(self):
        bookings = self._visible_bookings()
        if not bookings:
            bookings = self._all_bookings or []

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Bookings", "jayraldines_bookings_export.xlsx", "Excel Spreadsheet (*.xlsx);;CSV Files (*.csv)"
        )
        if not path:
            return

        ext = os.path.splitext(path)[1].lower()
        if not ext:
            path = f"{path}.xlsx"
            ext = ".xlsx"

        headers = [
            "Reference ID", "Customer Name", "Contact Number", "Email Address",
            "Event Date", "Event Time", "Occasion", "Venue", "Pax",
            "Total Amount (₱)", "Down Payment (₱)", "Balance (₱)", "Status", "Notes"
        ]

        rows = []
        for b in bookings:
            total_amt = float(b.get("total_amount") or 0.0)
            down_amt = float(b.get("down_payment") or 0.0)
            balance = max(0.0, total_amt - down_amt)
            rows.append([
                b.get("ref_id") or b.get("id") or "—",
                b.get("client_name") or b.get("customer_name") or "—",
                b.get("contact") or b.get("phone") or "—",
                b.get("email") or "—",
                str(b.get("event_date") or ""),
                str(b.get("event_time") or ""),
                b.get("occasion") or "—",
                b.get("venue") or "—",
                int(b.get("pax") or 0),
                f"{total_amt:,.2f}",
                f"{down_amt:,.2f}",
                f"{balance:,.2f}",
                str(b.get("status") or "PENDING").upper(),
                b.get("notes") or ""
            ])

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Bookings & Orders"

                header_fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")
                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                data_font = Font(name="Arial", size=10)
                thin_border = Border(
                    left=Side(style='thin', color='CBD5E1'),
                    right=Side(style='thin', color='CBD5E1'),
                    top=Side(style='thin', color='CBD5E1'),
                    bottom=Side(style='thin', color='CBD5E1')
                )

                all_rows = [headers] + rows
                for row_idx, r in enumerate(all_rows, start=1):
                    ws.append(r)
                    for col_idx in range(1, len(r) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        if row_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.font = data_font
                            cell.alignment = Alignment(vertical="center")

                ws.row_dimensions[1].height = 28
                for r_idx in range(2, len(all_rows) + 1):
                    ws.row_dimensions[r_idx].height = 22

                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or "")
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

                wb.save(path)
                prompt_file_saved(self, path, title="Export Complete", message=f"Exported {len(rows)} booking(s) to Excel successfully.")
            except Exception as e:
                QMessageBox.warning(self, "Export Failed", f"Could not export bookings: {e}")
        else:
            try:
                import csv
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)
                prompt_file_saved(self, path, title="Export Complete", message=f"Exported {len(rows)} booking(s) to CSV successfully.")
            except Exception as e:
                QMessageBox.warning(self, "Export Failed", f"Could not export bookings: {e}")

