import os
from datetime import datetime, date as _dt_date
from typing import Optional, List, Dict, Any

_dt_datetime = datetime

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, Image, KeepTogether
    )
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


if REPORTLAB_OK:
    _C_RED    = colors.HexColor("#E11D48")
    _C_DARK   = colors.HexColor("#0B1220")
    _C_GRAY   = colors.HexColor("#374151")
    _C_LIGHT  = colors.HexColor("#F9FAFB")
    _C_WHITE  = colors.white
    _C_MUTED  = colors.HexColor("#6B7280")
    _C_BORDER = colors.HexColor("#E5E7EB")
    _C_GREEN  = colors.HexColor("#22C55E")
    _C_AMBER  = colors.HexColor("#F59E0B")
    _C_BLUE   = colors.HexColor("#3B82F6")
else:
    _C_RED = _C_DARK = _C_GRAY = _C_LIGHT = _C_WHITE = _C_MUTED = _C_BORDER = _C_GREEN = _C_AMBER = _C_BLUE = None

from utils.paths import resource_path


def _logo_path() -> str:
    p = resource_path("assets", "logo.png")
    if not os.path.exists(p):
        p = resource_path("assets", "logo.jpg")
    return p

_PAGE_W = A4[0] if REPORTLAB_OK else 595
_MARGIN = 1.5 * cm if REPORTLAB_OK else 0
_CONTENT_W = _PAGE_W - 2 * _MARGIN


def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle("Brand",
        fontName="Helvetica-Bold", fontSize=22, textColor=_C_RED,
        spaceAfter=2, alignment=TA_LEFT, leading=26))
    s.add(ParagraphStyle("BrandSub",
        fontName="Helvetica", fontSize=9.5, textColor=_C_MUTED,
        spaceAfter=0, alignment=TA_LEFT, leading=13))
    s.add(ParagraphStyle("BrandSubRight",
        fontName="Helvetica", fontSize=9.5, textColor=_C_MUTED,
        spaceAfter=0, alignment=TA_RIGHT, leading=13))
    s.add(ParagraphStyle("SectionHead",
        fontName="Helvetica-Bold", fontSize=12, textColor=_C_DARK,
        spaceBefore=16, spaceAfter=8, alignment=TA_LEFT, leading=16))
    s.add(ParagraphStyle("KpiLabel",
        fontName="Helvetica-Bold", fontSize=8.5, textColor=_C_MUTED,
        spaceAfter=2, alignment=TA_CENTER, leading=11))
    s.add(ParagraphStyle("KpiValue",
        fontName="Helvetica-Bold", fontSize=16, textColor=_C_DARK,
        spaceAfter=0, alignment=TA_CENTER, leading=19))
    s.add(ParagraphStyle("Footer",
        fontName="Helvetica", fontSize=8, textColor=_C_MUTED,
        alignment=TA_CENTER, leading=11))
    s.add(ParagraphStyle("ReceiptTitle",
        fontName="Helvetica-Bold", fontSize=18, textColor=_C_WHITE,
        alignment=TA_LEFT, leading=22, spaceAfter=4))
    s.add(ParagraphStyle("ReceiptSub",
        fontName="Helvetica", fontSize=9.5, textColor=colors.HexColor("#CBD5E1"),
        alignment=TA_LEFT, leading=13))
    s.add(ParagraphStyle("DetailLabel",
        fontName="Helvetica-Bold", fontSize=9.5, textColor=_C_GRAY,
        leading=13, wordWrap="CJK"))
    s.add(ParagraphStyle("DetailValue",
        fontName="Helvetica", fontSize=9.5, textColor=_C_DARK,
        leading=13, wordWrap="CJK"))
    s.add(ParagraphStyle("DetailValueBold",
        fontName="Helvetica-Bold", fontSize=9.5, textColor=_C_DARK,
        leading=13, wordWrap="CJK"))
    s.add(ParagraphStyle("TableHead",
        fontName="Helvetica-Bold", fontSize=9, textColor=_C_WHITE,
        alignment=TA_CENTER, leading=12))
    s.add(ParagraphStyle("TableCell",
        fontName="Helvetica", fontSize=8.5, textColor=_C_DARK,
        leading=12, wordWrap="CJK"))
    s.add(ParagraphStyle("TableCellRight",
        fontName="Helvetica", fontSize=8.5, textColor=_C_DARK,
        alignment=TA_RIGHT, leading=12, wordWrap="CJK"))
    s.add(ParagraphStyle("TableCellCenter",
        fontName="Helvetica", fontSize=8.5, textColor=_C_DARK,
        alignment=TA_CENTER, leading=12, wordWrap="CJK"))
    s.add(ParagraphStyle("StatusPaid",
        fontName="Helvetica-Bold", fontSize=8.5, textColor=_C_GREEN,
        alignment=TA_CENTER, leading=12))
    s.add(ParagraphStyle("StatusPartial",
        fontName="Helvetica-Bold", fontSize=8.5, textColor=_C_AMBER,
        alignment=TA_CENTER, leading=12))
    s.add(ParagraphStyle("StatusUnpaid",
        fontName="Helvetica-Bold", fontSize=8.5, textColor=_C_RED,
        alignment=TA_CENTER, leading=12))
    s.add(ParagraphStyle("CalDayHead",
        fontName="Helvetica-Bold", fontSize=8.5, textColor=_C_WHITE,
        alignment=TA_CENTER, leading=11))
    s.add(ParagraphStyle("CalDateNum",
        fontName="Helvetica-Bold", fontSize=9, textColor=_C_DARK,
        alignment=TA_LEFT, leading=11))
    s.add(ParagraphStyle("CalEventChip",
        fontName="Helvetica", fontSize=7, textColor=_C_DARK,
        alignment=TA_LEFT, leading=9))
    return s


def _status_style(status: str, styles):
    mapping = {
        "Paid":    styles["StatusPaid"],
        "Partial": styles["StatusPartial"],
        "Unpaid":  styles["StatusUnpaid"],
    }
    return mapping.get(status, styles["TableCell"])


def _header_block(story, styles, biz_name: str, title: str, period: str = "All Time"):
    header_left = []
    _lp = _logo_path()
    if os.path.exists(_lp):
        try:
            logo = Image(_lp, width=2*cm, height=2*cm)
            header_left.append(logo)
        except Exception:
            header_left.append(Paragraph(biz_name[:1], styles["Brand"]))
    else:
        header_left.append(Paragraph(biz_name[:1], styles["Brand"]))

    header_center = [
        Paragraph(biz_name, styles["Brand"]),
        Paragraph("Professional Catering Services", styles["BrandSub"]),
    ]

    now_str = datetime.now().strftime("%b %d, %Y  %I:%M %p")
    header_right = [
        Paragraph(f"<b>{title}</b>", styles["BrandSubRight"]),
        Paragraph(f"Period: {period}", styles["BrandSubRight"]),
        Paragraph(f"Generated: {now_str}", styles["BrandSubRight"]),
    ]

    tbl = Table([[header_left, header_center, header_right]],
                colWidths=[2.2*cm, 10.5*cm, 5.5*cm])
    tbl.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (2, 0), (2, 0),   "RIGHT"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(HRFlowable(width="100%", thickness=2.5, color=_C_RED, spaceAfter=12))


def _kpi_row(story, styles, kpis: dict):
    items = [
        ("Total Bookings",  str(kpis.get("total_bookings", 0))),
        ("Total Pax",       f"{int(kpis.get('total_pax', 0)):,}"),
        ("Total Revenue",   f"PHP {float(kpis.get('total_revenue', 0)):,.0f}"),
        ("Unpaid Amount",   f"PHP {float(kpis.get('unpaid_amount', 0)):,.0f}"),
    ]
    ncols = len(items)
    col_w = [_CONTENT_W / ncols] * ncols

    labels_row = [Paragraph(lbl, styles["KpiLabel"]) for lbl, _ in items]
    values_row = [Paragraph(val, styles["KpiValue"]) for _, val in items]

    t = Table([labels_row, values_row], colWidths=col_w)
    t.setStyle(TableStyle([
        ("BOX",           (0, 0), (-1, -1), 0.5, _C_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, _C_BORDER),
        ("BACKGROUND",    (0, 0), (-1, 0),  _C_LIGHT),
        ("BACKGROUND",    (0, 1), (-1, 1),  _C_WHITE),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))


def _bookings_table(story, styles, bookings: list):
    story.append(Paragraph("Booking Statistics", styles["SectionHead"]))

    headers = ["Booking Ref", "Client", "Event Date", "Pax", "Total Amount", "Status"]
    col_w   = [2.8*cm, 5.5*cm, 2.8*cm, 1.5*cm, 3.2*cm, 2.4*cm]

    header_row = [Paragraph(h, styles["TableHead"]) for h in headers]
    rows = [header_row]

    status_styles = {
        "CONFIRMED": styles["StatusPaid"],
        "PENDING":   styles["StatusPartial"],
        "CANCELLED": styles["StatusUnpaid"],
    }

    for b in bookings:
        st_key = b.get("status", "").upper()
        st_style = status_styles.get(st_key, styles["TableCell"])
        rows.append([
            Paragraph(b.get("id", ""), styles["TableCell"]),
            Paragraph(b.get("name", ""), styles["TableCell"]),
            Paragraph(b.get("date", ""), styles["TableCellCenter"]),
            Paragraph(str(b.get("pax", "")), styles["TableCellCenter"]),
            Paragraph(b.get("total", ""), styles["TableCellRight"]),
            Paragraph(b.get("status", "").capitalize(), st_style),
        ])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  _C_RED),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [_C_WHITE, _C_LIGHT]),
        ("BOX",           (0, 0), (-1, -1), 0.4, _C_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, _C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tbl)


def _footer(story, styles, biz_name: str = "Jayraldine's Catering"):
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_C_BORDER))
    story.append(Spacer(1, 5))
    story.append(Paragraph(
        f"{biz_name}  •  This report is system-generated and confidential.",
        styles["Footer"]
    ))


def build_analytics_sections() -> list:
    """Analytics data tables for exports: (title, headers, rows).
    Mirrors what the Reports page charts show."""
    import utils.repository as repo
    sections = []

    def _safe(fn):
        try:
            return fn() or []
        except Exception:
            return []

    rows = _safe(repo.get_profit_summary)
    if rows:
        sections.append((
            "Monthly Revenue, Expenses & Profit (This Year)",
            ["Month", "Revenue", "Expenses", "Net Profit"],
            [[r["month"], f"PHP {r['revenue']:,.0f}", f"PHP {r['expense']:,.0f}",
              f"PHP {r['profit']:,.0f}"] for r in rows]))

    rows = _safe(repo.get_payment_methods)
    if rows:
        sections.append(("Payment Methods", ["Method", "Bookings"],
                         [[r["method"], str(r["total"])] for r in rows]))

    rows = _safe(repo.get_top_menu_items)
    if rows:
        sections.append(("Top Menu Items", ["Item", "Orders"],
                         [[r["item"], str(r["count"])] for r in rows]))

    rows = _safe(lambda: repo.get_top_locations(limit=10))
    if rows:
        sections.append((
            "Top Event Locations", ["Location", "Bookings"],
            [[str(r.get("location") or r.get("name") or "?"),
              str(r.get("count") or r.get("total") or 0)] for r in rows]))

    rows = _safe(lambda: repo.get_top_occasions(limit=10))
    if rows:
        sections.append((
            "Bookings by Occasion", ["Occasion", "Bookings"],
            [[str(r.get("occasion") or r.get("name") or "?"),
              str(r.get("count") or r.get("total") or 0)] for r in rows]))

    rows = _safe(repo.get_customer_order_frequency)
    if rows:
        sections.append(("Customer Order Frequency", ["Customer", "Bookings"],
                         [[r["name"], str(r["count"])] for r in rows]))

    rows = _safe(lambda: repo.get_expense_breakdown(datetime.now().year))
    if rows:
        sections.append(("Expenses by Category (This Year)", ["Category", "Amount"],
                         [[r["category"], f"PHP {r['total']:,.0f}"] for r in rows]))

    rows = _safe(repo.get_yearly_summary)
    if rows:
        sections.append((
            "Year-over-Year Summary", ["Year", "Revenue", "Expenses", "Net Profit"],
            [[str(r["year"]), f"PHP {r['revenue']:,.0f}", f"PHP {r['expense']:,.0f}",
              f"PHP {r['profit']:,.0f}"] for r in rows]))

    return sections


def _section_table(story, styles, title: str, headers: list, rows: list):
    story.append(Paragraph(title, styles["SectionHead"]))
    data = [[Paragraph(h, styles["TableHead"]) for h in headers]]
    for r in rows:
        data.append([Paragraph(str(c), styles["TableCell"]) for c in r])
    tbl = Table(data, colWidths=[_CONTENT_W / len(headers)] * len(headers),
                repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _C_DARK),
        ("GRID", (0, 0), (-1, -1), 0.4, _C_BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_C_WHITE, _C_LIGHT]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 10))


def export_pdf(path: str, kpis: dict, bookings: list,
               title: str = "Business Report", period: str = "All Time",
               biz_name: str = "Jayraldine's Catering",
               sections: list = None, chart_images: list = None) -> bool:
    """sections: [(title, headers, rows)] — analytics tables.
    chart_images: [(title, png_path)] — chart screenshots from the live page."""
    if not REPORTLAB_OK:
        return False
    try:
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=_MARGIN, rightMargin=_MARGIN,
            topMargin=_MARGIN, bottomMargin=_MARGIN,
            title=f"{biz_name} — {title}",
            author=biz_name,
        )
        styles = _styles()
        story = []
        _header_block(story, styles, biz_name, title, period)
        story.append(Paragraph("Key Performance Indicators", styles["SectionHead"]))
        _kpi_row(story, styles, kpis)

        # Charts, as images grabbed from the live Reports page
        for chart_title, png in (chart_images or []):
            try:
                from reportlab.lib.utils import ImageReader
                iw, ih = ImageReader(png).getSize()
                w = _CONTENT_W
                h = ih * (w / iw)
                if h > 9 * cm:
                    h = 9 * cm
                    w = iw * (h / ih)
                story.append(KeepTogether([
                    Paragraph(chart_title, styles["SectionHead"]),
                    Image(png, width=w, height=h),
                    Spacer(1, 8),
                ]))
            except Exception as exc:
                print(f"[exporter] chart image skipped ({chart_title}): {exc}")

        _bookings_table(story, styles, bookings)

        for sec_title, headers, rows in (sections or []):
            _section_table(story, styles, sec_title, headers, rows)

        _footer(story, styles, biz_name)
        doc.build(story)
        return True
    except Exception as exc:
        print(f"[exporter] PDF failed: {exc}")
        return False


def export_tablet_master_data(save_path: str) -> dict:
    """PC -> Tablet master data transfer.
    Writes a standalone SQLite file containing packages, package_items, menu_items,
    customers directory, and address lookup tables - strictly without past bookings
    or transaction records - so the Tablet App gets the complete customer list,
    menu catalog, and address dropdown data."""
    import utils.db as db
    stats = {"packages": 0, "menu_items": 0, "package_items": 0, "customers": 0, "addresses": 0, "errors": []}
    try:
        db.execute("ATTACH DATABASE ? AS dst", (save_path,))
    except Exception as exc:
        stats["errors"].append(f"Could not create export file: {exc}")
        return stats
    try:
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.packages (
                pkg_id INTEGER PRIMARY KEY, pkg_name TEXT NOT NULL UNIQUE, pkg_description TEXT,
                pkg_price_per_pax REAL NOT NULL, pkg_min_pax INTEGER DEFAULT 30,
                pkg_created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.menu_items (
                mi_id INTEGER PRIMARY KEY, mi_name TEXT NOT NULL, mi_category TEXT NOT NULL,
                mi_price REAL NOT NULL, mi_status TEXT DEFAULT 'Available', mi_description TEXT,
                mi_created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.package_items (
                pi_id INTEGER PRIMARY KEY, pi_package_id INTEGER, pi_menu_item_id INTEGER,
                pi_item_name TEXT, pi_category TEXT, pi_custom_price REAL DEFAULT 0.0, pi_quantity INTEGER DEFAULT 1
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.customers (
                cus_id INTEGER PRIMARY KEY, cus_name TEXT NOT NULL, cus_contact TEXT,
                cus_email TEXT, cus_address TEXT, cus_address_id INTEGER,
                cus_loyalty_tier TEXT DEFAULT 'Bronze', cus_total_events INTEGER DEFAULT 0,
                cus_total_spent REAL DEFAULT 0.0, cus_status TEXT DEFAULT 'Active',
                cus_notes TEXT, cus_created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.address_provinces (
                ap_id INTEGER PRIMARY KEY AUTOINCREMENT, ap_name TEXT NOT NULL UNIQUE
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.address_cities (
                ac_id INTEGER PRIMARY KEY AUTOINCREMENT, ac_province_id INTEGER, ac_name TEXT NOT NULL
            )
        """)
        db.execute("""
            CREATE TABLE IF NOT EXISTS dst.address_barangays (
                ab_id INTEGER PRIMARY KEY AUTOINCREMENT, ab_city_id INTEGER, ab_name TEXT NOT NULL
            )
        """)

        db.execute("DELETE FROM dst.packages")
        db.execute("DELETE FROM dst.menu_items")
        db.execute("DELETE FROM dst.package_items")
        db.execute("DELETE FROM dst.customers")
        db.execute("DELETE FROM dst.address_barangays")
        db.execute("DELETE FROM dst.address_cities")
        db.execute("DELETE FROM dst.address_provinces")

        db.execute("INSERT INTO dst.packages SELECT pkg_id, pkg_name, pkg_description, pkg_price_per_pax, pkg_min_pax, pkg_created_at FROM main.packages")
        db.execute("INSERT INTO dst.menu_items (mi_id, mi_name, mi_category, mi_price, mi_status, mi_description, mi_created_at) "
                   "SELECT mi_id, mi_name, mi_category, mi_price, mi_status, mi_description, mi_created_at FROM main.menu_items")
        db.execute("INSERT INTO dst.package_items SELECT pi_id, pi_package_id, pi_menu_item_id, pi_item_name, pi_category, pi_custom_price, pi_quantity FROM main.package_items")

        # Copy customer profiles without orders
        db.execute("""
            INSERT INTO dst.customers (cus_id, cus_name, cus_contact, cus_email, cus_address, cus_address_id, cus_loyalty_tier, cus_total_events, cus_total_spent, cus_status, cus_notes, cus_created_at)
            SELECT cus_id, cus_name, cus_contact, cus_email, cus_address, cus_address_id, cus_loyalty_tier, cus_total_events, cus_total_spent, cus_status, cus_notes, cus_created_at
            FROM main.customers WHERE cus_status = 'Active' OR cus_status IS NULL
        """)

        # Copy address lookup hierarchy
        db.execute("INSERT INTO dst.address_provinces SELECT ap_id, ap_name FROM main.address_provinces")
        db.execute("INSERT INTO dst.address_cities SELECT ac_id, ac_province_id, ac_name FROM main.address_cities")
        db.execute("INSERT INTO dst.address_barangays SELECT ab_id, ab_city_id, ab_name FROM main.address_barangays")

        stats["packages"] = db.fetchone("SELECT COUNT(*) AS c FROM dst.packages")["c"]
        stats["menu_items"] = db.fetchone("SELECT COUNT(*) AS c FROM dst.menu_items")["c"]
        stats["package_items"] = db.fetchone("SELECT COUNT(*) AS c FROM dst.package_items")["c"]
        stats["customers"] = db.fetchone("SELECT COUNT(*) AS c FROM dst.customers")["c"]
        stats["addresses"] = db.fetchone("SELECT COUNT(*) AS c FROM dst.address_barangays")["c"]
    except Exception as exc:
        stats["errors"].append(str(exc))
    finally:
        try:
            db.execute("DETACH DATABASE dst")
        except Exception:
            pass
    return stats


def export_daily_activity_report_pdf(path: str, entries: list, business: dict,
                                     period_label: str = "Today") -> bool:
    """PDF version of the Daily Activity / Audit Report: who did what, to
    which customer/order, for how much, and when - so the owner can review
    every action taken during the covered period."""
    if not REPORTLAB_OK:
        return False
    try:
        biz_name = business.get("name", "Jayraldine's Catering")
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=_MARGIN, rightMargin=_MARGIN,
            topMargin=_MARGIN, bottomMargin=_MARGIN,
            title=f"{biz_name} — Daily Activity Report",
            author=biz_name,
        )
        styles = _styles()
        story = []
        _header_block(story, styles, biz_name, "Daily Activity Report", period_label)

        if entries:
            headers = ["Date", "Time", "User", "Action", "Details"]
            rows = []
            for e in entries:
                rows.append([
                    e.get("date", ""), e.get("time", ""), e.get("actor", ""),
                    e.get("action", ""), e.get("description", ""),
                ])
            _section_table(story, styles, f"Activity Log ({len(entries)} action{'s' if len(entries) != 1 else ''})", headers, rows)
        else:
            story.append(Paragraph("No activity recorded for this period.", styles["TableCell"]))

        _footer(story, styles, biz_name)
        doc.build(story)
        return True
    except Exception as exc:
        print(f"[exporter] Daily Activity Report PDF failed: {exc}")
        return False


def export_receipt_pdf(path: str, inv: dict, business: dict,
                       additional_charges: list = None,
                       payment_records: list = None,
                       down_payment: float = None) -> bool:
    """Generate a professional PDF receipt for a single invoice.

    additional_charges: list of {"description", "amount", "date_added"} - positive
        amounts are extra charges, negative amounts are discounts.
    payment_records: full payment history (down payment + subsequent payments),
        each {"amount", "payment_date", "method", "note"} - each payment keeps its
        own recorded date so the receipt shows exactly when each amount was paid.
    down_payment: the original down payment amount, shown as its own line.
    """
    if not REPORTLAB_OK:
        return False
    try:
        additional_charges = additional_charges or []
        payment_records = payment_records or []
        charges = [c for c in additional_charges if float(c.get("amount", 0)) > 0]
        discounts = [c for c in additional_charges if float(c.get("amount", 0)) < 0]
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=_MARGIN, rightMargin=_MARGIN,
            topMargin=_MARGIN, bottomMargin=_MARGIN,
            title=f"Receipt — {inv.get('invoice', '')}",
        )
        styles = _styles()
        story = []

        biz_name    = business.get("name", "Jayraldine's Catering")
        biz_address = business.get("address", "")
        biz_contact = business.get("contact", "")
        biz_email   = business.get("email", "")

        total   = float(inv.get("amount", 0))
        paid    = float(inv.get("paid", 0))
        balance = total - paid
        status  = inv.get("status", "Unpaid")

        status_color = {"Paid": _C_GREEN, "Partial": _C_AMBER, "Unpaid": _C_RED}.get(status, _C_MUTED)

        header_w = _CONTENT_W

        logo_cell = ""
        _lp = _logo_path()
        if os.path.exists(_lp):
            try:
                logo_cell = Image(_lp, width=2.0*cm, height=2.0*cm)
            except Exception:
                logo_cell = ""

        biz_cell = [
            Paragraph(biz_name, styles["Brand"]),
            Paragraph("Professional Catering Services", styles["BrandSub"]),
            Spacer(1, 2),
            Paragraph(biz_address, styles["BrandSub"]),
            Paragraph(f"Tel: {biz_contact}  ·  {biz_email}", styles["BrandSub"]),
        ]

        receipt_cell = [
            Paragraph("OFFICIAL RECEIPT", ParagraphStyle(
                "rcpt_lbl", fontName="Helvetica-Bold", fontSize=8.5,
                textColor=_C_MUTED, alignment=TA_RIGHT, leading=11,
                spaceAfter=4)),
            Paragraph(inv.get("invoice", "—"), ParagraphStyle(
                "rcpt_no", fontName="Helvetica-Bold", fontSize=18,
                textColor=_C_RED, alignment=TA_RIGHT, leading=22)),
        ]

        hdr_cols = [2.2*cm, 10*cm, 6*cm] if logo_cell else [12.2*cm, 6*cm]
        hdr_data = [[logo_cell, biz_cell, receipt_cell]] if logo_cell else [[biz_cell, receipt_cell]]

        hdr_tbl = Table(hdr_data, colWidths=hdr_cols)
        hdr_tbl.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("ALIGN",         (-1, 0), (-1, 0), "RIGHT"),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=2.5, color=_C_RED, spaceAfter=0.4*cm))

        def _det_row(label, value, alt=False, value_style=None):
            bg = colors.HexColor("#F8FAFC") if alt else _C_WHITE
            vs = value_style or styles["DetailValue"]
            return [
                Paragraph(label, styles["DetailLabel"]),
                Paragraph(str(value), vs),
            ], bg

        detail_rows_data = [
            ("Receipt #",    inv.get("invoice", "—"),                   False, None),
            ("Customer",     inv.get("customer", "—"),                   True,  None),
            ("Event Date",   inv.get("event_date", "—"),                 False, None),
        ]

        detail_rows = []
        tbl_style_cmds = [
            ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("TOPPADDING",   (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 7),
            ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]

        for i, (lbl, val, alt, vstyle) in enumerate(detail_rows_data):
            row_data, bg = _det_row(lbl, val, alt, vstyle)
            detail_rows.append(row_data)
            if alt:
                tbl_style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAFC")))

        det_tbl = Table(detail_rows, colWidths=[4.5*cm, _CONTENT_W - 4.5*cm])
        det_tbl.setStyle(TableStyle(tbl_style_cmds))
        story.append(det_tbl)
        story.append(Spacer(1, 0.3*cm))

        # Additional Charges / Additional Items breakdown - each stays a
        # separate, explained line item, never silently merged into the total.
        if charges:
            story.append(Paragraph("Additional Charges / Items", styles["DetailLabel"]))
            story.append(Spacer(1, 4))
            rows = [[Paragraph("Description", styles["DetailLabel"]), Paragraph("Amount", styles["DetailLabel"])]]
            for c in charges:
                rows.append([
                    Paragraph(str(c.get("description", "")), styles["DetailValue"]),
                    Paragraph(f"PHP {float(c['amount']):,.2f}", ParagraphStyle(
                        "chg_amt", fontName="Helvetica", fontSize=9.5,
                        textColor=_C_DARK, alignment=TA_RIGHT, leading=12)),
                ])
            chg_tbl = Table(rows, colWidths=[_CONTENT_W * 0.7, _CONTENT_W * 0.3])
            chg_tbl.setStyle(TableStyle([
                ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
                ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
                ("BACKGROUND",   (0, 0), (-1, 0),  _C_LIGHT),
                ("TOPPADDING",   (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
                ("LEFTPADDING",  (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]))
            story.append(chg_tbl)
            story.append(Spacer(1, 0.3*cm))

        if discounts:
            story.append(Paragraph("Discounts", styles["DetailLabel"]))
            story.append(Spacer(1, 4))
            rows = [[Paragraph("Description", styles["DetailLabel"]), Paragraph("Amount", styles["DetailLabel"])]]
            for c in discounts:
                rows.append([
                    Paragraph(str(c.get("description", "")), styles["DetailValue"]),
                    Paragraph(f"- PHP {abs(float(c['amount'])):,.2f}", ParagraphStyle(
                        "disc_amt", fontName="Helvetica", fontSize=9.5,
                        textColor=_C_GREEN, alignment=TA_RIGHT, leading=12)),
                ])
            disc_tbl = Table(rows, colWidths=[_CONTENT_W * 0.7, _CONTENT_W * 0.3])
            disc_tbl.setStyle(TableStyle([
                ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
                ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
                ("BACKGROUND",   (0, 0), (-1, 0),  _C_LIGHT),
                ("TOPPADDING",   (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
                ("LEFTPADDING",  (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]))
            story.append(disc_tbl)
            story.append(Spacer(1, 0.3*cm))

        amount_rows = [
            [Paragraph("Total Amount", styles["DetailLabel"]),
             Paragraph(f"PHP {total:,.2f}", ParagraphStyle(
                 "amt", fontName="Helvetica", fontSize=10,
                 textColor=_C_DARK, alignment=TA_RIGHT, leading=13))],
        ]
        if down_payment:
            amount_rows.append([
                Paragraph("Down Payment", styles["DetailLabel"]),
                Paragraph(f"PHP {float(down_payment):,.2f}", ParagraphStyle(
                    "dp", fontName="Helvetica", fontSize=10,
                    textColor=_C_GREEN, alignment=TA_RIGHT, leading=13)),
            ])
        amount_rows.append([
            Paragraph("Total Amount Paid", styles["DetailLabel"]),
            Paragraph(f"PHP {paid:,.2f}", ParagraphStyle(
                "paid", fontName="Helvetica", fontSize=10,
                textColor=_C_GREEN, alignment=TA_RIGHT, leading=13)),
        ])
        amount_rows.append([
            Paragraph("Balance Due", ParagraphStyle(
                 "bal_lbl", fontName="Helvetica-Bold", fontSize=10,
                 textColor=_C_DARK, leading=13)),
             Paragraph(f"PHP {balance:,.2f}", ParagraphStyle(
                 "bal_val", fontName="Helvetica-Bold", fontSize=12,
                 textColor=_C_RED if balance > 0 else _C_GREEN,
                 alignment=TA_RIGHT, leading=15)),
        ])

        amt_style = [
            ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("BACKGROUND",   (0, 0), (-1, 0),  _C_LIGHT),
            ("BACKGROUND",   (0, len(amount_rows) - 1), (-1, len(amount_rows) - 1),
                colors.HexColor("#FFF1F2") if balance > 0 else colors.HexColor("#F0FDF4")),
            ("TOPPADDING",   (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 9),
            ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]

        amt_tbl = Table(amount_rows, colWidths=[_CONTENT_W * 0.55, _CONTENT_W * 0.45])
        amt_tbl.setStyle(TableStyle(amt_style))
        story.append(amt_tbl)
        story.append(Spacer(1, 0.3*cm))

        # Payment History - every payment keeps its own recorded date.
        if payment_records:
            story.append(Paragraph("Payment History", styles["DetailLabel"]))
            story.append(Spacer(1, 4))
            rows = [[
                Paragraph("Date", styles["DetailLabel"]),
                Paragraph("Method", styles["DetailLabel"]),
                Paragraph("Amount", styles["DetailLabel"]),
            ]]
            for p in payment_records:
                rows.append([
                    Paragraph(str(p.get("payment_date", "")), styles["DetailValue"]),
                    Paragraph(str(p.get("method", "")), styles["DetailValue"]),
                    Paragraph(f"PHP {float(p.get('amount', 0)):,.2f}", ParagraphStyle(
                        "pr_amt", fontName="Helvetica", fontSize=9.5,
                        textColor=_C_GREEN, alignment=TA_RIGHT, leading=12)),
                ])
            pr_tbl = Table(rows, colWidths=[_CONTENT_W * 0.3, _CONTENT_W * 0.35, _CONTENT_W * 0.35])
            pr_tbl.setStyle(TableStyle([
                ("BOX",          (0, 0), (-1, -1), 0.4, _C_BORDER),
                ("INNERGRID",    (0, 0), (-1, -1), 0.3, _C_BORDER),
                ("BACKGROUND",   (0, 0), (-1, 0),  _C_LIGHT),
                ("TOPPADDING",   (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
                ("LEFTPADDING",  (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]))
            story.append(pr_tbl)
            story.append(Spacer(1, 0.3*cm))

        status_row = Table(
            [[Paragraph("Payment Status:", ParagraphStyle(
                "st_lbl", fontName="Helvetica-Bold", fontSize=9,
                textColor=_C_GRAY, leading=12)),
              Paragraph(status.upper(), ParagraphStyle(
                "st_val", fontName="Helvetica-Bold", fontSize=11,
                textColor=status_color, alignment=TA_RIGHT, leading=14))]],
            colWidths=[_CONTENT_W * 0.5, _CONTENT_W * 0.5],
        )
        status_row.setStyle(TableStyle([
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("BOX",           (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("BACKGROUND",    (0, 0), (-1, 0),  _C_LIGHT),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(status_row)

        story.append(Spacer(1, 0.6*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=_C_BORDER))
        story.append(Spacer(1, 0.3*cm))

        from datetime import datetime as _dt
        printed_str = _dt.now().strftime("%B %d, %Y at %I:%M %p")
        story.append(Paragraph(
            f"Printed: {printed_str}  ·  Thank you for choosing <b>{biz_name}</b>!",
            styles["Footer"],
        ))
        if balance > 0:
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                f"Please settle the remaining balance of PHP {balance:,.2f} before your event date.",
                ParagraphStyle("bal_note", fontName="Helvetica", fontSize=7.5,
                    textColor=_C_RED, alignment=TA_CENTER, leading=10)
            ))

        doc.build(story)
        return True
    except Exception as exc:
        print(f"[exporter] Receipt PDF failed: {exc}")
        return False


def export_excel(path: str, kpis: dict, bookings: list,
                 title: str = "Business Report", period: str = "All Time",
                 biz_name: str = "Jayraldine's Catering",
                 sections: list = None) -> bool:
    if not OPENPYXL_OK:
        return False
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        RED   = "E11D48"
        DARK  = "0B1220"
        GRAY  = "374151"
        LIGHT = "F9FAFB"
        WHITE = "FFFFFF"

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _border():
            s = Side(style="thin", color="E5E7EB")
            return Border(left=s, right=s, top=s, bottom=s)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 14

        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = biz_name.upper()
        c.font = Font(name="Calibri", bold=True, size=18, color=RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = f"{title}  |  Period: {period}  |  Generated: {datetime.now().strftime('%b %d, %Y %I:%M %p')}"
        c.font = Font(name="Calibri", size=9, color=GRAY)
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 16
        row += 2

        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "KEY PERFORMANCE INDICATORS"
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = _fill(DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        kpi_pairs = [
            ("Total Bookings",  kpis.get("total_bookings", 0)),
            ("Total Pax",       kpis.get("total_pax", 0)),
            ("Total Revenue",   f"PHP {float(kpis.get('total_revenue', 0)):,.0f}"),
            ("Unpaid Amount",   f"PHP {float(kpis.get('unpaid_amount', 0)):,.0f}"),
            ("Today's Bookings", kpis.get("today_bookings", 0)),
            ("This Week's Bookings", kpis.get("week_bookings", 0)),
        ]
        for label, value in kpi_pairs:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(name="Calibri", bold=True, size=10, color=DARK)
            lc.fill = _fill(LIGHT)
            lc.border = _border()
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = Font(name="Calibri", size=10, color=DARK)
            vc.fill = _fill(WHITE)
            vc.border = _border()
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = "BOOKING STATISTICS"
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = _fill(RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        headers = ["Booking Ref", "Client", "Event Date", "Pax", "Total Amount", "Status"]
        for col, hdr in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
            c.fill = _fill(GRAY)
            c.border = _border()
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for i, b in enumerate(bookings):
            bg = LIGHT if i % 2 == 0 else WHITE
            vals = [
                b.get("id", ""), b.get("name", ""), b.get("date", ""),
                b.get("pax", ""), b.get("total", ""), b.get("status", "").capitalize()
            ]
            aligns = ["left","left","center","center","right","center"]
            for col, (val, aln) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Calibri", size=9)
                c.fill = _fill(bg)
                c.border = _border()
                c.alignment = Alignment(horizontal=aln, vertical="center", indent=1 if aln=="left" else 0)
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = f"{biz_name}  •  System-generated report  •  Confidential"
        c.font = Font(name="Calibri", size=8, italic=True, color=GRAY)
        c.alignment = Alignment(horizontal="center")

        # Analytics sections — one worksheet per section
        for sec_title, sec_headers, sec_rows in (sections or []):
            sheet_name = sec_title[:28].replace("/", "-")
            ws2 = wb.create_sheet(sheet_name)
            for col in range(1, len(sec_headers) + 1):
                ws2.column_dimensions[get_column_letter(col)].width = 24
            r2 = 1
            ws2.merge_cells(start_row=r2, start_column=1,
                            end_row=r2, end_column=len(sec_headers))
            tc = ws2.cell(row=r2, column=1, value=sec_title.upper())
            tc.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
            tc.fill = _fill(DARK)
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[r2].height = 20
            r2 += 1
            for col, hdr in enumerate(sec_headers, 1):
                hc = ws2.cell(row=r2, column=col, value=hdr)
                hc.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
                hc.fill = _fill(GRAY)
                hc.border = _border()
                hc.alignment = Alignment(horizontal="center", vertical="center")
            r2 += 1
            for i, sec_row in enumerate(sec_rows):
                bg = LIGHT if i % 2 == 0 else WHITE
                for col, val in enumerate(sec_row, 1):
                    cc = ws2.cell(row=r2, column=col, value=val)
                    cc.font = Font(name="Calibri", size=9)
                    cc.fill = _fill(bg)
                    cc.border = _border()
                    cc.alignment = Alignment(
                        horizontal="left" if col == 1 else "right",
                        vertical="center", indent=1)
                r2 += 1

        wb.save(path)
        return True
    except Exception as exc:
        print(f"[exporter] Excel failed: {exc}")
        return False


def _parse_amount(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("₱", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _record_in_month(rec: dict, date_key: str, year: int, month: int) -> bool:
    """True if rec[date_key] falls within the given (year, month). Handles
    date/datetime objects and common string formats used across the app."""
    raw = rec.get(date_key)
    if not raw:
        return False
    if isinstance(raw, (_dt_date, _dt_datetime)):
        return raw.year == year and raw.month == month
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%b %d, %Y", "%B %d, %Y", "%m/%d/%Y"):
        try:
            d = _dt_datetime.strptime(s, fmt)
            return d.year == year and d.month == month
        except ValueError:
            continue
    return s.startswith(f"{year:04d}-{month:02d}")


def _filter_by_month(records: list, date_key: str, year: int, month: int) -> list:
    if not year or not month:
        return records
    return [r for r in records if _record_in_month(r, date_key, year, month)]


def export_custom_entity_data(entity_name: str, is_excel: bool, save_path: str,
                              year: int = None, month: int = None) -> bool:
    """Export Bookings, Customers, Expenses, Menu, Billings, or Master Data to Excel or CSV.

    When year and month are given, only records whose relevant date column falls
    in that month are included (bookings/invoices by event date, expenses/cash
    flow by their own date). Customers and Menu Items have no date to filter by
    and are always exported in full.
    """
    import utils.repository as repo
    import csv

    headers = []
    rows = []
    sheets = {}

    if "Bookings" in entity_name:
        headers = ["Booking Ref", "Customer Name", "Contact Number", "Email Address", "Event Date", "Event Time", "Venue / Location", "Occasion", "Guest Count (Pax)", "Total Amount (₱)", "Down Paid (₱)", "Balance (₱)", "Status", "Payment Mode", "Notes / Theme"]
        b_list = _filter_by_month(repo.get_all_bookings_for_export() or [], "event_date", year, month)
        for b in b_list:
            try:
                tot = _parse_amount(b.get('total') or b.get('total_amount') or 0)
                paid = _parse_amount(b.get('amount_paid') or b.get('down_payment') or 0)
                bal = max(0.0, tot - paid)
                rows.append([
                    b.get("id") or b.get("booking_ref") or "",
                    b.get("name") or b.get("customer_name") or "",
                    b.get("contact") or b.get("phone") or "",
                    b.get("email") or "",
                    b.get("date") or b.get("event_date") or "",
                    b.get("time") or b.get("event_time") or "",
                    b.get("venue") or "",
                    b.get("occasion") or "",
                    str(b.get("pax") or 0),
                    f"₱{tot:,.2f}",
                    f"₱{paid:,.2f}",
                    f"₱{bal:,.2f}",
                    str(b.get("status") or "PENDING").upper(),
                    b.get("payment_mode") or "Cash",
                    b.get("notes") or ""
                ])
            except Exception as row_exc:
                print(f"[exporter] Bookings row error (skipping): {row_exc}")
        sheets["Bookings"] = (headers, rows)

    elif "Customers" in entity_name:
        headers = ["Customer ID", "Customer Name", "Contact Number", "Email Address", "Address", "Notes / History"]
        c_list = repo.get_all_customers() or []
        for c in c_list:
            rows.append([
                c.get("id", ""), c.get("name", ""), c.get("contact", ""),
                c.get("email", ""), c.get("address", "") or c.get("imported_address", ""),
                c.get("notes", "")
            ])
        sheets["Customers"] = (headers, rows)

    elif "Expenses" in entity_name:
        headers = ["Expense ID", "Expense Date", "Category", "Description", "Amount (₱)"]
        e_list = _filter_by_month(repo.get_all_expenses() or [], "date", year, month)
        for e in e_list:
            rows.append([
                e.get("id", ""), e.get("date", ""), e.get("category", ""),
                e.get("description", ""), f"₱{_parse_amount(e.get('amount', 0)):,.2f}"
            ])
        sheets["Expenses"] = (headers, rows)

    elif "Menu" in entity_name:
        headers = ["Item ID", "Item / Package Name", "Category", "Price / Rate (₱)", "Description / Inclusions"]
        m_list = repo.get_all_menu_items() or []
        for m in m_list:
            rows.append([
                m.get("id", ""), m.get("name", "") or m.get("item", ""), m.get("category", ""),
                f"₱{_parse_amount(m.get('price', 0)):,.2f}", m.get("description", "")
            ])
        sheets["Menu Items"] = (headers, rows)

    elif "Billing" in entity_name:
        headers = ["Invoice Ref", "Booking Ref", "Customer Name", "Event Date", "Total Amount (₱)", "Paid Amount (₱)", "Balance Due (₱)", "Payment Status"]
        i_list = _filter_by_month(repo.get_all_invoices() or [], "event_date", year, month)
        for inv in i_list:
            rows.append([
                inv.get("invoice", ""), inv.get("booking_ref", ""), inv.get("customer", ""),
                inv.get("event_date", ""), f"₱{_parse_amount(inv.get('amount', 0)):,.2f}",
                f"₱{_parse_amount(inv.get('paid', 0)):,.2f}", f"₱{_parse_amount(inv.get('balance', 0)):,.2f}",
                inv.get("status", "")
            ])
        sheets["Invoices"] = (headers, rows)

    elif "Cash" in entity_name or "Flow" in entity_name:
        headers = ["Date", "Check #", "Particulars (Account / Detail)", "Deposit (₱)", "Withdrawal (₱)", "Running Balance (₱)", "Actual Sales (₱)", "Variance / Difference (₱)", "Remarks / Notes"]
        tx_list = _filter_by_month(repo.get_cash_flow_transactions() or [], "date", year, month)
        for tx in tx_list:
            dep = float(tx.get("deposit") or 0.0)
            withd = float(tx.get("withdrawal") or 0.0)
            bal = float(tx.get("balance") or 0.0)
            act_sales = float(tx.get("actual_sales") or 0.0)
            diff = bal - act_sales
            diff_str = f"₱{diff:,.2f}" if diff >= 0 else f"(₱{abs(diff):,.2f})"
            rows.append([
                str(tx.get("date", "")),
                str(tx.get("check_no", "")),
                str(tx.get("particulars", "")),
                f"₱{dep:,.2f}" if dep > 0 else "—",
                f"₱{withd:,.2f}" if withd > 0 else "—",
                f"₱{bal:,.2f}" if bal >= 0 else f"(₱{abs(bal):,.2f})",
                f"₱{act_sales:,.2f}" if act_sales > 0 else "—",
                diff_str if act_sales > 0 else "—",
                str(tx.get("notes", ""))
            ])
        sheets["Cash Flow Ledger"] = (headers, rows)

    else: # Master Export
        b_hdrs = ["Booking Ref", "Customer Name", "Contact Number", "Email Address", "Event Date", "Event Time", "Venue", "Occasion", "Pax", "Total Amount (₱)", "Down Paid (₱)", "Balance (₱)", "Status", "Payment Mode", "Notes / Theme"]
        b_rows = []
        for b in _filter_by_month(repo.get_all_bookings_for_export() or [], "event_date", year, month):
            try:
                tot = _parse_amount(b.get('total') or b.get('total_amount') or 0)
                paid = _parse_amount(b.get('amount_paid') or b.get('down_payment') or 0)
                bal = max(0.0, tot - paid)
                b_rows.append([
                    b.get("id") or b.get("booking_ref") or "",
                    b.get("name") or b.get("customer_name") or "",
                    b.get("contact") or b.get("phone") or "",
                    b.get("email") or "",
                    b.get("date") or b.get("event_date") or "",
                    b.get("time") or b.get("event_time") or "",
                    b.get("venue") or "",
                    b.get("occasion") or "",
                    str(b.get("pax") or 0),
                    f"₱{tot:,.2f}",
                    f"₱{paid:,.2f}",
                    f"₱{bal:,.2f}",
                    str(b.get("status") or "PENDING").upper(),
                    b.get("payment_mode") or "Cash",
                    b.get("notes") or ""
                ])
            except Exception as row_exc:
                print(f"[exporter] Master Bookings row error (skipping): {row_exc}")
        sheets["Bookings"] = (b_hdrs, b_rows)

        c_hdrs = ["Customer ID", "Customer Name", "Contact Number", "Email Address", "Address", "Notes / History"]
        c_rows = [[c.get("id", ""), c.get("name", ""), c.get("contact", ""), c.get("email", ""), c.get("address", "") or c.get("imported_address", ""), c.get("notes", "")] for c in (repo.get_all_customers() or [])]
        sheets["Customers"] = (c_hdrs, c_rows)

        e_hdrs = ["Expense ID", "Expense Date", "Category", "Description", "Amount (₱)"]
        e_rows = [[e.get("id", ""), e.get("date", ""), e.get("category", ""), e.get("description", ""), f"₱{_parse_amount(e.get('amount', 0)):,.2f}"] for e in _filter_by_month(repo.get_all_expenses() or [], "date", year, month)]
        sheets["Expenses"] = (e_hdrs, e_rows)

        m_hdrs = ["Item ID", "Item / Package Name", "Category", "Price / Rate (₱)", "Description / Inclusions"]
        m_rows = [[m.get("id", ""), m.get("name", "") or m.get("item", ""), m.get("category", ""), f"₱{_parse_amount(m.get('price', 0)):,.2f}", m.get("description", "")] for m in (repo.get_all_menu_items() or [])]
        sheets["Menu Items"] = (m_hdrs, m_rows)

        cf_hdrs = ["Date", "Check #", "Particulars (Account / Detail)", "Deposit (₱)", "Withdrawal (₱)", "Running Balance (₱)", "Actual Sales (₱)", "Variance / Difference (₱)", "Remarks / Notes"]
        cf_rows = [[
            str(tx.get("date", "")),
            str(tx.get("check_no", "")),
            str(tx.get("particulars", "")),
            f"₱{float(tx.get('deposit') or 0.0):,.2f}" if float(tx.get('deposit') or 0.0) > 0 else "—",
            f"₱{float(tx.get('withdrawal') or 0.0):,.2f}" if float(tx.get('withdrawal') or 0.0) > 0 else "—",
            f"₱{float(tx.get('balance') or 0.0):,.2f}" if float(tx.get('balance') or 0.0) >= 0 else f"(₱{abs(float(tx.get('balance') or 0.0)):,.2f})",
            f"₱{float(tx.get('actual_sales') or 0.0):,.2f}" if float(tx.get('actual_sales') or 0.0) > 0 else "—",
            f"₱{(float(tx.get('balance') or 0.0) - float(tx.get('actual_sales') or 0.0)):,.2f}" if float(tx.get('actual_sales') or 0.0) > 0 and (float(tx.get('balance') or 0.0) - float(tx.get('actual_sales') or 0.0)) >= 0 else (f"(₱{abs(float(tx.get('balance') or 0.0) - float(tx.get('actual_sales') or 0.0)):,.2f})" if float(tx.get('actual_sales') or 0.0) > 0 else "—"),
            str(tx.get("notes", ""))
        ] for tx in _filter_by_month(repo.get_cash_flow_transactions() or [], "date", year, month)]
        sheets["Cash Flow Ledger"] = (cf_hdrs, cf_rows)

        i_hdrs = ["Invoice Ref", "Booking Ref", "Customer Name", "Event Date", "Total Amount (₱)", "Paid Amount (₱)", "Balance Due (₱)", "Payment Status"]
        i_rows = [[inv.get("invoice", ""), inv.get("booking_ref", ""), inv.get("customer", ""), inv.get("event_date", ""), f"₱{_parse_amount(inv.get('amount', 0)):,.2f}", f"₱{_parse_amount(inv.get('paid', 0)):,.2f}", f"₱{_parse_amount(inv.get('balance', 0)):,.2f}", inv.get("status", "")] for inv in _filter_by_month(repo.get_all_invoices() or [], "event_date", year, month)]
        sheets["Invoices & Payments"] = (i_hdrs, i_rows)

    if not is_excel:
        try:
            with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                first_sheet = next(iter(sheets.values()))
                writer.writerow(first_sheet[0])
                writer.writerows(first_sheet[1])
            return True
        except Exception as exc:
            print(f"[exporter] CSV export failed: {exc}")
            return False

    if not OPENPYXL_OK:
        return False

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        RED   = "E11D48"
        DARK  = "0B1220"
        LIGHT = "F8FAFC"
        WHITE = "FFFFFF"

        for sheet_title, (hdrs, data_rows) in sheets.items():
            ws = wb.create_sheet(title=sheet_title)

            ws.row_dimensions[1].height = 24
            for col_idx, h_text in enumerate(hdrs, 1):
                cell = ws.cell(row=1, column=col_idx, value=h_text)
                cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=RED if "Booking" in sheet_title or "Invoice" in sheet_title else DARK)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                s = Side(style="thin", color="E5E7EB")
                cell.border = Border(left=s, right=s, top=s, bottom=s)

            for row_idx, r_data in enumerate(data_rows, 2):
                ws.row_dimensions[row_idx].height = 18
                bg = LIGHT if row_idx % 2 == 0 else WHITE
                for col_idx, val in enumerate(r_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = Font(name="Calibri", size=9)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    s = Side(style="thin", color="E5E7EB")
                    cell.border = Border(left=s, right=s, top=s, bottom=s)
                    align_right = ("₱" in str(val) or "Amount" in hdrs[col_idx-1] or "Paid" in hdrs[col_idx-1])
                    cell.alignment = Alignment(horizontal="right" if align_right else "left", vertical="center")

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(save_path)
        return True
    except Exception as exc:
        print(f"[exporter] Excel export failed: {exc}")
        return False


def _format_time_short(t_raw) -> str:
    """Format time string compactly for calendar day cells: e.g. 5:00 PM -> 5PM, 11:30 AM -> 11:30AM."""
    if not t_raw:
        return ""
    from datetime import datetime as _dt
    if hasattr(t_raw, "strftime"):
        s = t_raw.strftime("%I:%M %p").lstrip("0")
    else:
        s = str(t_raw).strip()
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
            try:
                parsed = _dt.strptime(s, fmt).time()
                s = parsed.strftime("%I:%M %p").lstrip("0")
                break
            except ValueError:
                continue
    # Replace :00 AM/PM with AM/PM (e.g. 5:00 PM -> 5PM, but keep 5:30 PM -> 5:30PM)
    s = s.replace(":00 ", " ").replace(" ", "").upper()
    return s


def _fit_string(c, text: str, font_name: str, font_size: float, max_w: float) -> str:
    """Safely truncate string with ellipsis if it exceeds max_w in ReportLab."""
    if not text:
        return ""
    if c.stringWidth(text, font_name, font_size) <= max_w:
        return text
    while len(text) > 3 and c.stringWidth(text + "…", font_name, font_size) > max_w:
        text = text[:-1]
    return text + "…"


# ─── shared helper: draw one landscape wall-calendar page onto a Canvas ──────
def _draw_calendar_page(c, LS_W, LS_H, year, month, month_events,
                        biz_name, MX, MY, colors_ns):
    """Draw a single landscape wall-calendar month page onto canvas `c`.
    `c` must already be in the correct page; caller is responsible for
    calling c.showPage() afterwards.
    `colors_ns` is the reportlab `colors` module (passed in to avoid
    re-importing inside the helper).
    """
    import calendar as _cal
    from datetime import datetime
    from reportlab.lib.units import mm
    colors = colors_ns

    month_name = _cal.month_name[month]

    # Colour palette
    _C_GR_BG  = colors.HexColor("#DCFCE7")
    _C_GR_TXT = colors.HexColor("#15803D")
    _C_AM_BG  = colors.HexColor("#FEF3C7")
    _C_AM_TXT = colors.HexColor("#B45309")
    _C_RD_BG  = colors.HexColor("#FEE2E2")
    _C_RD_TXT = colors.HexColor("#B91C1C")
    _C_NAV    = colors.HexColor("#0F172A")
    _C_DAY_H  = colors.HexColor("#1E293B")
    _C_CELL_B = colors.HexColor("#CBD5E1")
    _C_EMPTY  = colors.HexColor("#F8FAFC")
    _C_TODAY  = colors.HexColor("#EFF6FF")
    _C_MUTED  = colors.HexColor("#6B7280")

    draw_w = LS_W - 2 * MX

    # KPI pre-calc
    all_evs   = [(d, ev) for d, evl in (month_events or {}).items() for ev in (evl or [])]
    total_pax = sum(int(ev.get("pax", 0) or 0) for _, ev in all_evs)
    total_evs = len(all_evs)
    active_d  = len([d for d, evl in (month_events or {}).items() if evl])

    # ── Title bar ──────────────────────────────────────────────────────────
    TITLE_H = 50
    title_y  = LS_H - MY - TITLE_H

    c.setFillColor(_C_NAV)
    c.roundRect(MX, title_y, draw_w, TITLE_H, 7, fill=1, stroke=0)

    _lp = _logo_path()
    LOGO_S = 34
    if os.path.exists(_lp):
        try:
            c.drawImage(_lp, MX + 10, title_y + (TITLE_H - LOGO_S) / 2,
                        width=LOGO_S, height=LOGO_S,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 24)
    c.drawCentredString(LS_W / 2, title_y + TITLE_H / 2 - 8,
                        f"{month_name.upper()}  {year}")

    kpi_items = [("EVENTS", str(total_evs)),
                 ("PAX",    f"{total_pax:,}"),
                 ("DAYS",   str(active_d))]
    BW, BH, BG = 64, 30, 6
    kx = MX + draw_w - len(kpi_items) * (BW + BG) - 4
    for lbl, val in kpi_items:
        c.setFillColor(colors.HexColor("#1E3A5F"))
        c.roundRect(kx, title_y + (TITLE_H - BH) / 2, BW, BH, 4, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#93C5FD"))
        c.setFont("Helvetica-Bold", 6.5)
        c.drawCentredString(kx + BW / 2, title_y + (TITLE_H + BH) / 2 - 8, lbl)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(kx + BW / 2, title_y + (TITLE_H - BH) / 2 + 5, val)
        kx += BW + BG

    # ── Weekday headers ────────────────────────────────────────────────────
    WDAY_H   = 24
    wday_y   = title_y - WDAY_H
    col_w    = draw_w / 7
    DAYS_FULL = ["SUNDAY", "MONDAY", "TUESDAY", "WEDNESDAY",
                 "THURSDAY", "FRIDAY", "SATURDAY"]
    c.setFillColor(_C_DAY_H)
    c.rect(MX, wday_y, draw_w, WDAY_H, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    for i, dn in enumerate(DAYS_FULL):
        c.drawCentredString(MX + col_w * i + col_w / 2, wday_y + WDAY_H / 2 - 3.5, dn)

    # ── Day cells ──────────────────────────────────────────────────────────
    _cal.setfirstweekday(_cal.SUNDAY)
    month_matrix = _cal.monthcalendar(year, month)
    n_rows  = len(month_matrix)
    avail_h = wday_y - MY - 14
    cell_h  = avail_h / n_rows
    today_dt = datetime.now().date()

    for r_idx, week in enumerate(month_matrix):
        row_bot = wday_y - (r_idx + 1) * cell_h

        for c_idx, day in enumerate(week):
            cx = MX + c_idx * col_w
            cy = row_bot

            if day == 0:
                c.setFillColor(_C_EMPTY)
                c.setStrokeColor(_C_CELL_B)
                c.setLineWidth(0.5)
                c.rect(cx, cy, col_w, cell_h, fill=1, stroke=1)
                continue

            is_today = (today_dt.year == year and
                        today_dt.month == month and
                        today_dt.day == day)

            c.setFillColor(_C_TODAY if is_today else colors.white)
            c.setStrokeColor(_C_CELL_B)
            c.setLineWidth(0.5)
            c.rect(cx, cy, col_w, cell_h, fill=1, stroke=1)

            day_fs = max(10, min(14, cell_h * 0.17))
            c.setFillColor(colors.HexColor("#1D4ED8") if is_today else _C_NAV)
            c.setFont("Helvetica-Bold", day_fs)
            c.drawString(cx + 4, cy + cell_h - day_fs - 3, str(day))
            if is_today:
                c.setFillColor(colors.HexColor("#1D4ED8"))
                c.setFont("Helvetica-Bold", 5.5)
                c.drawString(cx + 4 + day_fs + 3, cy + cell_h - 8, "TODAY")

            day_evs = (month_events or {}).get(day, [])
            if day_evs:
                d_count = len(day_evs)
                d_pax   = sum(int(e.get("pax", 0) or 0) for e in day_evs)

                top_space_y = cy + cell_h - day_fs - 5
                bot_space_y = cy + 2
                avail_h_box = top_space_y - bot_space_y
                bw = col_w - 6
                bx = cx + 3

                # Visual palette per event row (distinct colored accent borders)
                accent_colors = [
                    (colors.HexColor("#F1F5F9"), colors.HexColor("#2563EB"), colors.HexColor("#1E3A8A")),  # Slate/Blue
                    (colors.HexColor("#FFFBEB"), colors.HexColor("#D97706"), colors.HexColor("#92400E")),  # Amber
                    (colors.HexColor("#F0FDF4"), colors.HexColor("#16A34A"), colors.HexColor("#14532D")),  # Emerald
                    (colors.HexColor("#FAF5FF"), colors.HexColor("#9333EA"), colors.HexColor("#581C87")),  # Purple
                    (colors.HexColor("#FFF1F2"), colors.HexColor("#E11D48"), colors.HexColor("#881337")),  # Rose
                    (colors.HexColor("#F0FDFA"), colors.HexColor("#0D9488"), colors.HexColor("#115E59")),  # Teal
                    (colors.HexColor("#FEF2F2"), colors.HexColor("#DC2626"), colors.HexColor("#991B1B")),  # Red
                ]

                N = d_count
                if N == 1:
                    gap = 0.0
                    bh = min(avail_h_box, 30.0)
                    fs_t, fs_p = 7.2, 6.8
                    two_lines = True
                elif N == 2:
                    gap = 2.5
                    bh = min(26.0, (avail_h_box - gap) / 2)
                    fs_t, fs_p = 6.8, 6.2
                    two_lines = True
                elif N == 3:
                    gap = 1.8
                    bh = (avail_h_box - 2 * gap) / 3
                    fs_t, fs_p = 5.8, 5.2
                    two_lines = True
                elif N == 4:
                    gap = 1.2
                    bh = (avail_h_box - 3 * gap) / 4
                    fs_t, fs_p = 5.0, 4.6
                    two_lines = True
                elif N == 5:
                    gap = 1.0
                    bh = (avail_h_box - 4 * gap) / 5
                    fs_t, fs_p = 4.8, 4.4
                    two_lines = (bh >= 13.0)
                else: # N >= 6
                    gap = 0.8
                    bh = (avail_h_box - (N - 1) * gap) / N
                    fs_t = max(4.0, min(4.8, bh * 0.55))
                    fs_p = max(3.8, fs_t * 0.9)
                    two_lines = False

                for i, ev in enumerate(day_evs):
                    by = top_space_y - (i + 1) * bh - i * gap
                    occ = str(ev.get("occasion") or ev.get("name") or ev.get("customer_name") or "EVENT").strip().upper()
                    t_short = _format_time_short(ev.get("time") or ev.get("event_time"))
                    pax = int(ev.get("pax", 0) or 0)
                    header_txt = f"{occ} {t_short}".strip()
                    pax_txt = f"{pax:,} PAX"

                    col_hex = str(ev.get("color_theme") or ev.get("color") or "").strip()
                    if col_hex and col_hex.startswith("#"):
                        try:
                            bar_c = colors.HexColor(col_hex)
                            c_r, c_g, c_b = bar_c.red, bar_c.green, bar_c.blue
                            bg_c = colors.Color(0.93 + 0.07 * c_r, 0.93 + 0.07 * c_g, 0.93 + 0.07 * c_b)
                            txt_c = bar_c
                        except Exception:
                            bg_c, bar_c, txt_c = accent_colors[i % len(accent_colors)]
                    else:
                        bg_c, bar_c, txt_c = accent_colors[i % len(accent_colors)]

                    stripe = 2.2 if N <= 3 else 1.8

                    c.setFillColor(bg_c)
                    c.roundRect(bx, by, bw, bh, 2.0, fill=1, stroke=0)
                    c.setFillColor(bar_c)
                    c.roundRect(bx, by, stripe, bh, 1.0, fill=1, stroke=0)

                    if two_lines:
                        c.setFillColor(colors.HexColor("#0F172A"))
                        c.setFont("Helvetica-Bold", fs_t)
                        fit_h = _fit_string(c, header_txt, "Helvetica-Bold", fs_t, bw - stripe - 5)
                        c.drawString(bx + stripe + 3, by + bh - fs_t - 2.0, fit_h)

                        c.setFillColor(txt_c)
                        c.setFont("Helvetica-Bold", fs_p)
                        fit_p = _fit_string(c, pax_txt, "Helvetica-Bold", fs_p, bw - stripe - 5)
                        c.drawString(bx + stripe + 3, by + 1.8, fit_p)
                    else:
                        line_txt = f"{header_txt} · {pax}p"
                        c.setFillColor(colors.HexColor("#0F172A"))
                        c.setFont("Helvetica-Bold", fs_t)
                        fit_l = _fit_string(c, line_txt, "Helvetica-Bold", fs_t, bw - stripe - 4)
                        c.drawString(bx + stripe + 3, by + (bh - fs_t) / 2, fit_l)

    # ── Legend ─────────────────────────────────────────────────────────────
    leg_y = MY + 1
    leg_items = [
        (_C_GR_BG, _C_GR_TXT, "Available (< 400 pax)"),
        (_C_AM_BG, _C_AM_TXT, "Near Full (400–599 pax)"),
        (_C_RD_BG, _C_RD_TXT, "Fully Booked (600+ pax)"),
    ]
    lx = MX
    for bg_c, tc, lbl in leg_items:
        BOX = 9
        c.setFillColor(bg_c)
        c.roundRect(lx, leg_y, BOX, BOX, 2, fill=1, stroke=0)
        c.setFillColor(tc)
        c.setFont("Helvetica-Bold", 6.5)
        c.drawString(lx + BOX + 3, leg_y + 2, lbl)
        lx += BOX + 3 + c.stringWidth(lbl, "Helvetica-Bold", 6.5) + 18

    now_str = datetime.now().strftime("%b %d, %Y  %I:%M %p")
    c.setFillColor(_C_MUTED)
    c.setFont("Helvetica", 6)
    c.drawRightString(MX + draw_w, leg_y + 2,
                      f"Generated: {now_str}  •  {biz_name}")


# ─── shared helper: build agenda story for one month ─────────────────────────
def _build_agenda_story(year, month, month_events, styles, biz_name):
    """Return a ReportLab Platypus story list for one month's agenda page(s)."""
    import calendar as _cal
    month_name = _cal.month_name[month]

    story = []
    _header_block(story, styles, biz_name,
                  f"Booking Agenda — {month_name} {year}",
                  period=f"{month_name} {year}")
    story.append(Paragraph("Itemized Event Agenda", styles["SectionHead"]))

    all_events = [(d, ev)
                  for d, evl in sorted((month_events or {}).items())
                  for ev in (evl or [])]

    agenda_hdrs = ["Date & Time", "Ref / Occasion", "Customer & Menu", "Venue", "Pax", "Theme / Notes", "Status"]
    agenda_w    = [2.6*cm, 2.6*cm, 4.4*cm, 3.8*cm, 1.4*cm, 3.2*cm, 2.0*cm]
    agenda_rows = [[Paragraph(h, styles["TableHead"]) for h in agenda_hdrs]]

    sstyles = {
        "CONFIRMED": styles["StatusPaid"],
        "COMPLETED": styles["StatusPaid"],
        "PENDING":   styles["StatusPartial"],
        "CANCELLED": styles["StatusUnpaid"],
    }

    if not all_events:
        agenda_rows.append([
            Paragraph("No events scheduled for this month.", styles["TableCellCenter"]),
            "", "", "", "", "", ""
        ])
    else:
        for day, ev in all_events:
            time_str = ev.get("time") or "6:00 PM"
            st_key   = str(ev.get("status") or "CONFIRMED").upper()
            c_name   = ev.get("customer_name") or ev.get("name") or "Valued Client"
            occ      = ev.get("occasion") or "Event"
            menu_txt = ev.get("menu") or ev.get("package_name") or "Standard Menu"
            theme_notes = str(ev.get("notes") or ev.get("theme") or ev.get("description_theme") or ev.get("description") or "Standard Setup").strip()
            if not theme_notes:
                theme_notes = "Standard Setup"

            agenda_rows.append([
                Paragraph(
                    f"<b>{month_name[:3]} {day}, {year}</b><br/>"
                    f"<font color='#6B7280' size=7>{time_str}</font>",
                    styles["TableCell"]),
                Paragraph(
                    f"<font color='#E11D48'><b>{ev.get('ref') or '—'}</b></font><br/>"
                    f"<font color='#4B5563' size=7>{occ}</font>",
                    styles["TableCell"]),
                Paragraph(
                    f"<b>{c_name}</b><br/>"
                    f"<font color='#6B7280' size=7>{menu_txt}</font>",
                    styles["TableCell"]),
                Paragraph(str(ev.get("venue") or ev.get("location") or "—"), styles["TableCell"]),
                Paragraph(str(ev.get("pax", 0)), styles["TableCellCenter"]),
                Paragraph(f"<font color='#374151'>{theme_notes}</font>", styles["TableCell"]),
                Paragraph(str(st_key).capitalize(),
                          sstyles.get(st_key, styles["TableCellCenter"])),
            ])

    tbl = Table(agenda_rows, colWidths=agenda_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  _C_DARK),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [_C_WHITE, _C_LIGHT]),
        ("BOX",           (0, 0), (-1, -1), 0.4, _C_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, _C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tbl)
    _footer(story, styles, biz_name)
    return story


def _draw_agenda_canvas_pages(c, year, month, month_events, biz_name, styles):
    """Draw one or more LANDSCAPE A4 agenda pages onto canvas `c` with full booking details."""
    import calendar as _cal
    from datetime import datetime
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm, mm

    LS = landscape(A4)
    PW, PH = LS          # 841.89 x 595.28
    MX = 16 * mm         # matches calendar page margins
    MY = 12 * mm
    CW = PW - 2 * MX

    # Colour aliases
    C_DARK   = _C_DARK
    C_MUTED  = _C_MUTED
    C_BORDER = _C_BORDER
    C_LIGHT  = _C_LIGHT
    C_GREEN  = _C_GREEN
    C_AMBER  = _C_AMBER
    C_RED    = _C_RED
    C_WHITE  = _C_WHITE

    month_name = _cal.month_name[month]
    now_str    = datetime.now().strftime("%b %d, %Y  %I:%M %p")

    all_events = [
        (d, ev)
        for d, evl in sorted((month_events or {}).items())
        for ev in (evl or [])
    ]

    # ── Column layout (landscape, 7 columns) ─────────────────────────────
    COL_X = [
        MX + 4,          # 0: Date & Time (84pt)
        MX + 92,         # 1: Ref & Occasion (100pt)
        MX + 196,        # 2: Customer & Menu (165pt)
        MX + 366,        # 3: Venue & Address (145pt)
        MX + 516,        # 4: Pax (40pt)
        MX + 560,        # 5: Description / Theme (115pt)
        MX + 680,        # 6: Status (68pt)
    ]
    COL_HDRS   = ["DATE & TIME", "REF / OCCASION", "CUSTOMER & MENU", "VENUE / ADDRESS", "PAX", "THEME / NOTES", "STATUS"]
    COL_MAXW   = [84, 98, 160, 140, 36, 110, 65]

    HEADER_H   = 50   # title bar height
    COL_HDR_H  = 22   # column-header row height
    ROW_H      = 36   # data row height (2-line layout)
    PAGE_BOT   = MY + 14

    def start_page():
        c.setPageSize(LS)

        # Dark title bar
        c.setFillColor(C_DARK)
        c.roundRect(MX, PH - MY - HEADER_H, CW, HEADER_H, 7, fill=1, stroke=0)

        # Logo
        _lp = _logo_path()
        LOGO_S = 34
        if os.path.exists(_lp):
            try:
                c.drawImage(_lp, MX + 10,
                            PH - MY - HEADER_H + (HEADER_H - LOGO_S) / 2,
                            width=LOGO_S, height=LOGO_S,
                            preserveAspectRatio=True, mask="auto")
            except Exception:
                pass

        # Title
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 17)
        c.drawCentredString(PW / 2,
                            PH - MY - HEADER_H / 2 - 6,
                            f"Booking Agenda — {month_name.upper()}  {year}")

        # Biz name right-aligned in header
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#93C5FD"))
        c.drawRightString(MX + CW - 6,
                          PH - MY - HEADER_H + 8, biz_name)

        # Column header row
        hdr_y = PH - MY - HEADER_H - COL_HDR_H
        c.setFillColor(colors.HexColor("#1E293B"))
        c.rect(MX, hdr_y, CW, COL_HDR_H, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 7.5)
        for cx, hdr in zip(COL_X, COL_HDRS):
            c.drawString(cx, hdr_y + 7, hdr)

        # Footer line
        c.setFillColor(C_MUTED)
        c.setFont("Helvetica", 6)
        c.drawCentredString(PW / 2, MY - 2,
                            f"Generated: {now_str}  •  {biz_name}")

        return hdr_y - 1

    def status_color(st):
        st = str(st or "").upper()
        if st in ("CONFIRMED", "COMPLETED"):
            return C_GREEN
        if st == "PENDING":
            return C_AMBER
        if st == "CANCELLED":
            return C_RED
        return C_MUTED

    def trunc(text, font, size, max_w):
        t = str(text or "")
        if c.stringWidth(t, font, size) <= max_w:
            return t
        while t and c.stringWidth(t + "…", font, size) > max_w:
            t = t[:-1]
        return t + "…"

    y       = start_page()
    row_num = 0

    if not all_events:
        c.setFillColor(C_MUTED)
        c.setFont("Helvetica-Oblique", 10)
        c.drawCentredString(PW / 2, y - ROW_H,
                            "No events scheduled for this month.")
        c.showPage()
        return

    for day, ev in all_events:
        if y - ROW_H < PAGE_BOT:
            c.showPage()
            y = start_page()
            row_num = 0

        # Row background
        c.setFillColor(C_WHITE if row_num % 2 == 0 else C_LIGHT)
        c.rect(MX, y - ROW_H, CW, ROW_H, fill=1, stroke=0)

        # Bottom border
        c.setStrokeColor(C_BORDER)
        c.setLineWidth(0.3)
        c.line(MX, y - ROW_H, MX + CW, y - ROW_H)

        # ── 0: DATE & TIME ────────────────────────────────────────────────
        c.setFillColor(C_DARK)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(COL_X[0], y - 14, f"{month_name[:3]} {day}, {year}")
        c.setFont("Helvetica", 7)
        c.setFillColor(C_MUTED)
        c.drawString(COL_X[0], y - 26, str(ev.get("time") or "6:00 PM"))

        # ── 1: REF / OCCASION ─────────────────────────────────────────────
        c.setFillColor(colors.HexColor("#E11D48"))
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(COL_X[1], y - 14, trunc(ev.get("ref") or "—", "Helvetica-Bold", 8.5, COL_MAXW[1]))
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#4B5563"))
        c.drawString(COL_X[1], y - 26, trunc(ev.get("occasion") or "Event", "Helvetica", 7, COL_MAXW[1]))

        # ── 2: CUSTOMER & MENU ────────────────────────────────────────────
        c.setFillColor(C_DARK)
        c.setFont("Helvetica-Bold", 8.5)
        c_name = ev.get("customer_name") or ev.get("name") or "Valued Client"
        c.drawString(COL_X[2], y - 14, trunc(c_name, "Helvetica-Bold", 8.5, COL_MAXW[2]))
        c.setFont("Helvetica", 7)
        c.setFillColor(C_MUTED)
        menu_desc = ev.get("menu") or ev.get("package_name") or "Standard Package"
        c.drawString(COL_X[2], y - 26, trunc(menu_desc, "Helvetica", 7, COL_MAXW[2]))

        # ── 3: VENUE / ADDRESS ────────────────────────────────────────────
        venue_str = ev.get("venue") or ev.get("location") or "Client Venue"
        c.setFillColor(C_DARK)
        c.setFont("Helvetica", 8)
        c.drawString(COL_X[3], y - 14, trunc(venue_str, "Helvetica", 8, COL_MAXW[3]))
        addr_str = ev.get("address") or ""
        if addr_str and addr_str != venue_str:
            c.setFont("Helvetica", 6.5)
            c.setFillColor(C_MUTED)
            c.drawString(COL_X[3], y - 26, trunc(addr_str, "Helvetica", 6.5, COL_MAXW[3]))

        # ── 4: PAX ────────────────────────────────────────────────────────
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(C_DARK)
        c.drawCentredString(COL_X[4] + 18, y - 15, str(ev.get("pax") or 0))
        c.setFont("Helvetica", 6.5)
        c.setFillColor(C_MUTED)
        c.drawCentredString(COL_X[4] + 18, y - 26, "guests")

        # ── 5: THEME / NOTES ──────────────────────────────────────────────
        theme_txt = str(ev.get("notes") or ev.get("theme") or ev.get("description_theme") or ev.get("description") or "Standard Setup").strip()
        if not theme_txt:
            theme_txt = "Standard Setup"

        c.setFillColor(C_DARK)
        if c.stringWidth(theme_txt, "Helvetica", 7.5) <= COL_MAXW[5]:
            c.setFont("Helvetica", 7.5)
            c.drawString(COL_X[5], y - 18, theme_txt)
        else:
            words = theme_txt.split()
            line1, line2 = "", ""
            for word in words:
                test_l1 = (line1 + " " + word).strip()
                if c.stringWidth(test_l1, "Helvetica", 7.5) <= COL_MAXW[5]:
                    line1 = test_l1
                else:
                    line2 = (line2 + " " + word).strip()
            if not line1:
                line1 = trunc(theme_txt, "Helvetica", 7.5, COL_MAXW[5])
            c.setFont("Helvetica", 7.5)
            c.drawString(COL_X[5], y - 14, line1)
            if line2:
                c.setFont("Helvetica", 7)
                c.setFillColor(C_MUTED)
                c.drawString(COL_X[5], y - 26, trunc(line2, "Helvetica", 7, COL_MAXW[5]))

        # ── 6: STATUS ─────────────────────────────────────────────────────
        st_key = str(ev.get("status") or "CONFIRMED").upper()
        st_col = status_color(st_key)

        badge_lbl = st_key.capitalize()
        badge_w   = max(56, c.stringWidth(badge_lbl, "Helvetica-Bold", 7.5) + 14)
        badge_x   = COL_X[6]
        badge_y   = y - ROW_H + 9
        badge_h   = 18

        if st_key in ("CONFIRMED", "COMPLETED"):
            pill_bg = colors.HexColor("#DCFCE7")
        elif st_key == "PENDING":
            pill_bg = colors.HexColor("#FEF3C7")
        elif st_key == "CANCELLED":
            pill_bg = colors.HexColor("#FEE2E2")
        else:
            pill_bg = colors.HexColor("#F1F5F9")

        c.setFillColor(pill_bg)
        c.roundRect(badge_x, badge_y, badge_w, badge_h, 4, fill=1, stroke=0)
        c.setFillColor(st_col)
        c.setFont("Helvetica-Bold", 7.5)
        c.drawCentredString(badge_x + badge_w / 2, badge_y + 5, badge_lbl)

        y -= ROW_H
        row_num += 1

    c.showPage()




def export_calendar_pdf_range(
    save_path: str,
    months: list,               # list of (year, month) tuples
    events_by_month: dict,      # {(year, month): {day: [event_dicts]}}
    biz_name: str = "Jayraldine's Catering",
    include_agenda: bool = True,
    include_empty: bool = False,
) -> bool:
    """Export a multi-month printable wall-calendar PDF.

    Uses a SINGLE ReportLab Canvas with setPageSize() between pages — no
    external PDF merge library (pypdf / PyPDF2) is required.

    Each month gets one landscape A4 calendar page. When *include_agenda*
    is True, landscape A4 agenda pages follow each calendar page.
    When *include_empty* is False, months with zero bookings are skipped.
    """
    if not REPORTLAB_OK:
        return False
    try:
        import calendar as _cal
        from reportlab.lib.pagesizes import landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as _canvas_mod

        LS     = landscape(A4)
        LS_W, LS_H = LS
        MX, MY = 16 * mm, 12 * mm
        styles = _styles()

        # One canvas, one file
        c = _canvas_mod.Canvas(save_path)
        c.setTitle(f"{biz_name} — Calendar Export")
        c.setAuthor(biz_name)

        wrote_any = False

        for (year, month) in months:
            month_events = events_by_month.get((year, month), {})
            has_bookings = any(bool(v) for v in month_events.values())

            if not has_bookings and not include_empty:
                continue

            # ── Landscape calendar page ───────────────────────────────────
            c.setPageSize(LS)
            _draw_calendar_page(c, LS_W, LS_H, year, month,
                                month_events, biz_name, MX, MY, colors)
            c.showPage()
            wrote_any = True

            # ── Landscape agenda pages ───────────────────────────────────
            if include_agenda:
                _draw_agenda_canvas_pages(c, year, month, month_events,
                                          biz_name, styles)

        if not wrote_any:
            # Nothing to export — write a single placeholder page
            c.setPageSize(LS)
            c.setFont("Helvetica-Bold", 16)
            c.setFillColorRGB(0.4, 0.4, 0.4)
            c.drawCentredString(LS_W / 2, LS_H / 2,
                                "No bookings found for the selected date range.")
            c.showPage()

        c.save()
        return True

    except Exception as exc:
        print(f"[exporter] export_calendar_pdf_range failed: {exc}")
        import traceback; traceback.print_exc()
        return False



def export_calendar_pdf(arg1, arg2, arg3, month_events: dict = None,
                        biz_name: str = "Jayraldine's Catering") -> bool:
    """Generate a printable monthly wall-calendar PDF (single month).
    Flexibly supports (save_path, year, month, events) or (year, month, save_path, events).
    """
    if isinstance(arg1, str) and (arg1.endswith(".pdf") or "/" in arg1 or "\\" in arg1):
        save_path = str(arg1)
        year = int(arg2)
        month = int(arg3)
    elif isinstance(arg3, str) and (arg3.endswith(".pdf") or "/" in arg3 or "\\" in arg3):
        year = int(arg1)
        month = int(arg2)
        save_path = str(arg3)
    else:
        save_path = str(arg1)
        year = int(arg2)
        month = int(arg3)

    return export_calendar_pdf_range(
        save_path       = save_path,
        months          = [(year, month)],
        events_by_month = {(year, month): month_events or {}},
        biz_name        = biz_name,
        include_agenda  = True,
        include_empty   = True,   # always export even if empty (single-month call)
    )



def export_cash_flow_pdf(save_path: str, transactions: Optional[list] = None,
                         summary: Optional[dict] = None,
                         biz_name: str = "Jayraldine's Catering") -> bool:
    """Generate a clean, high-quality A4 PDF of the Cash Flow Ledger with KPI summary and itemized entries."""
    if not REPORTLAB_OK:
        return False
    try:
        import utils.repository as repo
        tx_list = transactions if transactions is not None else (repo.get_cash_flow_transactions() or [])
        smry = summary if summary is not None else repo.get_cash_flow_summary()

        doc = SimpleDocTemplate(
            save_path, pagesize=A4,
            leftMargin=_MARGIN, rightMargin=_MARGIN,
            topMargin=_MARGIN, bottomMargin=_MARGIN,
            title=f"{biz_name} — Cash Flow Statement",
            author=biz_name,
        )
        styles = _styles()
        story = []

        now_str = datetime.now().strftime("%B %d, %Y")
        _header_block(story, styles, biz_name, "Cash Flow Statement & Ledger", period=now_str)

        # Top KPI Summary Card
        tot_dep = float(smry.get("total_deposits", 0.0))
        tot_with = float(smry.get("total_withdrawals", 0.0))
        cur_bal = float(smry.get("current_balance", 0.0))
        tot_sales = float(smry.get("total_actual_sales", 0.0))
        tot_diff = cur_bal - tot_sales

        cf_kpis = [
            ("Total Deposits (In)", f"₱{tot_dep:,.2f}"),
            ("Total Withdrawals (Out)", f"₱{tot_with:,.2f}"),
            ("Running Balance", f"₱{cur_bal:,.2f}"),
            ("Total Actual Sales", f"₱{tot_sales:,.2f}"),
        ]
        ncols = len(cf_kpis)
        col_w = [_CONTENT_W / ncols] * ncols
        labels_row = [Paragraph(lbl, styles["KpiLabel"]) for lbl, _ in cf_kpis]
        values_row = [Paragraph(val, styles["KpiValue"]) for _, val in cf_kpis]
        t = Table([labels_row, values_row], colWidths=col_w)
        t.setStyle(TableStyle([
            ("BOX",           (0, 0), (-1, -1), 0.5, _C_BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.5, _C_BORDER),
            ("BACKGROUND",    (0, 0), (-1, 0),  _C_LIGHT),
            ("BACKGROUND",    (0, 1), (-1, 1),  _C_WHITE),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(t)
        story.append(Spacer(1, 14))

        # Ledger Table
        story.append(Paragraph("Transaction Journal & Running Balances", styles["SectionHead"]))
        table_hdrs = ["Date", "Check #", "Particulars (Account / Detail)", "Deposit", "Withdrawal", "Balance", "Actual Sales", "Variance"]
        table_w = [2.0*cm, 1.8*cm, 4.4*cm, 2.1*cm, 2.1*cm, 2.2*cm, 2.1*cm, 2.1*cm]
        table_rows = [[Paragraph(h, styles["TableHead"]) for h in table_hdrs]]

        if not tx_list:
            table_rows.append([Paragraph("No cash flow transactions recorded.", styles["TableCellCenter"]), "", "", "", "", "", "", ""])
        else:
            for tx in tx_list:
                d_str = str(tx.get("date") or "")
                chk_str = str(tx.get("check_no") or "—")
                part_str = str(tx.get("particulars") or "")
                dep_val = float(tx.get("deposit") or 0.0)
                with_val = float(tx.get("withdrawal") or 0.0)
                bal_val = float(tx.get("balance") or 0.0)
                sales_val = float(tx.get("actual_sales") or 0.0)
                diff_val = bal_val - sales_val

                dep_txt = f"<font color='#16A34A'>₱{dep_val:,.2f}</font>" if dep_val > 0 else "—"
                with_txt = f"<font color='#DC2626'>₱{with_val:,.2f}</font>" if with_val > 0 else "—"
                bal_txt = f"<b>₱{bal_val:,.2f}</b>" if bal_val >= 0 else f"<font color='#DC2626'><b>(₱{abs(bal_val):,.2f})</b></font>"
                sales_txt = f"₱{sales_val:,.2f}" if sales_val > 0 else "—"
                diff_txt = f"₱{diff_val:,.2f}" if diff_val >= 0 else f"<font color='#DC2626'>(₱{abs(diff_val):,.2f})</font>"
                if sales_val <= 0:
                    diff_txt = "—"

                table_rows.append([
                    Paragraph(d_str, styles["TableCellCenter"]),
                    Paragraph(chk_str, styles["TableCellCenter"]),
                    Paragraph(part_str, styles["TableCell"]),
                    Paragraph(dep_txt, styles["TableCellRight"]),
                    Paragraph(with_txt, styles["TableCellRight"]),
                    Paragraph(bal_txt, styles["TableCellRight"]),
                    Paragraph(sales_txt, styles["TableCellRight"]),
                    Paragraph(diff_txt, styles["TableCellRight"]),
                ])

        tbl = Table(table_rows, colWidths=table_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  _C_DARK),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [_C_WHITE, _C_LIGHT]),
            ("BOX",           (0, 0), (-1, -1), 0.4, _C_BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)

        _footer(story, styles, biz_name)
        doc.build(story)
        return True
    except Exception as exc:
        print(f"[exporter] export_cash_flow_pdf failed: {exc}")
        return False
