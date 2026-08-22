"""
Data Importer Engine — handles intelligent auto-mapping, file parsing (CSV & Excel),
field normalization, data validation preview, sample template generation, and
batch database insertion for Jayraldine's Catering System.
"""
import os
import re
import csv
import math
from datetime import datetime, date, timedelta
from typing import List, Dict, Tuple, Any, Optional

import utils.repository as repo

# ─────────────────────────────────────────────────────────────────────────────
# ENTITY SCHEMAS
# ─────────────────────────────────────────────────────────────────────────────

ENTITY_SCHEMAS = {
    "all_in_one": {
        "title": "All-in-One Master File (Bookings, Customers, Expenses & Menu)",
        "fields": {
            "customer_name": {"label": "Customer Name", "required": True},
            "contact": {"label": "Contact Number", "required": False},
            "event_date": {"label": "Event Date", "required": True},
            "event_time": {"label": "Event Time", "required": False},
            "venue": {"label": "Venue / Location", "required": False},
            "occasion": {"label": "Occasion", "required": False},
            "pax": {"label": "Guest Count (Pax)", "required": True},
            "total_amount": {"label": "Booking Total (₱)", "required": True},
            "expense_date": {"label": "Expense Date", "required": False},
            "expense_category": {"label": "Expense Category", "required": False},
            "expense_description": {"label": "Expense Description", "required": False},
            "expense_amount": {"label": "Expense Amount (₱)", "required": False},
        },
        "sample": [
            ["Customer Name", "Contact Number", "Event Date", "Event Time", "Venue", "Occasion", "Pax", "Total Amount", "Expense Date", "Expense Category", "Expense Description", "Expense Amount"],
            ["Maria Santos", "09171234567", datetime.now().strftime("%Y-%m-%d"), "6:00 PM", "Grand Ballroom", "Wedding", "150", "45000.00", datetime.now().strftime("%Y-%m-%d"), "Food Cost", "Fresh Pork Lechon", "15500.00"],
            ["Juan Dela Cruz", "09189876543", datetime.now().strftime("%Y-%m-%d"), "12:00 PM", "Lahug Clubhouse", "Birthday", "80", "28000.00", datetime.now().strftime("%Y-%m-%d"), "Transport", "Gas for Van", "2400.00"],
        ]
    },
    "bookings": {
        "title": "Bookings & Orders",
        "fields": {
            "name": {"label": "Customer Name", "required": True},
            "contact": {"label": "Contact Number", "required": False},
            "email": {"label": "Email Address", "required": False},
            "address": {"label": "Address", "required": False},
            "occasion": {"label": "Occasion", "required": False},
            "venue": {"label": "Venue / Location", "required": False},
            "date": {"label": "Event Date", "required": True},
            "time": {"label": "Event Time", "required": False},
            "pax": {"label": "Guest Count (Pax)", "required": True},
            "total": {"label": "Total Amount (₱)", "required": True},
            "notes": {"label": "Special Notes / Add-ons", "required": False},
            "status": {"label": "Status", "required": False},
        },
        "sample": [
            ["Customer Name", "Contact Number", "Email Address", "Address", "Occasion", "Venue", "Event Date", "Event Time", "Guest Count (Pax)", "Total Amount (₱)", "Special Notes / Add-ons", "Status"],
            ["Engr. Rodrigo Tan", "09178889900", "rodrigo.tan@example.com", "Cebu City", "Wedding", "Grand Ballroom Cebu", datetime.now().strftime("%Y-%m-%d"), "6:00 PM", "150", "45000.00", "Includes Lechon & Backdrop", "CONFIRMED"],
            ["Capt. Juanito Dela Cruz", "09182223344", "juanito.dc@example.com", "Mandaue City", "Birthday", "Lahug Clubhouse", datetime.now().strftime("%Y-%m-%d"), "12:00 PM", "80", "28000.00", "Buffet Setup", "CONFIRMED"],
        ]
    },
    "customers": {
        "title": "Customers",
        "fields": {
            "name": {"label": "Customer Name", "required": True},
            "contact": {"label": "Contact Number", "required": False},
            "email": {"label": "Email Address", "required": False},
            "address": {"label": "Address", "required": False},
            "status": {"label": "Status", "required": False},
            "notes": {"label": "Notes / History", "required": False},
        },
        "sample": [
            ["Customer Name", "Contact Number", "Email Address", "Address", "Status", "Notes"],
            ["Maria Santos", "09171234567", "maria@example.com", "Cebu City", "Active", "VIP Client"],
            ["Juan Dela Cruz", "09189876543", "juan@example.com", "Mandaue City", "Active", "Prefers Pork Lechon"],
        ]
    },
    "expenses": {
        "title": "Expenses",
        "fields": {
            "date": {"label": "Expense Date", "required": True},
            "category": {"label": "Category", "required": True},
            "description": {"label": "Description", "required": True},
            "amount": {"label": "Amount (₱)", "required": True},
        },
        "sample": [
            ["Expense Date", "Category", "Description", "Amount (₱)"],
            [datetime.now().strftime("%Y-%m-%d"), "Food Cost", "Fresh Pork & Spices", "15500.00"],
            [datetime.now().strftime("%Y-%m-%d"), "Transport", "Gas for Delivery Van", "2400.00"],
            [datetime.now().strftime("%Y-%m-%d"), "Labor", "Assistant Cook Daily Pay", "3500.00"],
        ]
    },
    "menu_items": {
        "title": "Menu Items & Packages",
        "fields": {
            "name": {"label": "Item / Package Name", "required": True},
            "category": {"label": "Category", "required": True},
            "price": {"label": "Price / Rate (₱)", "required": True},
            "description": {"label": "Description / Inclusions", "required": False},
        },
        "sample": [
            ["Item / Package Name", "Category", "Price / Rate (₱)", "Description / Inclusions"],
            ["Lechon Belly Package A", "Packages", "12500.00", "1 Whole Lechon Belly + 3 Side Dishes"],
            ["Special Pork Humba", "Main Course", "450.00", "Serves 8-10 pax"],
            ["Biko with Latik", "Dessert", "250.00", "1 Large Tray"],
        ]
    }
}

ENTITY_HEADER_ALIASES = {
    "all_in_one": {
        "customer_name": ["customer name", "client name", "customer", "client", "name", "full name"],
        "contact": ["contact number", "contact", "phone number", "phone", "mobile", "cellphone", "tel"],
        "event_date": ["event date", "booking date", "date", "fecha"],
        "event_time": ["event time", "time", "start time", "schedule"],
        "venue": ["venue location", "venue", "event venue", "location", "place", "event location", "site", "address"],
        "occasion": ["occasion", "event type", "event", "celebration", "party", "theme"],
        "pax": ["guest count pax", "guest count", "pax", "guests", "number of guests", "attendees", "headcount", "capacity"],
        "total_amount": ["booking total", "total amount", "total", "amount", "cost", "price", "grand total"],
        "expense_date": ["expense date", "exp date"],
        "expense_category": ["expense category", "category", "exp category", "expense type", "type"],
        "expense_description": ["expense description", "particulars", "description", "details", "memo", "notes", "inclusions"],
        "expense_amount": ["expense amount", "exp amount", "cost php", "price php"],
    },
    "bookings": {
        "name": ["customer name", "client name", "customer", "client", "full name", "name"],
        "contact": ["contact number", "contact", "phone number", "phone", "mobile", "cellphone", "tel"],
        "email": ["email address", "email", "e-mail", "mail"],
        "address": ["address", "home address", "client address", "city", "street"],
        "date": ["event date", "booking date", "date", "fecha"],
        "time": ["event time", "time", "schedule", "start time"],
        "venue": ["venue location", "venue", "event venue", "location", "place", "event location", "site"],
        "occasion": ["occasion", "event", "event type", "celebration", "party", "theme"],
        "pax": ["guest count pax", "guest count", "pax", "guests", "number of guests", "attendees", "headcount", "capacity"],
        "total": ["total amount", "total", "amount", "booking total", "cost", "price", "subtotal", "grand total", "fee"],
        "notes": ["special notes add-ons", "special notes", "notes", "add-ons", "remarks", "inclusions", "details", "comments"],
        "status": ["status", "booking status", "state", "order status"],
    },
    "customers": {
        "name": ["customer name", "client name", "customer", "client", "full name", "name"],
        "contact": ["contact number", "contact", "phone number", "phone", "mobile", "cellphone", "tel"],
        "email": ["email address", "email", "e-mail", "mail"],
        "address": ["home address", "address", "location", "city", "street"],
        "status": ["status", "customer status", "state", "account status"],
        "notes": ["notes", "history", "remarks", "memo", "details", "comments"],
    },
    "expenses": {
        "date": ["expense date", "date", "transaction date", "fecha"],
        "category": ["expense category", "category", "expense type", "type", "item category", "group"],
        "description": ["expense description", "particulars", "description", "details", "memo", "notes", "inclusions", "summary"],
        "amount": ["expense amount", "amount", "cost", "fee", "cost php", "price php", "price", "subtotal", "rate"],
    },
    "menu_items": {
        "name": ["item package name", "item name", "package name", "dish name", "name", "title", "item"],
        "category": ["category", "item category", "menu category", "type", "group"],
        "price": ["price rate", "price", "rate", "cost", "unit price", "amount", "price php"],
        "description": ["description inclusions", "description", "inclusions", "details", "notes", "summary"],
    }
}


def _clean_header_str(text: str) -> str:
    """Strip symbols, punctuation, parentheses, and extra spaces for resilient matching."""
    s = str(text or "").lower().strip()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def normalize_expense_category(cat: Any) -> str:
    if not cat:
        return "Other"
    s = str(cat).strip().lower()
    if any(w in s for w in ["food", "meat", "pork", "beef", "chicken", "fish", "vegetable", "spice", "ingredient", "grocery", "market", "rice", "dish", "beverage", "drink"]):
        return "Food Cost"
    if any(w in s for w in ["labor", "salary", "wage", "cook", "chef", "crew", "helper", "staff", "assistant", "payroll", "waiter"]):
        return "Labor"
    if any(w in s for w in ["transport", "transpo", "gas", "fuel", "diesel", "delivery", "van", "truck", "fare", "parking"]):
        return "Transport"
    if any(w in s for w in ["util", "electric", "water", "power", "gasul", "lpg", "internet", "bill", "phone"]):
        return "Utilities"
    if any(w in s for w in ["equip", "rent", "table", "chair", "tent", "chafing", "utensil", "plate", "pot", "pan", "appliance", "sound", "light"]):
        return "Equipment"
    return "Other"


def normalize_booking_status(status: Any) -> str:
    if not status:
        return "PENDING"
    s = str(status).strip().upper()
    if "CONFIRM" in s:
        return "CONFIRMED"
    if "COMPLET" in s:
        return "COMPLETED"
    if "CANCEL" in s:
        return "CANCELLED"
    return "PENDING"


def normalize_customer_status(status: Any) -> str:
    if not status:
        return "Active"
    s = str(status).strip().lower()
    if "inact" in s or "dorm" in s:
        return "Inactive"
    if "pend" in s:
        return "Pending"
    return "Active"


def normalize_menu_category(cat: Any) -> str:
    if not cat:
        return "Main Course"
    s = str(cat).strip().lower()
    if "nood" in s or "pasta" in s or "pancit" in s or "spaghetti" in s:
        return "Noodles"
    if "soup" in s or "broth" in s or "sinigang" in s or "tinola" in s:
        return "Soup"
    if "veg" in s or "salad" in s:
        return "Vegetables"
    if "dessert" in s or "sweet" in s or "cake" in s or "biko" in s or "leche" in s:
        return "Dessert"
    if "drink" in s or "beverage" in s or "juice" in s or "soda" in s or "tea" in s:
        return "Drinks"
    if "bread" in s or "pastry" in s or "bun" in s:
        return "Bread"
    if "pack" in s or "set" in s:
        return "Other"
    if any(w in s for w in ["main", "pork", "beef", "chicken", "fish", "seafood", "meat", "lechon", "humba", "caldereta"]):
        return "Main Course"
    return "Other"


def normalize_amount(raw: Any) -> float:
    if raw is None:
        return 0.0
    s = str(raw).strip()
    s = re.sub(r"[^\d.-]", "", s.replace(",", ""))
    try:
        return float(s)
    except ValueError:
        return 0.0


def normalize_date(raw: Any) -> str:
    from datetime import timedelta
    if not raw:
        return datetime.now().strftime("%Y-%m-%d")

    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, date):
        return raw.strftime("%Y-%m-%d")

    # Numeric Excel serial date (e.g. 45000 -> date in 2023-2030)
    if isinstance(raw, (int, float)):
        try:
            if 20000 <= raw <= 80000:
                dt = datetime(1899, 12, 30) + timedelta(days=float(raw))
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass

    s = str(raw).strip()
    if not s:
        return datetime.now().strftime("%Y-%m-%d")

    # If already YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s

    # Strip time suffix if present e.g. " 00:00:00", "T18:30:00", " 12:00:00 PM"
    clean_s = re.sub(r"[T\s]+\d{1,2}:\d{2}(:\d{2})?(\.\d+)?(\s*[AP]M)?.*$", "", s, flags=re.IGNORECASE).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", clean_s):
        return clean_s

    # Also check if clean_s is numeric excel serial as string
    if clean_s.replace(".", "", 1).isdigit() and len(clean_s) in (5, 6, 7):
        try:
            num = float(clean_s)
            if 20000 <= num <= 80000:
                dt = datetime(1899, 12, 30) + timedelta(days=num)
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass

    date_formats = [
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y",
        "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y",
        "%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d-%B-%Y",
        "%b-%d-%Y", "%B-%d-%Y",
        "%m/%d/%y", "%d/%m/%y", "%y-%m-%d"
    ]
    for fmt in date_formats:
        try:
            dt = datetime.strptime(clean_s, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    try:
        dt = datetime.fromisoformat(s[:10])
        return dt.strftime("%Y-%m-%d")
    except Exception:
        pass

    return datetime.now().strftime("%Y-%m-%d")


def normalize_pax(raw: Any) -> int:
    if not raw:
        return 50
    s = str(raw).strip()
    s = re.sub(r"\D", "", s)
    try:
        v = int(s)
        return max(1, v)
    except ValueError:
        return 50


# ─────────────────────────────────────────────────────────────────────────────
# FILE PARSING (CSV & EXCEL)
# ─────────────────────────────────────────────────────────────────────────────

def parse_file(file_path: str) -> Tuple[List[str], List[Dict[str, str]], Optional[str]]:
    if not os.path.exists(file_path):
        return [], [], "File does not exist."

    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return _parse_excel(file_path)
    return _parse_csv(file_path)


def _parse_csv(file_path: str) -> Tuple[List[str], List[Dict[str, str]], Optional[str]]:
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252", "utf-16"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                sample = f.read(4096)
                f.seek(0)
                delimiter = ","
                if ";" in sample and sample.count(";") > sample.count(","):
                    delimiter = ";"
                elif "\t" in sample and sample.count("\t") > sample.count(","):
                    delimiter = "\t"

                reader = csv.reader(f, delimiter=delimiter)
                rows_raw = [row for row in reader if any(cell.strip() for cell in row)]
                if not rows_raw:
                    return [], [], "CSV file is empty."

                headers = [h.strip() for h in rows_raw[0]]
                data_rows = []
                for idx, raw_row in enumerate(rows_raw[1:], start=2):
                    row_dict = {}
                    for h_idx, h_name in enumerate(headers):
                        val = raw_row[h_idx].strip() if h_idx < len(raw_row) else ""
                        row_dict[h_name] = val
                    data_rows.append(row_dict)

                return headers, data_rows, None
        except UnicodeDecodeError:
            continue
        except Exception as e:
            return [], [], f"Failed to read CSV: {e}"

    return [], [], "Unsupported text encoding."


def _parse_excel(file_path: str) -> Tuple[List[str], List[Dict[str, str]], Optional[str]]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        rows_raw = []
        for row in sheet.iter_rows(values_only=True):
            if row and any(cell is not None and str(cell).strip() for cell in row):
                rows_raw.append([str(c).strip() if c is not None else "" for c in row])

        if not rows_raw:
            return [], [], "Excel worksheet is empty."

        headers = rows_raw[0]
        data_rows = []
        for raw_row in rows_raw[1:]:
            row_dict = {}
            for h_idx, h_name in enumerate(headers):
                val = raw_row[h_idx] if h_idx < len(raw_row) else ""
                row_dict[h_name] = val
            data_rows.append(row_dict)

        return headers, data_rows, None
    except Exception as e:
        return [], [], f"Failed to read Excel file: {e}"


def parse_master_file(file_path: str) -> Tuple[Dict[str, Tuple[List[str], List[Dict[str, str]]]], Optional[str]]:
    """Parse a multi-sheet Excel file or single-sheet master file into entity sections.
    Returns ({entity_type: (headers, data_rows)}, error_message)."""
    if not os.path.exists(file_path):
        return {}, "File does not exist."

    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if len(wb.sheetnames) > 1:
                result = {}
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    rows_raw = []
                    for row in sheet.iter_rows(values_only=True):
                        if row and any(cell is not None and str(cell).strip() for cell in row):
                            rows_raw.append([str(c).strip() if c is not None else "" for c in row])
                    if not rows_raw:
                        continue

                    headers = rows_raw[0]
                    data_rows = []
                    for raw_row in rows_raw[1:]:
                        row_dict = {headers[i]: (raw_row[i] if i < len(raw_row) else "") for i in range(len(headers))}
                        data_rows.append(row_dict)

                    sn_lower = sheetname.lower().strip()
                    entity = "customers"
                    if "expense" in sn_lower:
                        entity = "expenses"
                    elif "booking" in sn_lower or "order" in sn_lower or "event" in sn_lower:
                        entity = "bookings"
                    elif "menu" in sn_lower or "item" in sn_lower or "package" in sn_lower:
                        entity = "menu_items"

                    result[entity] = (headers, data_rows)

                if result:
                    return result, None
        except Exception:
            pass

    # Single sheet / CSV file auto-splitting
    headers, rows, err = parse_file(file_path)
    if err:
        return {}, err

    result = {}
    headers_lower = [h.lower() for h in headers]

    # Check for bookings data
    has_bkg = any("booking" in h or "event" in h or "pax" in h or "occasion" in h or "venue" in h for h in headers_lower)
    if has_bkg:
        bkg_rows = []
        for r in rows:
            has_name = any(r.get(h, "").strip() for h in headers if "name" in h.lower() or "customer" in h.lower())
            has_date = any(r.get(h, "").strip() for h in headers if "event" in h.lower() or "date" in h.lower())
            if has_name and has_date:
                bkg_rows.append(r)
        if bkg_rows:
            result["bookings"] = (headers, bkg_rows)

    # Check for customers data
    has_cus = any("customer" in h or "client" in h or "contact" in h for h in headers_lower)
    if has_cus or not result:
        cus_rows = []
        for r in rows:
            has_name = any(r.get(h, "").strip() for h in headers if "customer" in h.lower() or "client" in h.lower() or "name" in h.lower())
            if has_name:
                cus_rows.append(r)
        if cus_rows:
            result["customers"] = (headers, cus_rows)

    # Check for expenses data
    has_exp = any("expense" in h or "category" in h or "cost" in h for h in headers_lower)
    if has_exp:
        exp_rows = []
        for r in rows:
            has_exp_amt = any(normalize_amount(r.get(h, "")) > 0 for h in headers if "expense" in h.lower() or "amount" in h.lower() or "cost" in h.lower())
            has_exp_desc = any(r.get(h, "").strip() for h in headers if "expense" in h.lower() or "description" in h.lower() or "category" in h.lower())
            if has_exp_amt or has_exp_desc:
                exp_rows.append(r)
        if exp_rows:
            result["expenses"] = (headers, exp_rows)

    # Check for menu items data
    has_menu = any("menu" in h or "dish" in h or "package" in h or "item" in h for h in headers_lower)
    if has_menu:
        menu_rows = []
        for r in rows:
            has_item_name = any(r.get(h, "").strip() for h in headers if "item" in h.lower() or "dish" in h.lower())
            has_item_price = any(normalize_amount(r.get(h, "")) > 0 for h in headers if "price" in h.lower() or "rate" in h.lower())
            if has_item_name and has_item_price:
                menu_rows.append(r)
        if menu_rows:
            result["menu_items"] = (headers, menu_rows)

    if not result:
        result["customers"] = (headers, rows)

    return result, None


# ─────────────────────────────────────────────────────────────────────────────
# INTELLIGENT AUTO-MAPPING & NORMALIZATION
# ─────────────────────────────────────────────────────────────────────────────

def auto_map_headers(headers: List[str], entity_type: str) -> Dict[str, str]:
    """Auto-detect which uploaded column header maps to system fields for entity_type."""
    schema = ENTITY_SCHEMAS.get(entity_type, {})
    fields = schema.get("fields", {})
    mapping = {}

    entity_aliases = ENTITY_HEADER_ALIASES.get(entity_type, {})
    headers_clean = {h: _clean_header_str(h) for h in headers}

    for field_key in fields.keys():
        aliases = [_clean_header_str(a) for a in entity_aliases.get(field_key, [field_key])]
        matched_header = ""

        # 1. Exact match in order of alias priority
        for alias in aliases:
            for h_orig, h_clean in headers_clean.items():
                if h_clean == alias:
                    matched_header = h_orig
                    break
            if matched_header:
                break

        # 2. Substring match in order of alias priority
        if not matched_header:
            for alias in aliases:
                for h_orig, h_clean in headers_clean.items():
                    if alias in h_clean or h_clean in alias:
                        matched_header = h_orig
                        break
                if matched_header:
                    break

        mapping[field_key] = matched_header

    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATION & PREVIEW GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def validate_and_prepare_rows(
    data_rows: List[Dict[str, str]],
    mapping: Dict[str, str],
    entity_type: str
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    schema = ENTITY_SCHEMAS.get(entity_type, {})
    fields = schema.get("fields", {})

    prepared = []
    counts = {"valid": 0, "warning": 0, "error": 0}

    for idx, row in enumerate(data_rows, start=1):
        issues = []
        status = "valid"
        sanitized = {}

        for field_key, field_info in fields.items():
            mapped_h = mapping.get(field_key, "")
            raw_val = row.get(mapped_h, "").strip() if mapped_h else ""

            if field_info.get("required") and not raw_val:
                issues.append(f"Missing required field '{field_info['label']}'")
                status = "error"

            if field_key in ("amount", "total", "price", "total_amount", "expense_amount"):
                val = normalize_amount(raw_val)
                if val <= 0 and field_info.get("required"):
                    issues.append(f"Invalid amount '{raw_val}'")
                    status = "error"
                sanitized[field_key] = val
            elif field_key in ("date", "event_date", "expense_date"):
                if raw_val:
                    sanitized[field_key] = normalize_date(raw_val)
                else:
                    sanitized[field_key] = datetime.now().strftime("%b %d, %Y")
            elif field_key in ("pax",):
                sanitized[field_key] = normalize_pax(raw_val)
            elif field_key in ("category", "expense_category"):
                sanitized[field_key] = normalize_expense_category(raw_val) if entity_type in ("expenses", "all_in_one") else normalize_menu_category(raw_val)
            elif field_key in ("status",):
                sanitized[field_key] = normalize_booking_status(raw_val) if entity_type == "bookings" else normalize_customer_status(raw_val)
            else:
                sanitized[field_key] = raw_val

        # Warning checks
        if entity_type == "customers" and sanitized.get("contact"):
            if not re.search(r"\d{7,}", sanitized["contact"]):
                issues.append("Contact number format warning")
                if status == "valid":
                    status = "warning"

        if status == "valid":
            counts["valid"] += 1
        elif status == "warning":
            counts["warning"] += 1
        else:
            counts["error"] += 1

        prepared.append({
            "_row_index": idx,
            "_status": status,
            "_issues": issues,
            "_data": sanitized,
            "_raw": row,
        })

    return prepared, counts


# ─────────────────────────────────────────────────────────────────────────────
# BATCH DATABASE INSERTION
# ─────────────────────────────────────────────────────────────────────────────

def execute_batch_import(
    prepared_rows: List[Dict[str, Any]],
    entity_type: str,
    skip_errors: bool = True
) -> Tuple[int, int, List[str]]:
    success_count = 0
    fail_count = 0
    errors = []

    for row_info in prepared_rows:
        if row_info["_status"] == "error" and skip_errors:
            fail_count += 1
            err_details = ", ".join(row_info.get("_issues", ["Validation failed"]))
            errors.append(f"Row {row_info['_row_index']} skipped: {err_details}")
            continue

        data = row_info["_data"]
        try:
            if entity_type == "customers":
                cust_name = data.get("name", "").strip()
                if not cust_name:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Customer Name]: Value cannot be empty.")
                    continue

                res = repo.add_customer({
                    "name": cust_name,
                    "contact": data.get("contact", "").strip(),
                    "email": data.get("email", "").strip(),
                    "address": data.get("address", "").strip(),
                    "status": normalize_customer_status(data.get("status")),
                    "notes": data.get("notes", "").strip(),
                })
                if res:
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Customer: '{cust_name}']: Database insert failed.")

            elif entity_type == "expenses":
                amount_val = float(data.get("amount", 0.0))
                if amount_val <= 0:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Amount]: Expense amount must be greater than 0.")
                    continue

                res = repo.add_expense({
                    "category": normalize_expense_category(data.get("category")),
                    "description": data.get("description", "Imported Expense").strip() or "Imported Expense",
                    "amount": amount_val,
                    "date": data.get("date") or datetime.now().strftime("%Y-%m-%d"),
                })
                if res:
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Category: '{data.get('category')}']: Expense database insert failed.")

            elif entity_type == "bookings":
                cust_name = data.get("name", "").strip()
                total_val = float(data.get("total", 0.0))
                if not cust_name:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Customer Name]: Booking customer name cannot be empty.")
                    continue
                if total_val <= 0:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Total Amount]: Booking total amount must be greater than 0.")
                    continue

                bkg_payload = {
                    "name": cust_name,
                    "contact": data.get("contact", "").strip(),
                    "email": data.get("email", "").strip(),
                    "address": data.get("venue", "").strip() or data.get("address", "").strip() or "Cebu City",
                    "occasion": data.get("occasion", "").strip() or "Catering Event",
                    "venue": data.get("venue", "").strip() or "Catering Venue",
                    "date": data.get("date") or datetime.now().strftime("%Y-%m-%d"),
                    "time": data.get("time", "").strip() or "6:00 PM",
                    "pax": int(data.get("pax", 50)),
                    "notes": data.get("notes", "").strip(),
                    "menu_type": "package",
                    "total": total_val,
                    "payment_mode": "Cash",
                    "amount_paid": 0.0,
                }
                res = repo.create_booking(bkg_payload)
                if res and res.get("booking_id"):
                    target_status = normalize_booking_status(data.get("status"))
                    if target_status == "COMPLETED":
                        try:
                            repo.update_booking_status(res["booking_id"], "CONFIRMED")
                            repo.update_booking_status(res["booking_id"], "COMPLETED")
                        except Exception:
                            pass
                    elif target_status in ("CONFIRMED", "CANCELLED"):
                        try:
                            repo.update_booking_status(res["booking_id"], target_status)
                        except Exception:
                            pass
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Booking: '{cust_name}']: Booking database insert failed.")

            elif entity_type == "menu_items":
                item_name = data.get("name", "New Item").strip() or "New Item"
                price_val = float(data.get("price", 0.0))
                if price_val <= 0:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Price]: Menu item price must be greater than 0.")
                    continue

                item_payload = {
                    "item": item_name,
                    "category": normalize_menu_category(data.get("category")),
                    "package": "Standard",
                    "price": price_val,
                    "status": "Available",
                    "description": data.get("description", "").strip(),
                }
                res = repo.add_menu_item(item_payload)
                if res:
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']} [Item: '{item_name}']: Menu item database insert failed.")

            elif entity_type == "all_in_one":
                # Create customer & booking if present
                cust_name = (data.get("customer_name") or data.get("name", "")).strip()
                row_success = False
                if cust_name:
                    repo.add_customer({
                        "name": cust_name,
                        "contact": data.get("contact", "").strip(),
                        "email": data.get("email", "").strip(),
                        "address": data.get("venue", "").strip(),
                        "status": "Active",
                    })

                    total_amt = float(data.get("total_amount") or data.get("total", 0.0))
                    if total_amt > 0:
                        bkg_res = repo.create_booking({
                            "name": cust_name,
                            "contact": data.get("contact", "").strip(),
                            "email": data.get("email", "").strip(),
                            "address": data.get("venue", "").strip() or "Cebu City",
                            "occasion": data.get("occasion", "").strip() or "Event",
                            "venue": data.get("venue", "").strip() or "TBD",
                            "date": data.get("event_date") or data.get("date") or datetime.now().strftime("%Y-%m-%d"),
                            "time": data.get("event_time") or data.get("time") or "6:00 PM",
                            "pax": int(data.get("pax", 50)),
                            "total": total_amt,
                            "menu_type": "package",
                            "payment_mode": "Cash",
                            "amount_paid": 0.0,
                        })
                        if bkg_res and bkg_res.get("booking_id"):
                            row_success = True

                # Create expense if present
                exp_amt = float(data.get("expense_amount") or data.get("amount", 0.0))
                if exp_amt > 0:
                    exp_res = repo.add_expense({
                        "category": normalize_expense_category(data.get("expense_category") or data.get("category")),
                        "description": data.get("expense_description", "Master File Expense").strip() or "Master File Expense",
                        "amount": exp_amt,
                        "date": data.get("expense_date") or data.get("date") or datetime.now().strftime("%Y-%m-%d"),
                    })
                    if exp_res:
                        row_success = True

                if row_success:
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"Row {row_info['_row_index']}: No valid booking or expense data could be imported.")

        except Exception as e:
            fail_count += 1
            errors.append(f"Row {row_info['_row_index']}: {e}")

    # Emit app data change signals and run automatic deduplication
    if success_count > 0:
        try:
            if entity_type in ("customers", "bookings", "all_in_one"):
                repo.merge_duplicate_customers()
        except Exception:
            pass

        try:
            from utils.signals import app_events
            ev = app_events()
            ev.data_changed.emit()
            if entity_type == "expenses":
                ev.expense_saved.emit()
            elif entity_type == "customers":
                ev.customer_saved.emit()
            elif entity_type == "bookings":
                ev.booking_saved.emit()
            elif entity_type == "all_in_one":
                ev.customer_saved.emit()
                ev.booking_saved.emit()
                ev.expense_saved.emit()
        except Exception:
            pass

    return success_count, fail_count, errors


# ─────────────────────────────────────────────────────────────────────────────
# SAMPLE TEMPLATE GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_sample_csv(entity_type: str, save_path: str) -> Optional[str]:
    """Generate sample CSV or Excel file with clean headers and example data."""
    ext = os.path.splitext(save_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            if entity_type == "all_in_one":
                wb.remove(wb.active)  # remove default sheet
                sections = ["bookings", "customers", "expenses", "menu_items"]
                for sec in sections:
                    s_info = ENTITY_SCHEMAS.get(sec, {})
                    ws = wb.create_sheet(title=s_info.get("title", sec.title()))
                    for r in s_info.get("sample", []):
                        ws.append(r)
            else:
                ws = wb.active
                schema = ENTITY_SCHEMAS.get(entity_type, {})
                ws.title = schema.get("title", "Template")
                for r in schema.get("sample", []):
                    ws.append(r)
            wb.save(save_path)
            return None
        except Exception as e:
            return f"Failed to save Excel template: {e}"

    # CSV fallback
    schema = ENTITY_SCHEMAS.get(entity_type)
    if not schema:
        return "Unknown entity type."

    sample_rows = schema.get("sample", [])
    try:
        with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(sample_rows)
        return None
    except Exception as e:
        return f"Failed to save CSV template: {e}"
