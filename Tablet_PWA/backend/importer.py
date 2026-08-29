"""
Master-data import (PC -> kiosk) and template generation. Ported from the
original Tablet app's utils/importer.py — see that file for the fuller
design rationale comment this was distilled from. Two accepted formats:
a PC-exported .db (SQLite ATTACH) or a hand-edited .xlsx workbook.
"""
import os
from datetime import datetime

import db

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

_EMPTY_STATS = {"packages": 0, "menu_items": 0, "package_items": 0, "customers": 0, "addresses": 0, "errors": []}


def import_master_data(source_path: str) -> dict:
    if not source_path or not os.path.exists(source_path):
        return {**_EMPTY_STATS, "errors": ["Master data file not found."]}
    ext = os.path.splitext(source_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _import_master_data_excel(source_path)
    return _import_master_data_db(source_path)


def _replace_master_tables(packages: list, menu_items: list, package_items: list,
                            customers: list = None, provinces: list = None,
                            cities: list = None, barangays: list = None) -> dict:
    stats = {"packages": 0, "menu_items": 0, "package_items": 0, "customers": 0, "addresses": 0, "errors": []}
    try:
        db.execute("PRAGMA foreign_keys = OFF;")
        db.execute("DELETE FROM package_items")
        db.execute("DELETE FROM packages")
        db.execute("DELETE FROM menu_items")

        pkg_id_by_name = {}
        for p in packages:
            name = (p.get("name") or "").strip()
            if not name:
                continue
            new_id = db.execute(
                "INSERT INTO packages (pkg_name, pkg_description, pkg_price_per_pax, pkg_min_pax) VALUES (?, ?, ?, ?)",
                (name, p.get("description") or "", float(p.get("price_per_pax") or 0.0), int(p.get("min_pax") or 30)),
            )
            pkg_id_by_name[name.lower()] = new_id
            stats["packages"] += 1

        for m in menu_items:
            name = (m.get("name") or "").strip()
            if not name:
                continue
            db.execute(
                "INSERT INTO menu_items (mi_name, mi_category, mi_price, mi_status, mi_description) VALUES (?, ?, ?, ?, ?)",
                (name, m.get("category") or "Other", float(m.get("price") or 0.0), m.get("status") or "Available", m.get("description") or ""),
            )
            stats["menu_items"] += 1

        for pi in package_items:
            pkg_name = (pi.get("package_name") or "").strip().lower()
            pkg_id = pkg_id_by_name.get(pkg_name)
            item_name = (pi.get("item_name") or "").strip()
            if not pkg_id or not item_name:
                continue
            db.execute(
                "INSERT INTO package_items (pi_package_id, pi_item_name, pi_category, pi_custom_price, pi_quantity) VALUES (?, ?, ?, ?, ?)",
                (pkg_id, item_name, pi.get("category") or "Other", float(pi.get("price") or 0.0), int(pi.get("quantity") or 1)),
            )
            stats["package_items"] += 1

        if customers:
            for c in customers:
                c_name = (c.get("name") or c.get("cus_name") or "").strip()
                if not c_name:
                    continue
                c_id = c.get("id") or c.get("cus_id")
                c_contact = (c.get("contact") or c.get("cus_contact") or "").strip()
                c_email = (c.get("email") or c.get("cus_email") or "").strip()
                c_addr = (c.get("address") or c.get("cus_address") or "").strip()

                existing = None
                if c_id:
                    existing = db.fetchone("SELECT cus_id FROM customers WHERE cus_id = ?", (c_id,))
                if not existing and c_contact:
                    existing = db.fetchone("SELECT cus_id FROM customers WHERE cus_contact = ? AND cus_contact != ''", (c_contact,))
                if not existing:
                    existing = db.fetchone("SELECT cus_id FROM customers WHERE LOWER(cus_name) = LOWER(?)", (c_name,))

                if existing:
                    db.execute(
                        "UPDATE customers SET cus_name = ?, cus_contact = ?, cus_email = ?, cus_address = ? WHERE cus_id = ?",
                        (c_name, c_contact, c_email, c_addr, existing["cus_id"]),
                    )
                else:
                    db.execute(
                        "INSERT INTO customers (cus_id, cus_name, cus_contact, cus_email, cus_address, cus_status) VALUES (?, ?, ?, ?, ?, 'Active')",
                        (c_id, c_name, c_contact, c_email, c_addr),
                    )
                stats["customers"] += 1

            try:
                db.execute("""
                    DELETE FROM customers
                    WHERE cus_id NOT IN (
                        SELECT MIN(cus_id) FROM customers GROUP BY LOWER(cus_name), cus_contact
                    ) AND cus_id NOT IN (
                        SELECT DISTINCT bk_customer_id FROM bookings WHERE bk_customer_id IS NOT NULL
                    )
                """)
            except Exception:
                pass

        if provinces:
            for pr in provinces:
                db.execute("INSERT OR IGNORE INTO address_provinces (ap_id, ap_name) VALUES (?, ?)", (pr.get("ap_id"), pr.get("ap_name")))
        if cities:
            for ci in cities:
                db.execute("INSERT OR IGNORE INTO address_cities (ac_id, ac_province_id, ac_name) VALUES (?, ?, ?)", (ci.get("ac_id"), ci.get("ac_province_id"), ci.get("ac_name")))
        if barangays:
            for bg in barangays:
                db.execute("INSERT OR IGNORE INTO address_barangays (ab_id, ab_city_id, ab_name) VALUES (?, ?, ?)", (bg.get("ab_id"), bg.get("ab_city_id"), bg.get("ab_name")))
                stats["addresses"] += 1

        db.execute(
            "INSERT INTO tablet_master_sync (tms_source_export_version, tms_packages_count, tms_menu_items_count, tms_customers_count) VALUES (?, ?, ?, ?)",
            (datetime.now().isoformat(timespec="seconds"), stats["packages"], stats["menu_items"], stats["customers"]),
        )
    except Exception as exc:
        stats["errors"].append(str(exc))
    finally:
        try:
            db.execute("PRAGMA foreign_keys = ON;")
        except Exception:
            pass
    return stats


def _import_master_data_db(source_path: str) -> dict:
    try:
        db.execute("ATTACH DATABASE ? AS src", (source_path,))
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [f"Could not open master data file: {exc}"]}

    try:
        src_packages = db.fetchall("SELECT * FROM src.packages")
        src_items = db.fetchall("SELECT * FROM src.menu_items")
        try:
            src_pkg_items = db.fetchall("SELECT * FROM src.package_items")
        except Exception:
            src_pkg_items = []
        try:
            src_customers = db.fetchall("SELECT * FROM src.customers")
        except Exception:
            src_customers = []
        try:
            src_provinces = db.fetchall("SELECT * FROM src.address_provinces")
            src_cities = db.fetchall("SELECT * FROM src.address_cities")
            src_barangays = db.fetchall("SELECT * FROM src.address_barangays")
        except Exception:
            src_provinces, src_cities, src_barangays = [], [], []

        if not src_packages and not src_items and not src_customers:
            return {**_EMPTY_STATS, "errors": ["Master data file has no packages, menu items, or customers."]}

        pkg_name_by_id = {p["pkg_id"]: p["pkg_name"] for p in src_packages}
        packages = [
            {"name": p["pkg_name"], "description": p.get("pkg_description"), "price_per_pax": p["pkg_price_per_pax"], "min_pax": p.get("pkg_min_pax")}
            for p in src_packages
        ]
        menu_items = [
            {"name": m["mi_name"], "category": m["mi_category"], "price": m["mi_price"], "status": m.get("mi_status"), "description": m.get("mi_description")}
            for m in src_items
        ]
        package_items = [
            {
                "package_name": pkg_name_by_id.get(pi["pi_package_id"]),
                "category": pi.get("pi_category"), "item_name": pi.get("pi_item_name"),
                "price": pi.get("pi_custom_price"), "quantity": pi.get("pi_quantity"),
            }
            for pi in src_pkg_items
        ]
        return _replace_master_tables(
            packages, menu_items, package_items,
            customers=src_customers, provinces=src_provinces,
            cities=src_cities, barangays=src_barangays,
        )
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [str(exc)]}
    finally:
        try:
            db.execute("DETACH DATABASE src")
        except Exception:
            pass


def _import_master_data_excel(source_path: str) -> dict:
    if not OPENPYXL_OK:
        return {**_EMPTY_STATS, "errors": ["openpyxl is not installed on the server."]}
    try:
        wb = openpyxl.load_workbook(source_path, data_only=True)
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [f"Could not open Excel file: {exc}"]}

    if "Packages" not in wb.sheetnames:
        return {**_EMPTY_STATS, "errors": ["Excel file is missing a 'Packages' sheet. Use the sample template."]}

    def _rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip().lower() for h in rows[0]]
        out = []
        for r in rows[1:]:
            if not any(v not in (None, "") for v in r):
                continue
            out.append(dict(zip(headers, r)))
        return out

    pkg_rows = _rows("Packages")
    packages = [
        {"name": r.get("package name"), "description": r.get("description"),
         "price_per_pax": r.get("price per pax"), "min_pax": r.get("min pax")}
        for r in pkg_rows if r.get("package name")
    ]

    item_rows = _rows("Menu Items")
    menu_items, package_items = [], []
    seen = set()
    for r in item_rows:
        item_name = r.get("item name")
        if not item_name:
            continue
        category = r.get("category") or "Other"
        price = r.get("extra price") or 0
        pkg_name = r.get("package name")
        key = (str(item_name).strip().lower(), str(category).strip().lower())
        if key not in seen:
            menu_items.append({"name": item_name, "category": category, "price": price, "status": "Available", "description": ""})
            seen.add(key)
        if pkg_name:
            package_items.append({"package_name": pkg_name, "category": category, "item_name": item_name, "price": price, "quantity": r.get("quantity") or 1})

    if not packages:
        return {**_EMPTY_STATS, "errors": ["No packages found in the 'Packages' sheet."]}

    try:
        return _replace_master_tables(packages, menu_items, package_items)
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [str(exc)]}


def generate_sample_excel_template(save_path: str) -> bool:
    if not OPENPYXL_OK:
        return False
    try:
        wb = openpyxl.Workbook()
        ws_pkg = wb.active
        ws_pkg.title = "Packages"
        ws_pkg.append(["Package Name", "Description", "Price Per Pax", "Min Pax"])
        ws_pkg.append(["Classic Celebration Package", "Standard buffet with 4 main dishes, rice, dessert, drinks", 350.0, 30])
        ws_pkg.append(["Premium Grand Feast", "Deluxe buffet with 6 main dishes, lechon belly, 2 desserts", 550.0, 50])

        ws_items = wb.create_sheet("Menu Items")
        ws_items.append(["Package Name", "Category", "Item Name", "Extra Price", "Quantity"])
        ws_items.append(["Classic Celebration Package", "Main Dish", "Chicken BBQ", 0, 1])
        ws_items.append(["Classic Celebration Package", "Main Dish", "Pork Adobo", 0, 1])
        ws_items.append(["Classic Celebration Package", "Rice", "Steamed Rice", 0, 1])
        ws_items.append(["Premium Grand Feast", "Main Dish", "Lechon Belly", 0, 1])
        ws_items.append(["", "Add-on", "Additional Lechon (standalone item, no package)", 2500, 1])

        for ws in (ws_pkg, ws_items):
            for col in ws.columns:
                length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 12), 45)

        wb.save(save_path)
        return True
    except Exception as exc:
        print(f"[pwa importer] generate_sample_excel_template failed: {exc}")
        return False


def get_last_master_sync() -> dict | None:
    return db.fetchone("SELECT * FROM tablet_master_sync ORDER BY tms_id DESC LIMIT 1")


def export_local_database(dest_path: str) -> bool:
    import shutil
    from schema import DB_PATH
    try:
        shutil.copy2(str(DB_PATH), dest_path)
        return True
    except Exception as exc:
        print(f"[pwa importer] export_local_database failed: {exc}")
        return False
