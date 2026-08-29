"""
Receipt PDF + orders Excel archive export. Ported from the original Tablet
app's utils/exporter.py (reportlab receipt, openpyxl multi-sheet archive).
"""
from datetime import datetime

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

_C_RED = colors.HexColor("#E11D48") if REPORTLAB_OK else None
_C_DARK = colors.HexColor("#0B1220") if REPORTLAB_OK else None
_C_GRAY = colors.HexColor("#6B7280") if REPORTLAB_OK else None
_C_GREEN = colors.HexColor("#22C55E") if REPORTLAB_OK else None
_C_BORDER = colors.HexColor("#E5E7EB") if REPORTLAB_OK else None
_MARGIN = 1.5 * cm if REPORTLAB_OK else 0
_PAGE_W = A4[0] if REPORTLAB_OK else 0
_CONTENT_W = _PAGE_W - 2 * _MARGIN if REPORTLAB_OK else 0


def export_order_receipt_pdf(path: str, order: dict, business_name: str = "Jayraldine's Catering") -> bool:
    if not REPORTLAB_OK:
        return False
    try:
        doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=_MARGIN, rightMargin=_MARGIN,
                                 topMargin=_MARGIN, bottomMargin=_MARGIN, title=f"Receipt — {order.get('booking_ref', '')}")
        story = []
        title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=18, textColor=_C_RED, leading=22)
        sub_style = ParagraphStyle("sub", fontName="Helvetica", fontSize=9, textColor=_C_GRAY, leading=12)
        label_style = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=9.5, textColor=_C_GRAY, leading=13)
        value_style = ParagraphStyle("value", fontName="Helvetica", fontSize=9.5, textColor=_C_DARK, leading=13)
        right_style = ParagraphStyle("right", fontName="Helvetica", fontSize=9.5, textColor=_C_DARK, alignment=TA_RIGHT, leading=13)

        story.append(Paragraph(business_name, title_style))
        story.append(Paragraph("Order Receipt (Kiosk PWA)", sub_style))
        story.append(Spacer(1, 6))
        story.append(HRFlowable(width="100%", thickness=2, color=_C_RED))
        story.append(Spacer(1, 10))

        info_rows = [
            [Paragraph("Order Ref", label_style), Paragraph(str(order.get("booking_ref", "—")), value_style)],
            [Paragraph("Customer", label_style), Paragraph(str(order.get("customer_name", "—")), value_style)],
            [Paragraph("Event Date", label_style), Paragraph(str(order.get("event_date", "—")), value_style)],
            [Paragraph("Venue / Occasion", label_style), Paragraph(f"{order.get('venue', '—')} / {order.get('occasion', '—')}", value_style)],
            [Paragraph("Guests", label_style), Paragraph(str(order.get("pax", "—")), value_style)],
        ]
        info_tbl = Table(info_rows, colWidths=[4 * cm, _CONTENT_W - 4 * cm])
        info_tbl.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, _C_BORDER), ("INNERGRID", (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(info_tbl)
        story.append(Spacer(1, 10))

        menu_selections = order.get("menu_selections") or []
        if menu_selections:
            story.append(Paragraph("Selected Menu", label_style))
            story.append(Spacer(1, 4))
            rows = [[Paragraph("Category", label_style), Paragraph("Item", label_style)]]
            for m in menu_selections:
                rows.append([Paragraph(m.get("category", ""), value_style), Paragraph(m.get("item_name", ""), value_style)])
            tbl = Table(rows, colWidths=[4 * cm, _CONTENT_W - 4 * cm])
            tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.4, _C_BORDER), ("INNERGRID", (0, 0), (-1, -1), 0.3, _C_BORDER),
                ("BACKGROUND", (0, 0), (-1, 0), _C_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 10))

        charges = order.get("additional_charges") or []
        if charges:
            story.append(Paragraph("Additional Charges / Discounts", label_style))
            story.append(Spacer(1, 4))
            rows = [[Paragraph("Description", label_style), Paragraph("Amount", label_style)]]
            for c in charges:
                amt = float(c["amount"])
                amt_text = f"- PHP {abs(amt):,.2f}" if amt < 0 else f"PHP {amt:,.2f}"
                rows.append([Paragraph(c["description"], value_style), Paragraph(amt_text, right_style)])
            tbl = Table(rows, colWidths=[_CONTENT_W * 0.7, _CONTENT_W * 0.3])
            tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.4, _C_BORDER), ("INNERGRID", (0, 0), (-1, -1), 0.3, _C_BORDER),
                ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 10))

        total = float(order.get("total", 0))
        paid = float(order.get("paid", 0))
        balance = float(order.get("balance", max(0.0, total - paid)))
        status = order.get("status", "Unpaid")
        status_color = {"Paid": _C_GREEN, "Partial": colors.HexColor("#F59E0B")}.get(status, colors.HexColor("#EF4444"))

        amount_rows = [
            [Paragraph("Total Amount", label_style), Paragraph(f"PHP {total:,.2f}", right_style)],
            [Paragraph("Amount Paid", label_style), Paragraph(f"PHP {paid:,.2f}", ParagraphStyle("p", parent=right_style, textColor=_C_GREEN))],
            [Paragraph("Remaining Balance", ParagraphStyle("bl", fontName="Helvetica-Bold", fontSize=10, textColor=_C_DARK)),
             Paragraph(f"PHP {balance:,.2f}", ParagraphStyle("bv", fontName="Helvetica-Bold", fontSize=12, textColor=colors.HexColor("#EF4444") if balance > 0 else _C_GREEN, alignment=TA_RIGHT))],
            [Paragraph("Payment Status", label_style), Paragraph(status.upper(), ParagraphStyle("st", fontName="Helvetica-Bold", fontSize=10, textColor=status_color, alignment=TA_RIGHT))],
        ]
        amt_tbl = Table(amount_rows, colWidths=[_CONTENT_W * 0.55, _CONTENT_W * 0.45])
        amt_tbl.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, _C_BORDER), ("INNERGRID", (0, 0), (-1, -1), 0.3, _C_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(amt_tbl)
        story.append(Spacer(1, 14))
        story.append(HRFlowable(width="100%", thickness=0.5, color=_C_BORDER))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"Printed: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} · Thank you for choosing {business_name}!",
            ParagraphStyle("footer", fontName="Helvetica", fontSize=7.5, textColor=_C_GRAY, alignment=TA_CENTER),
        ))
        doc.build(story)
        return True
    except Exception as exc:
        print(f"[pwa exporter] Receipt PDF failed: {exc}")
        return False


def export_all_orders_to_excel(save_path: str) -> dict:
    import db
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"success": False, "orders_count": 0, "path": "", "error": "openpyxl is required for Excel export."}

    try:
        bookings = db.fetchall("""
            SELECT b.*, i.inv_id, i.inv_invoice_number, i.inv_total_amount, i.inv_amount_paid, i.inv_balance, i.inv_status
            FROM bookings b
            LEFT JOIN invoices i ON i.inv_booking_id = b.bk_id
            ORDER BY b.bk_created_at DESC
        """)

        wb = openpyxl.Workbook()
        ws_orders = wb.active
        ws_orders.title = "Orders Summary"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")

        ws_orders.append(["JAYRALDINE'S CATERING — KIOSK ORDERS ARCHIVE"])
        ws_orders.append([f"Export Date & Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_orders.append([])
        ws_orders.cell(1, 1).font = title_font
        ws_orders.cell(2, 1).font = sub_font

        order_headers = [
            "Booking Ref", "Customer Name", "Contact Number", "Email Address", "Delivery / Billing Address",
            "Event Date", "Event Time", "Venue", "Occasion", "Pax", "Package",
            "Base Total (PHP)", "Total Amount (PHP)", "Amount Paid (PHP)", "Balance Due (PHP)",
            "Status", "Payment Mode", "Date Created", "Notes",
        ]
        ws_orders.append(order_headers)
        for col_num in range(1, len(order_headers) + 1):
            c = ws_orders.cell(row=4, column=col_num)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for b in bookings:
            ws_orders.append([
                b.get("bk_booking_ref") or f"TB-{b['bk_id']:05d}",
                b.get("bk_customer_name", ""),
                b.get("bk_contact", ""),
                b.get("bk_email", ""),
                b.get("bk_address", ""),
                b.get("bk_event_date", ""),
                b.get("bk_event_time", ""),
                b.get("bk_venue", ""),
                b.get("bk_occasion", ""),
                b.get("bk_pax", 0),
                b.get("bk_package_name") or "Custom Package",
                float(b.get("bk_base_total") or 0.0),
                float(b.get("bk_total_amount") or 0.0),
                float(b.get("inv_amount_paid") or b.get("bk_amount_paid") or 0.0),
                float(b.get("inv_balance") or 0.0),
                b.get("inv_status") or b.get("bk_status") or "PENDING",
                b.get("bk_payment_mode", "Cash"),
                b.get("bk_created_at", ""),
                b.get("bk_notes", ""),
            ])

        ws_menu = wb.create_sheet(title="Menu Selections")
        ws_menu.append(["Booking Ref", "Customer Name", "Dish Name", "Category", "Quantity", "Extra Price (PHP)"])
        for col_num in range(1, 7):
            c = ws_menu.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        menu_items = db.fetchall("""
            SELECT bmi.*, b.bk_booking_ref, b.bk_customer_name
            FROM booking_menu_items bmi
            JOIN bookings b ON b.bk_id = bmi.bmi_booking_id
            ORDER BY b.bk_id, bmi.bmi_category
        """)
        for m in menu_items:
            ws_menu.append([
                m.get("bk_booking_ref", ""), m.get("bk_customer_name", ""),
                m.get("bmi_item_name", ""), m.get("bmi_category", ""),
                m.get("bmi_quantity", 1), float(m.get("bmi_price") or 0.0),
            ])

        ws_charges = wb.create_sheet(title="Additional Charges")
        ws_charges.append(["Booking Ref", "Customer Name", "Charge Description", "Amount (PHP)", "Date Added", "Added By"])
        for col_num in range(1, 7):
            c = ws_charges.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        charges = db.fetchall("""
            SELECT ac.*, b.bk_booking_ref, b.bk_customer_name
            FROM booking_additional_charges ac
            JOIN bookings b ON b.bk_id = ac.ac_booking_id
            ORDER BY b.bk_id, ac.ac_date_added
        """)
        for ch in charges:
            ws_charges.append([
                ch.get("bk_booking_ref", ""), ch.get("bk_customer_name", ""),
                ch.get("ac_description", ""), float(ch.get("ac_amount") or 0.0),
                ch.get("ac_date_added", ""), ch.get("ac_added_by", "Staff"),
            ])

        ws_payments = wb.create_sheet(title="Payment Records")
        ws_payments.append(["Booking Ref", "Customer Name", "Invoice Ref", "Payment Amount (PHP)", "Payment Date", "Payment Method", "Is Downpayment", "Notes"])
        for col_num in range(1, 9):
            c = ws_payments.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        payments = db.fetchall("""
            SELECT pr.*, i.inv_invoice_number, b.bk_booking_ref, b.bk_customer_name
            FROM payment_records pr
            JOIN invoices i ON i.inv_id = pr.pr_invoice_id
            JOIN bookings b ON b.bk_id = i.inv_booking_id
            ORDER BY pr.pr_payment_date DESC
        """)
        for p in payments:
            ws_payments.append([
                p.get("bk_booking_ref", ""), p.get("bk_customer_name", ""),
                p.get("inv_invoice_number", ""), float(p.get("pr_amount") or 0.0),
                p.get("pr_payment_date", ""), p.get("pr_payment_method") or p.get("pr_method") or "Cash",
                "Yes" if p.get("pr_is_downpayment") else "No",
                p.get("pr_notes") or p.get("pr_note") or "",
            ])

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(save_path)
        return {"success": True, "orders_count": len(bookings), "path": save_path, "error": None}
    except Exception as exc:
        print(f"[pwa exporter] Excel export failed: {exc}")
        return {"success": False, "orders_count": 0, "path": "", "error": str(exc)}
