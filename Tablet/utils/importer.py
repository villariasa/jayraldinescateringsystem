"""
Two distinct, separate data flows (Tablet-mode.md sections 14/16 — never mix
these up):

- import_master_data(): PC -> Tablet. Packages/menu/prices only. Sync, not
  destructive to transaction history: historical orders store their own
  snapshot values in booking_menu_items / bk_base_total / booking_additional_
  charges at the time they were placed, they never reference packages/
  menu_items live - so replacing the master tables wholesale on each import
  is safe and correctly satisfies "historical orders keep their original
  price" without any extra bookkeeping. Accepts either a PC-exported .db
  file OR a hand-edited .xlsx workbook (see generate_sample_excel_template).

- export_order() / the tablet's own local .db file: Tablet -> PC. Because
  the tablet's schema is an exact mirror of the PC's for the shared tables,
  the tablet's local database file *is* the export - it can be opened
  directly by the PC's Settings -> "Merge Backup File Into This Database"
  (utils/importer.py: merge_database_file in the PC app), no separate export
  format needed.
"""
import os
import shutil
from datetime import datetime

import utils.db as db

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

_EMPTY_STATS = {"packages": 0, "menu_items": 0, "package_items": 0, "errors": []}


def import_master_data(source_path: str) -> dict:
    """Dispatches to the .db or .xlsx importer based on file extension."""
    if not source_path or not os.path.exists(source_path):
        return {**_EMPTY_STATS, "errors": ["Master data file not found."]}

    ext = os.path.splitext(source_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _import_master_data_excel(source_path)
    return _import_master_data_db(source_path)


def _replace_master_tables(packages: list, menu_items: list, package_items: list) -> dict:
    """packages: [{"name","description","price_per_pax","min_pax"}]
    menu_items: [{"name","category","price","status","description"}]
    package_items: [{"package_name","category","item_name","price","quantity"}]
    Wipes and reloads packages/menu_items/package_items — safe, see module docstring."""
    stats = {"packages": 0, "menu_items": 0, "package_items": 0, "errors": []}

    db.execute("DELETE FROM package_items")
    db.execute("DELETE FROM packages")
    db.execute("DELETE FROM menu_items")

    pkg_id_by_name = {}
    for p in packages:
        name = (p.get("name") or "").strip()
        if not name:
            continue
        new_id = db.execute(
            "INSERT INTO packages (pkg_name, pkg_description, pkg_price_per_pax, pkg_min_pax) VALUES (?, ?, ?, ?)",
            (name, p.get("description") or "", float(p.get("price_per_pax") or 0.0), int(p.get("min_pax") or 30)),
        )
        pkg_id_by_name[name.lower()] = new_id
        stats["packages"] += 1

    for m in menu_items:
        name = (m.get("name") or "").strip()
        if not name:
            continue
        db.execute(
            "INSERT INTO menu_items (mi_name, mi_category, mi_price, mi_status, mi_description) VALUES (?, ?, ?, ?, ?)",
            (name, m.get("category") or "Other", float(m.get("price") or 0.0), m.get("status") or "Available", m.get("description") or ""),
        )
        stats["menu_items"] += 1

    for pi in package_items:
        pkg_name = (pi.get("package_name") or "").strip().lower()
        pkg_id = pkg_id_by_name.get(pkg_name)
        item_name = (pi.get("item_name") or "").strip()
        if not pkg_id or not item_name:
            continue
        db.execute(
            "INSERT INTO package_items (pi_package_id, pi_item_name, pi_category, pi_custom_price, pi_quantity) VALUES (?, ?, ?, ?, ?)",
            (pkg_id, item_name, pi.get("category") or "Other", float(pi.get("price") or 0.0), int(pi.get("quantity") or 1)),
        )
        stats["package_items"] += 1

    db.execute(
        "INSERT INTO tablet_master_sync (tms_source_export_version, tms_packages_count, tms_menu_items_count) VALUES (?, ?, ?)",
        (datetime.now().isoformat(timespec="seconds"), stats["packages"], stats["menu_items"]),
    )
    return stats


def _import_master_data_db(source_path: str) -> dict:
    """Import from a PC-exported SQLite .db file (Settings -> Export Tablet Master Data on the PC)."""
    try:
        db.execute("ATTACH DATABASE ? AS src", (source_path,))
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [f"Could not open master data file: {exc}"]}

    try:
        src_packages = db.fetchall("SELECT * FROM src.packages")
        src_items = db.fetchall("SELECT * FROM src.menu_items")
        try:
            src_pkg_items = db.fetchall("SELECT * FROM src.package_items")
        except Exception:
            src_pkg_items = []

        if not src_packages and not src_items:
            return {**_EMPTY_STATS, "errors": ["Master data file has no packages or menu items."]}

        pkg_name_by_id = {p["pkg_id"]: p["pkg_name"] for p in src_packages}
        packages = [
            {"name": p["pkg_name"], "description": p.get("pkg_description"), "price_per_pax": p["pkg_price_per_pax"], "min_pax": p.get("pkg_min_pax")}
            for p in src_packages
        ]
        menu_items = [
            {"name": m["mi_name"], "category": m["mi_category"], "price": m["mi_price"], "status": m.get("mi_status"), "description": m.get("mi_description")}
            for m in src_items
        ]
        package_items = [
            {
                "package_name": pkg_name_by_id.get(pi["pi_package_id"]),
                "category": pi.get("pi_category"), "item_name": pi.get("pi_item_name"),
                "price": pi.get("pi_custom_price"), "quantity": pi.get("pi_quantity"),
            }
            for pi in src_pkg_items
        ]
        return _replace_master_tables(packages, menu_items, package_items)
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [str(exc)]}
    finally:
        try:
            db.execute("DETACH DATABASE src")
        except Exception:
            pass


def _import_master_data_excel(source_path: str) -> dict:
    """Import from a hand-editable .xlsx workbook — see
    generate_sample_excel_template() for the exact expected column layout.
    Sheet 'Packages': Package Name, Description, Price Per Pax, Min Pax.
    Sheet 'Menu Items': Package Name, Category, Item Name, Extra Price, Quantity.
    A blank Package Name on a Menu Items row means a standalone menu item
    (not tied to any specific package's choice list)."""
    if not OPENPYXL_OK:
        return {**_EMPTY_STATS, "errors": ["openpyxl is not installed — run: pip install openpyxl"]}

    try:
        wb = openpyxl.load_workbook(source_path, data_only=True)
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [f"Could not open Excel file: {exc}"]}

    if "Packages" not in wb.sheetnames:
        return {**_EMPTY_STATS, "errors": ["Excel file is missing a 'Packages' sheet. Use the sample template."]}

    def _rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip().lower() for h in rows[0]]
        out = []
        for r in rows[1:]:
            if not any(v not in (None, "") for v in r):
                continue
            out.append(dict(zip(headers, r)))
        return out

    pkg_rows = _rows("Packages")
    packages = [
        {
            "name": r.get("package name"), "description": r.get("description"),
            "price_per_pax": r.get("price per pax"), "min_pax": r.get("min pax"),
        }
        for r in pkg_rows if r.get("package name")
    ]

    item_rows = _rows("Menu Items")
    menu_items, package_items = [], []
    seen_menu_items = set()
    for r in item_rows:
        item_name = r.get("item name")
        if not item_name:
            continue
        category = r.get("category") or "Other"
        price = r.get("extra price") or 0
        pkg_name = r.get("package name")

        key = (str(item_name).strip().lower(), str(category).strip().lower())
        if key not in seen_menu_items:
            menu_items.append({"name": item_name, "category": category, "price": price, "status": "Available", "description": ""})
            seen_menu_items.add(key)

        if pkg_name:
            package_items.append({
                "package_name": pkg_name, "category": category, "item_name": item_name,
                "price": price, "quantity": r.get("quantity") or 1,
            })

    if not packages:
        return {**_EMPTY_STATS, "errors": ["No packages found in the 'Packages' sheet."]}

    try:
        return _replace_master_tables(packages, menu_items, package_items)
    except Exception as exc:
        return {**_EMPTY_STATS, "errors": [str(exc)]}


def generate_sample_excel_template(save_path: str) -> bool:
    """Writes a starter .xlsx with the exact columns import_master_data()
    expects, pre-filled with example rows, so the owner can just edit it in
    Excel and re-import without guessing the format."""
    if not OPENPYXL_OK:
        return False
    try:
        wb = openpyxl.Workbook()
        ws_pkg = wb.active
        ws_pkg.title = "Packages"
        ws_pkg.append(["Package Name", "Description", "Price Per Pax", "Min Pax"])
        ws_pkg.append(["Classic Celebration Package", "Standard buffet with 4 main dishes, rice, dessert, drinks", 350.0, 30])
        ws_pkg.append(["Premium Grand Feast", "Deluxe buffet with 6 main dishes, lechon belly, 2 desserts", 550.0, 50])

        ws_items = wb.create_sheet("Menu Items")
        ws_items.append(["Package Name", "Category", "Item Name", "Extra Price", "Quantity"])
        ws_items.append(["Classic Celebration Package", "Main Dish", "Chicken BBQ", 0, 1])
        ws_items.append(["Classic Celebration Package", "Main Dish", "Pork Adobo", 0, 1])
        ws_items.append(["Classic Celebration Package", "Rice", "Steamed Rice", 0, 1])
        ws_items.append(["Premium Grand Feast", "Main Dish", "Lechon Belly", 0, 1])
        ws_items.append(["", "Add-on", "Additional Lechon (standalone item, no package)", 2500, 1])

        for ws in (ws_pkg, ws_items):
            for col in ws.columns:
                length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 12), 45)

        wb.save(save_path)
        return True
    except Exception as exc:
        print(f"[tablet importer] generate_sample_excel_template failed: {exc}")
        return False


def get_last_master_sync() -> dict | None:
    return db.fetchone("SELECT * FROM tablet_master_sync ORDER BY tms_id DESC LIMIT 1")


def export_local_database(dest_path: str) -> bool:
    """The tablet's local .db IS the Tablet -> PC export (see module docstring).
    We just checkpoint WAL and copy the file so the export is self-contained."""
    try:
        db.execute("PRAGMA wal_checkpoint(FULL)")
    except Exception:
        pass
    from utils.sqlite_schema import get_sqlite_db_path
    src_path = str(db._db_path or get_sqlite_db_path())
    try:
        shutil.copy2(src_path, dest_path)
        return True
    except Exception as exc:
        print(f"[tablet importer] export_local_database failed: {exc}")
        return False
